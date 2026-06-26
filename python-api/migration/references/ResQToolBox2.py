import win32com.client
import numpy as np
import pythoncom
import itertools
import time

# custom modules
from XLToolBox2 import *


# Connect to ResQ Automation Library
ResQApp = win32com.client.Dispatch("ResQ3Automation.ResQApplication")

# JGO_CO1SQLWPV22 = "C:\ProgramData\Willis Towers Watson\ResQ\Data\Local Database.mdb"
# Local_Database = "E:\MSSQL13.MSSQLSERVER\MSSQL\DATA\ResQNew"
# Server_Name = 'CO1SQLWPV22'
Connection_Name = 'JGO_CO1SQLWPV22'


# connection & administation

projects = {}

def connect(connection_name="JGO_CO1SQLWPV22", user_name="", password="", print_info=True):
    ResQApp.ConnectByName(connection_name, user_name, password)

    resq_user_name = ResQApp.LastUserName
    if resq_user_name == '':
        resq_user_name = u.getpass.getuser()
    if print_info:
        print(f"Connected as [{connection_name}] - {resq_user_name} \nResQ Version: {ResQApp.AppVersion}")


def connect_admin(print_info=True):
    """
    admin-robot
    """
    connect(connection_name='JGO_CO1SQLWPV22', user_name='admin-robot', password='098765', print_info=print_info)


def reconnect(print_info=False):
    con_name = ResQApp.LastConnection
    proj_name = proj_info['Project Name']
    resClass = proj_info['Reserving Class']

    ResQApp.Disconnect()
    
    if ResQApp.LastUserName == 'admin-robot':
        connect_admin(print_info=print_info)
    else:
        ResQApp.ConnectByName(con_name, "", "")
        
    try:
        set_proj_info(con_name, proj_name, resClass)
    except:
        pass


proj_info = {
    'Connection': "JGO_CO1SQLWPV22",
    'Project Name': None, 
    'Reserving Class': None,
    'xProject': None, 
    'xReservingClass': None
}

global_dict = {
    'DFM Ratio': {'Dev Period': None, 'Old': None, 'New': None}
}


def _project2com(name):

    if name in ['Default', 'default', '']:
        return proj_info['xProject']

    elif '.Project' in str(type(name)):
        return name.COMObject
    
    elif type(name) == str:
        return ResQApp.Projects().Item(name)
    
    else:
        return name
    
def _reserving_class2com(project, path):

    xProject = _project2com(project)

    if path in ['Default', 'default', '']:
        return proj_info['xReservingClass']

    elif '.Reserving_Class' in str(type(path)):
        return path.COMObject
    
    elif type(path) == str and path != 'Default':
        return xProject.GetReservingClass(path)
    
    else:
        return path



class _ResQ:
    
    def __init__(self):
        self.COM = self.COMObject = ResQApp
        self.dataset_format = {'Triangle': 0, 'Vector': 1}
        self.method_type = {'None': 0, 'DFM': 1, 'BF': 2, 'CC': 3, 'Result Selection': 4}
        self.percentage_developed_type = {'Latest/Ultimates': 0, 'Pattern vector': 1, 'DFM dev factors': 2}

    def get_connection_name(self):
        return ResQApp.LastConnection

    def get_user_name(self):
        if ResQApp.LastUserName == '':
            return user_name
        else:
            return ResQApp.LastUserName

    def is_connected(self):
        return ResQApp.Connected

    def Project(self, name='Default'):
        return Project(name)
    
    def add_folder(self, name, parent_folder_name):
        _add_project_folder(name, parent_folder_name)
    
    def reconnect(self):
        reconnect()


ResQ = _ResQ()


class Project:
    """
    Class: Project() can be used to generate a ResQ Project instance. It will contain all the properties and methods of any ResQ Project Object.

    If the input project name is 'Default', it will generate the the project instance based on the default value stored in the proj_info dictionary.

    The input value can be a string or a ResQ Project COMObject, or a Project class instance from ResQToolBox.

    Example Usage:
    >>> My_Project = Project()

    """

    def __init__(self, name='Default'):
   
        if name == 'Default':
            try:
                if projects == {}:
                    print('Loading all projects...')
                    for i in tq(ResQApp.Projects()):
                        projects[i.Name] = i
                search(projects)
                return
            except Exception as e:
                print(e)
                return

        self.COMObject = _project2com(name)
        self.COM = self.COMObject
        self.name = self.COMObject.Name
        self.location = self.COMObject.Folder.Path
        self.dataset_types = "Use 'get_dataset_types' method to initialize this property"
        self.import_wizard = self.get_imports()

    def add_reserving_class(self, parent_class, reserving_class_type):
        _add_reserving_class_path(self.COM, parent_class, reserving_class_type)

    def Reserving_Class(self, path = ''):
        """
        Class: Reserving_Class() can be used to generate a ResQ Reserving Class instance.
        """
        if path == "":
            RC_list = []
            for i in tq(self.COMObject.ReservingClasses(), "Loading Reserving Classes..."):
                if i.Path.count('\\') == 4:
                    RC_list.append(i.Path)
            search(RC_list)
        else:
            return Reserving_Class(path=path, project=self.name)
        

    def get_imports(self):   
        break_count = 0; imp_list = {}

        for i in self.COMObject.Imports():
            if break_count >= len(self.COMObject.Imports()):
                break
            try:
                imp_list[i.Name] = i
                break_count += 1
            except:
                pass

        return imp_list


    def get_reserving_classes(self, calculated='All'):
        """
        Return a dict which contains all reserving classes in the project
        calculated: 'All', True, False
        """
        RC_dict = {}
        if calculated == 'All':
            for i in tq(self.COMObject.ReservingClasses(), "Collecting reserving classes..."):
                if i.Path.count('\\') == 4:
                    RC_dict[i.Path] = i

        elif calculated == True:
            for i in tq(self.COMObject.ReservingClasses(), "Collecting calculated reserving classes..."):
                if i.Calculated == True and i.Path.count('\\') == 4:
                    RC_dict[i.Path] = i   

        elif calculated == False:
            for i in tq(self.COMObject.ReservingClasses(), "Collecting non-calculated reserving classes..."):
                if i.Calculated == False and i.Path.count('\\') == 4:
                    RC_dict[i.Path] = i          

        return RC_dict


    def duplicate(self, new_name, output_folder='current', write_log=False):
        _duplicate_project(self.COMObject.Name, new_name, output_folder, write_log)


    def add_dataset_type(self, dataset_name, dataset_cat, aggregate = True, unique = True):
        _add_dataset_type(self.COMObject, dataset_name, dataset_cat, aggregate, unique)


    def get_dataset_types(self):
        self.dataset_types = {i.Name: Dataset_Type(self.name, i.Name) for i in self.COMObject.DatasetTypes()}


    def Dataset_Type(self, name = ''):

        if name == '': 
            search([i.Name for i in self.COMObject.DatasetTypes()])
            return

        return Dataset_Type(self.name, name)
    

    def Dataset_Type_Category(self, name = ''):

        if name == '': 
            search([i.Name for i in self.COMObject.Categories()])
            return

        return Dataset_Type_Category(self.name, name)


class Reserving_Class:
    """
    Class: Reserving_Class() can be used to generate a ResQ Reserving Class instance.
    """
    def __init__(self, path = 'Default', project = 'Default'):
        if path == 'Default':
            set_project(_project2com(project).Name)
            # search(reserving_classes)
            select_reserving_class()
            return
        elif path == 'Default' and proj_info['Reserving Class'] != None:
            path = proj_info['Reserving Class']

        self.project = Project(_project2com(project).Name)
        self.COM = self.COMObject = _reserving_class2com(self.project, path)
        if self.COMObject is None:
            self = None
            return
        self.path = self.COMObject.Path
        self.name = self.COMObject.Name
        self.calculated = self.COMObject.Calculated

    def add_method(self, name, method_type, project='Default', reserving_class='Default'):
        _add_method(name, method_type, project, reserving_class)

    def get_child_class(self, name):
        return Reserving_Class(self.COMObject.GetChildClass(name).Path)
    

    def Triangle(self, name = ''):
        if name == '':
            search([i.Name for i in self.COMObject.Triangles()])
            return
        return Triangle(name, self.project, self.path)


    def Vector(self, name = ''):
        if name == '':
            search([i.Name for i in self.COMObject.Vectors()])
            return
        return Vector(name, self.project, self.path)        


    def DFM(self, name = ''):
        if name == '':
            search([i.Name for i in self.COMObject.DFMMethods()])
            return
        return DFM(name, self.project, self.path)
    
    def add_dataset(self, dataset_type, name: str = 'default', formula: str = 'default'):
        _add_dataset(self.project, self.path, dataset_type, name, formula)


class Vector():

    def __init__(self, name='', project='Default', reserving_class='Default'):

        xProject = _project2com(project)
        xReservingClass = _reserving_class2com(xProject, reserving_class)

        if name == '':
            search([i.Name for i in xReservingClass.Vectors()])  
            return

        self.reserving_class = Reserving_Class(xReservingClass.Path, xProject)
        self.project = Project(xProject.Name)
        self.COMObject = xReservingClass.GetVector(name)
        self.COM = self.COMObject
        self.name = self.COMObject.Name

        self.count = None
        self.period_length = None
        self.stored_period_length = None
        self.notes = None
        self.formula = None
        self.projected_ultimates = None

        self.update()

    def update(self):
        self.count = self.COMObject.Count
        self.period_length = self.COMObject.PeriodLength
        self.stored_period_length = self.COMObject.StoredPeriodLength
        self.notes = self.COMObject.Notes
        self.formula = self.COMObject.Formula

    def set_formula(self, formula):
        self.COMObject.Formula = formula
        self.update()

    def set_period_length(self, value):
        self.COMObject.PeriodLength = value
        self.update()
        # self.save()

    def value(self, index):
        self.COMObject.PeriodLength = self.period_length
        if index < 1000:
            return self.COMObject.ValuesByIndex(index)
        else:
            offset = int(self.COMObject.PeriodLabel(1)) - 1
            return self.COMObject.ValuesByIndex(index - offset)
            

    def set_value(self, index: int, value):
        self.COMObject.PeriodLength = self.period_length
        if index < 1000:
            self.COMObject.SetValuesByIndex(index, value)
        else:
            offset = int(self.COMObject.PeriodLabel(1)) - 1
            self.COMObject.SetValuesByIndex(index - offset, value)
        # self.values: list = [self.COMObject.ValuesByIndex(i) for i in range_(self.count)]

    def clear_value(self):
        self.set_period_length(self.stored_period_length)
        for i in range_(self.count):
            self.set_value(i, 0)
        self.update()
        # self.save()
        print('Values Cleared.')

    def clear_notes(self):
        self.COMObject.Notes = ''
        self.notes = ''

    def clear_values(self):
        display_period_length = self.period_length
        self.set_period_length(self.stored_period_length)
        for i in range_(self.COMObject.Count): self.set_value(i, 0)
        self.set_period_length(display_period_length)
        # self.values = [self.COMObject.ValuesByIndex(i) for i in range_(self.count)]

    def add_notes(self, content=''):
        self.COMObject.Notes = _format_notes(self.notes)
        self.COMObject.Notes += content + "\r\n"
        self.notes = self.COMObject.Notes

    def save(self):
        if option['save'] not in [True, 1]: return
        self.COMObject.Notes = _format_notes(self.notes)
        self.COMObject.PeriodLength = self.period_length
        try:
            self.COMObject.Save()
        except Exception as e:
            print(e)
        print(self.COMObject.Notes)
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"\nVector updated at {current_time}.")

    def view(self):
        print(self.formula)
        self.COMObject.PeriodLength = self.period_length
        df = pd.DataFrame({'Value':[f"{self.value(i): >8,.4f}" for i in range_(self.count)]},
                           index = [f"{self.COMObject.PeriodLabel(i)}" for i in range_(self.count)])
        # return df.to_string(index=False)
        return df

    def view_prior(self, index=-1):
        p1 = ResQ.Project(DATE(index*3-1).ResQ_Project_Name).COM
        r1 = p1.GetReservingClass(proj_info['Reserving Class'])
        v1 = r1.GetVector(self.name)
        print(v1.Formula)
        print('\n')
        print(v1.Notes)


class Triangle:

    """
    Class: Triangle() can be used to generate a ResQ Triangle instance. 
    It will contain all the properties and methods of any ResQ Triangle Object.
    """

    def __init__(self, name = '', project = 'Default', reserving_class = 'Default'):

        if name == '':
            try:
                search(triangles)
                return
            except:
                print("Please select a Project and Reserving Class!")

        xProject = _project2com(project)
        xReservingClass = _reserving_class2com(project, reserving_class)

        if type(name) == str:
            self.COMObject = xReservingClass.GetTriangle(name)
        else:
            self.COMObject = name

        self.COM = self.COMObject  
        
        self.name = self.COMObject.Name
        self.calculated = self.COMObject.Calculated
        self.period_length = self.COMObject.DevelopmentLength
        self.project = xProject.Name
        self.reserving_class = xReservingClass.Path
        self.dataset_type = Dataset_Type(self.COMObject.DatasetType.Name, xProject)
        self.dataset_category = Dataset_Type_Category(self.COMObject.DatasetType.Category.Name, xProject)
        self.dev_count = self.development_count = self.COMObject.DevelopmentCount(datetime(new.year-9, 1, 1))
        self.dev_index = self.development_index = [i for i in range_(self.dev_count)]
        self.dev_label = self.development_label = [self.COM.DevelopmentLabel(i) for i in range_(self.dev_count)]
        self.org_index = self.origin_index = [i for i in range_(self.COM.OriginCount)]
        self.org_label = self.origin_label = [self.COM.OriginLabel(i) for i in range_(self.COM.OriginCount)]


    def update(self):
        self.period_length = self.COMObject.DevelopmentLength


    def value(self, row, column):
        if type(row) == str:
            row = self.origin_index[self.origin_label.index(row)]
        if type(column) == str:
            column = self.dev_index[self.dev_label.index(column)]   

        return self.COM.ValuesByIndex(row, column)
    

    def set_value(self, row, column, value):
        if type(row) == str:
            row = self.origin_index[self.origin_label.index(row)]
        if type(column) == str:
            column = self.dev_index[self.dev_label.index(column)] 

        self.COMObject.SetValuesByIndex(row, column, value)


    def diagonal(self):

        data_dict = pd.DataFrame()
        
        for index in range(10, 0, -1):
            value_list = []
            j = self.dev_count
            for i in self.origin_index:
                if i <= self.dev_count - index + 1:
                    value_list.append(self.COM.ValuesByIndex(i, j-index+1))
                else:
                    value_list.append(None)
                j = j - 1
            
            data_dict[self.origin_label[self.dev_count-index]] = value_list

        data_dict['Latest'] = data_dict[self.origin_label[-1]]
            
        data_dict.index = [self.origin_label[i] for i in range(self.dev_count)]
        return data_dict

    

    def view(self, view_in_excel = False, period_length = 'Default', pull_latest_diagonal = False, decimal = 2):

        if period_length == 'Default': 
            period_length = self.period_length

        current_time = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
        df = _get_triangle(
            self.name, 
            project = self.project,
            reserving_class = self.reserving_class,
            period_length = period_length, 
            pull_latest_diagonal = pull_latest_diagonal, 
            decimal = decimal
        )

        if view_in_excel in [True, 1]:
            pass
            # view_triangle(df, self.project, self.reserving_class, self.name)
        else:
            return df


    def set_period_length(self, value):
        self.COMObject.OriginLength = value
        self.COMObject.DevelopmentLength = value
        self.update()


    def save(self):
        self.COMObject.Save()
  

class Dataset_Type:

    """
    Class: Dataset_Type() can be used to generate a ResQ Dataset Type instance.
    """

    def __init__(self, dataset_name, project = 'Default'):

        xProject = _project2com(project)

        if type(dataset_name) == str:
            self.COMObject = xProject.GetDatasetType(dataset_name)
        else:
            self.COMObject = dataset_name

        self.name = self.COMObject.Name
        self.dataset_category = self.COMObject.Category.Name
        self.project = xProject.Name


    def delete(self):
        self.COMObject.Delete()


class Dataset_Type_Category:

    def __init__(self, category_name, project = 'Default'):

        xProject = _project2com(project)

        if type(category_name) == str:
            self.COMObject = xProject.GetCategory(category_name)
        else:
            self.COMObject = category_name

        self.name = self.COMObject.Name
        # self.dataset_types = {i.Name: Dataset_Type(i.Name, xProject) for i in xProject.DatasetTypes()}


# Auto login

try:
    if 'xwei' in user_name:
        connect_admin()
    else:
        connect()
    ResQApp.Connected # Test Connection
except:
    print('Unable to connect.')



def Help():
    Excel['Application'].Visible = 1
    Excel['Application'].Workbooks.Open("E:\\ResQ\\Automations\\library\\ToolBoxResource\\ResQToolBox_doc.xlsm", ReadOnly=True)


def update_dev_date(project_name): 

    """
    For Jan ResQ Project: Origin Period (Yr +1), Development Date (Mon +1)
    For Other Months: Development Date (Mon +1)
    """
    
    New_Project = ResQ.Project(project_name).COM
    d = New_Project.DevelopmentEndDate
    
    if d.month + 1 > 12:
    
        d1 = New_Project.OriginStartDate
        d2 = New_Project.OriginEndDate
    
        New_Project.OriginStartDate = datetime(d1.year + 1, d1.month, d1.day)
        New_Project.OriginEndDate = datetime(d2.year + 1, d2.month, d2.day)
        New_Project.DevelopmentEndDate = datetime(d.year + 1, d.month + 1 - 12, int(new.ed))
    
    else:
        New_Project.DevelopmentEndDate = datetime(d.year, d.month + 1, int(new.ed))

    New_Project.Save()


# Enable/Disable functions: method.save(), method.view() 
option = {
    'view': True, 
    'save': True, 
    'keep DFM patterns': False, 
    'keep RS patterns': False,
    'prior quarter index': -4,
    }



proj_info = {
    'Connection': "JGO_CO1SQLWPV22",
    'Project Name': None, 
    'Reserving Class': None,
    'xProject': None, 
    'xReservingClass': None
}

project_info = {
    'Connection': "JGO_CO1SQLWPV22",
    'Project Name': None, 
    'Reserving Class': None,
}

vectors = {}
triangles = {}
dataset_type = {}

reserving_classes = {}; r = reserving_classes


def set_project(name):

    if type(name) == str:
        project_info['Project Name'] = proj_info['Project Name'] = name
        proj_info['xProject'] = ResQApp.Projects().Item(name)
    else:
        project_info['Project Name'] = proj_info['Project Name'] = name.Name
        proj_info['xProject'] = name


def set_reserving_class(path: str):
    project_info['Reserving Class'] = proj_info['Reserving Class'] = path
    proj_info['xReservingClass'] = proj_info['xProject'].GetReservingClass(path)

    xReservingClass = proj_info['xReservingClass']

    for i in xReservingClass.Vectors():
        vectors[i.Name] = i

    for i in xReservingClass.Triangles():
        triangles[i.Name] = i


def set_proj_info(
        connection_name="JGO_CO1SQLWPV22", 
        project='', 
        reserving_class=''):
    """
        Default Connection Name: JGO_CO1SQLWPV22  
    """

    proj_info['Connection'] = connection_name

    if ResQApp.Connected is False:
        connect(connection_name)

    set_project(project)
    set_reserving_class(reserving_class)



class BS():
    """
    Berquist Sherman method_type: SR/CRA
    """
    def __init__(self, name, method_type='Default', project="Default", reserving_class="Default"):
        
        xProject = _project2com(project)
        xReservingClass = _reserving_class2com(project, reserving_class)

        self.xReservingClass = xReservingClass

        if type(name) == str:
            if method_type=="SR":
                self.COMObject = self.xReservingClass.GetBerquistShermanSR(name)
            elif method_type=="CRA":
                self.COMObject = self.xReservingClass.GetBerquistShermanCRA(name)
            elif method_type=="Default":
                try:
                    self.COMObject = self.xReservingClass.GetBerquistShermanSR(name)
                except:
                    self.COMObject = self.xReservingClass.GetBerquistShermanCRA(name)
            else:
                print('invalid input: method_type')
        else: # input = COM Vector
            self.COMObject = name.Method

        self.COM = self.COMObject
        self.name = self.COMObject.Name
        self.project = xProject.Name
        self.reserving_class = xReservingClass.Path

    def set_user_avg_inflation(self, dev_period, value):
        self.COM.SetUserAvgInflation(dev_period, value)
        self.COM.SetSelectedAvgInflation(dev_period, 4) # Row 4: User Value

    def save(self):
        if option['save'] not in [True, 1]: return
        self.COMObject.Save()
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"Method saved at {current_time}.")



class RS():
    """
    Returns a modified Result Selection Method Object
    """
    def __init__(self, name='', project="Default", reserving_class="Default"):

        xProject = _project2com(project)
        xReservingClass = _reserving_class2com(xProject, reserving_class)

        if name == '' and reserving_class == "Default":
            search([i.Name for i in xReservingClass.ResultSelections()])
            return

        self.name = name
        if type(name) == str:
            self.COMObject = xReservingClass.GetResultSelection(name)
        else:
            self.COMObject = xReservingClass.GetResultSelection(name.Name)
        self.COM = self.COMObject

        self.project = xProject.Name
        self.reserving_class = xReservingClass.Path
        self.notes = self.COMObject.Notes
        self.origin_length = self.COMObject.OriginLength
        self.count = self.COMObject.OriginCount
        self.datasets = {f"[{i}]": self.COMObject.Dataset(i).Name for i in range_(self.COMObject.DatasetCount)}
        self.table = None
        self.user = self.COMObject.OutputVector.User
        self.origin_label = self.label = [self.COM.OriginLabel(i) for i in range_(self.count)]


    def ensure_positive_reserve(self, basis_dataset='default'):
        """
        Remove weights for methods which produce negative reserves (IBNR + OS < 0)
        basis_dataset examples: 'Claim Counts--CWP'/'Gross Loss--Paid '...
        Give weights (1.0) to the basis dataset if no other methods selected.
        """
        if basis_dataset == 'default':
            basis_dataset = self.datasets['[1]']
            
        for AY in self.origin_label:
            for dataset_name in self.datasets.values():
                if self.dataset(dataset_name)[AY] < self.dataset(basis_dataset)[AY] or dataset_name == basis_dataset:
                    self.set_weights(dataset_name, AY, 0)
            
            if self.ultimates(AY) == 0:
                print(f"AY {AY}: No method selected, set ultimate values equal to {basis_dataset}")
                weights_to_basis = 1
            else:
                weights_to_basis = 0
                
            self.set_weights(basis_dataset, AY, weights_to_basis)


    def dataset(self, dataset_index):

        if type(dataset_index) == str:
            name2index = {self.COMObject.Dataset(i).Name: i for i in range_(self.COMObject.DatasetCount)}
            dataset_index = name2index[dataset_index]

        return {self.COMObject.OriginLabel(row_index): self.COMObject.DatasetValues(dataset_index, row_index, self.origin_length) for row_index in range_(self.COMObject.OriginCount)}


    def weights(self, dataset_index: int) -> list:
        return [self.COMObject.Weights(dataset_index, i) for i in range_(self.count)]


    def set_weights(self, dataset, row_index, value = 1.0):
        if option['keep RS patterns'] in [True, 1]: return

        name2index = {self.COMObject.Dataset(i).Name: i for i in range_(self.COMObject.DatasetCount)}
        if type(dataset) == str:
            try:
                dataset = name2index[dataset]
            except:
                for name in name2index.keys():
                    if dataset in name:
                        dataset = name2index[name]
                        break
        elif type(dataset) == int:
            pass
        else: # Input a COMObject
            dataset = name2index[dataset.Name]

        if type(row_index) == int:
            if row_index > 1000:
                row_index = row_index - int(self.COMObject.OriginLabel(1)) + 1
        elif type(row_index) == str:
            label2index = {self.COMObject.OriginLabel(i): i for i in range_(self.COMObject.OriginCount)}
            row_index = label2index[row_index]

        self.COMObject.SetWeights(dataset, row_index, value)


    def set_weights_pattern(self, pattern = 'prior Result Selection'):
        """
        Copy weight patterns from another Result_Selection, default pattern is the prior Result Selection Method
        """
        if pattern == 'prior Result Selection':
            pattern_object = self.prior()
        else:
            pattern_object = pattern
        
        def get_weights_pattern(aResultSelection):
            pattern_dict = {}
            for i in range_(aResultSelection.DatasetCount):
                pattern_dict[aResultSelection.Dataset(i).Name] = [aResultSelection.Weights(i, j) for j in range_(aResultSelection.OriginCount)]
            return pattern_dict
            
        def apply_weights_pattern(aResultSelection, pattern_dict):
            for i in range_(aResultSelection.DatasetCount):
                for j in range_(aResultSelection.OriginCount):
                    try:
                        aResultSelection.SetWeights(i, j, pattern_dict[aResultSelection.Dataset(i).Name][j-1])
                    except Exception as e:
                        print(f"✪ Cannot set weights for dataset [{aResultSelection.Dataset(i).Name}], error code: {str(e)}")

        pattern_dict = get_weights_pattern(pattern_object.COMObject)
        apply_weights_pattern(self.COMObject, pattern_dict)


    def clear_weights(self, row='All'):
        if option['keep RS patterns'] in [True, 1]: return

        if row=='All':
            row = range_(self.count)

        for dataset in self.datasets.values():
            for r_idx in row:
                self.set_weights(dataset, r_idx, 0)


    def ultimates(self, origin_index, origin_length='default'):
        if origin_length == 'default':
            origin_length = self.COMObject.OriginLength
        if type(origin_index)==str:
            origin_index = self.origin_label.index(origin_index)+1
        return self.COMObject.Ultimates(origin_index, origin_length)
        

    def set_ultimates(self, row_index: int, value):
        if option['keep RS patterns'] in [True, 1]: return
        if row_index > 1000:
            row_index = row_index - (new.year-10)
        self.COMObject.SetUltimates(row_index, self.origin_length, value)


    def revert_all_ultimate_values(self):
        """
        The UltimateOverridden property indicates whether the selected ultimate value in a Result Selection has been manually overridden by origin index.
        """
        if option['keep RS patterns'] in [True, 1]: return
        self.COMObject.ClearOverriddenUltimates()


    def set_orgin_length(self, length: int):
        self.COMObject.OriginLength = length
        self.origin_length = length


    def add_notes(self, contents: str):
        if option['keep RS patterns'] in [True, 1]: return
        if self.COMObject.Notes != '': self.COMObject.Notes += '\r\n'
        self.COMObject.Notes += contents + '\r\n'
        self.notes = self.COMObject.Notes


    def clear_notes(self):
        if option['keep RS patterns'] in [True, 1]: return
        self.COMObject.Notes = ''
        self.notes = ''


    def save(self):
        if option['save'] not in [True, 1]: return
        self.COMObject.Notes = self.notes
        self.COMObject.Save()
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"\nMethod saved at {current_time}.")
    

    def view(self, view_in_excel=True):
        if option['view'] not in [True, 1]: return
        self.COMObject.Notes = self.notes
        if view_in_excel in [True, 1]:
            _view_result_selection(self.COMObject, True)
        elif view_in_excel in [False, 0]:
            self.table = _view_result_selection(self.COMObject, False)
        Excel['Application'].Visible = 1


    def prior(self, prior_index = 'automatic'):
        if prior_index == 'automatic':
            prior_index = option['prior quarter index']
        prior_proj_name = DATE(prior_index).ResQ_Project_Name
        return RS(self.name, prior_proj_name, self.reserving_class)


    def plot(self):
        import plotly.graph_objects as go
        methods = [self.COM.Dataset(i).Name for i in range_(self.COMObject.DatasetCount)]
        
        org_count = self.COM.OriginCount
        
        np.random.seed(0)
        
        data = {}
        data['Time'] = [f"{self.COMObject.OriginLabel(i+1)}" for i in range(org_count)]
        data['Selected'] = [f"{self.ultimates(i+1)}" for i in range(org_count)]
        
        for method in methods:
            data[method] = list(self.dataset(method).values())[:org_count]
        df = pd.DataFrame(data)
        
        fig = go.Figure()
        
        for method in methods+['Selected']:
            fig.add_trace(go.Scatter(x=df['Time'], y=df[method], mode='lines+markers', name=method))
        
        fig.update_layout(
            title='Projected Ultimates by Methods',
            xaxis_title='Accident Year',
            yaxis_title='Ultimates',
            hovermode='closest',
            width=1000,
            height=500
        )
        fig.show()


    def edit(self):
        if option['view'] not in [True, 1]: return
        self.view()

        tmp_file = excel.ActiveWorkbook
        tmp_file_path = tmp_file.FullName
        sheet = tmp_file.Worksheets(1)
        info_cell = sheet.Range('C1')

        def load_weights_to_resq(sheet, self):
            weights = []
            for column in range(4, 4 + self.COMObject.DatasetCount*2, 2):
                weights.append([])
                for row in range(3, 3 + self.COMObject.OriginCount):
                    weights[-1].append(sheet.Cells(row, column).Value)
                
            for j in range(self.COMObject.DatasetCount):
                for i in range(self.COMObject.OriginCount):
                    self.COMObject.SetWeights(j + 1, i + 1, weights[j][i])
                    
        while True: 
            try:
                if os.path.basename(tmp_file_path) not in get_active_workbook_name():
                    break
            except Exception as e:
                print(e)
                break
            try:
                if info_cell.Value == 'Updating...':
                    info_cell.Value = 'Updating......'
                    load_weights_to_resq(sheet, self)
                    self.save()
                    info_cell.Value = f"Updated @ {str(datetime.now())}"
            except:
                pass
            time.sleep(2)


class DFM:
    """
    Returns a ResQ DFM Object (Instance);
    """

    def __init__(self, name='', project="Default", reserving_class="Default"):

        xProject = _project2com(project)
        xReservingClass = _reserving_class2com(xProject, reserving_class)

        if name == '' and reserving_class == "Default":
            search([i.Name for i in xReservingClass.DFMMethods()])
            return

        self.xReservingClass = xReservingClass

        if type(name) == str:
            self.COMObject = xReservingClass.GetDFMMethod(name)
        else: # input == COM Vector
            self.COMObject = name.Method

        self.COM = self.COMObject
        self.name = self.COMObject.Name
        self.project = xProject.Name
        self.reserving_class = xReservingClass.Path
        self.notes = self.COMObject.Notes
        self.origin_label = [self.COM.OriginLabel(i+1) for i in range(len(self))]

        if self.COMObject.SummaryRatioBasis is not None:
            self.summary_ratio_basis = Vector(self.COM.SummaryRatioBasis.Name, self.project, self.reserving_class)
        else:
            self.summary_ratio_basis = "Not Selected"


    def __len__(self):
        return self.COM.OriginCount
    

    def selected_cumulative_factor(self, dev_index=1):
        return self.COM.CumulativeEstimateValues(dev_index)
    

    def dev_period(self, index, format=1):
        """
        Return a development label in a string format.

        format: 
        0: (index) month-month
        1: month-month
        'start': first month of the period
        'end': last month of the period
        """
        if type(index) != int:
            start_month = self.COMObject.DevelopmentLabel(index[0]).split()[1].split('-')[0]
            end_month = self.COMObject.DevelopmentLabel(index[-1]).split()[1].split('-')[1]
            return f"{start_month}-{end_month}"
        if format==0:
            return self.COMObject.DevelopmentLabel(index)
        elif format==1:
            return self.COMObject.DevelopmentLabel(index).split()[1]
        elif format=='start':
            return self.COMObject.DevelopmentLabel(index).split()[1].split('-')[0]
        elif format=='end':
            return self.COMObject.DevelopmentLabel(index).split()[1].split('-')[1]


    def dev_month(self, index):
        return self.dev_period(index).split('-')[0]


    def info(self):
        print(f'Default Project: {self.project}\
              \nReserving Class: {self.reserving_classe}')


    def add_notes(self, text, add_space = True):
        if option['keep DFM patterns'] in [True, 1]: return
        if self.COMObject.Notes != '': 
            if add_space in [True, 0]:   
                self.COMObject.Notes += '\r\n\r\n'
            else:
                self.COMObject.Notes += '\r\n'
        text = text.replace("\n", "\r\n")
        text = text.replace("\r\r\n", "\r\n")
        if text[:2] == '\r\n':
            text = text[2:]
        if text[-2:] == '\r\n':
            text = text[:-2]
        self.COMObject.Notes += text
        self.notes = self.COMObject.Notes


    def clear(self):
        """
        Clear all notes and included all ratios.
        """
        if option['keep DFM patterns'] in [True, 1]: return
        self.COMObject.Notes = ''
        self.notes = ''
        self.include_all_ratios()
        global_dict['DFM Ratio']['Old'] = None
        global_dict['DFM Ratio']['New'] = None
        global_dict['DFM Ratio']['Dev Period'] = None


    def ratio(self, row, column): 
        if type(row) == int and type(column) == int:
            if row > 1900:
                row = row - int(self.origin_label[0]) + 1
            return self.COMObject.Ratios(row, column)
        elif type(row) == str:
            return self.COMObject.Ratios(self.origin_label.index(row)+1, column)
        

    def ultimate(self, row):
        if type(row) == int: 
            if row > 1900:
                row = row - int(self.origin_label[0]) + 1
            return self.COMObject.Ultimates(row)
        elif type(row) == str:
            return self.COMObject.Ultimates(self.origin_label.index(row)+1)
    

    def selected_ratio(self, dev_period=1):
        return round(self.COMObject.SelectedEstimateValues(dev_period), 4)


    def set_selected_estimate(self, avg_factor, dev_period='All', avg_factor_row_count=9):
        """
        avg_factor: 'Simple - 3', 'Simple - 5', 'High', 'Low'...
        dev_period: int, list, range
        avg_factor_row_count: The max row index for avg_factor
        """
        if option['keep DFM patterns'] in [True, 1]: return

        avg_index = "Not Selected"

        if type(avg_factor) == str:
            f_list = [self.COMObject.AverageFormula(i) for i in range_(self.COMObject.RatioAverageCount)]
            
            for f in f_list: # exact search
                if avg_factor == f:
                    avg_index = f_list.index(f) + 1
                    break

            if avg_index == "Not Selected": # not exact match
                for f in f_list:
                    if avg_factor in f:
                        avg_index = f_list.index(f) + 1
                        break
                        
            if avg_factor in ['High', 'Low']: # search for the highese/lowest value
                avg_index = avg_factor
            
            if avg_index == "Not Selected":
                print(f"invalid input -- avg_formula: select one of them from {f_list}")
                avg_index = 1

        elif type(avg_factor) == int:
            avg_index = avg_factor
        
        def auto_select(dev_index, avg_index):
            
            v_list = [self.COMObject.AverageRatioValues(dev_index, i) for i in range_(avg_factor_row_count)]

            if avg_index == 'High':
                v_max = max(v_list)
                for i in range_(avg_factor_row_count):
                    ratio_value = self.COMObject.AverageRatioValues(dev_index, i)
                    if ratio_value == v_max:
                        selected_avg_index = v_list.index(ratio_value) + 1
                        break
                        
            elif avg_index == 'Low':
                v_min = min(v_list)
                for i in range_(avg_factor_row_count):
                    ratio_value = self.COMObject.AverageRatioValues(dev_index, i)
                    if ratio_value == v_min:
                        selected_avg_index = v_list.index(ratio_value) + 1
                        break
            else:
                selected_avg_index = avg_index
                
            self.COMObject.SetSelectedRatios(dev_index, selected_avg_index)


        if type(dev_period) == int:
            auto_select(dev_period, avg_index)
            
        elif dev_period == 'All':
            for dev in range_(self.COMObject.DevelopmentCount(1)):
                auto_select(dev, avg_index)
        else: # range or list
            for dev in dev_period:
                auto_select(dev, avg_index)


    def set_user_value(self, value, dev_period=1, row_index=10):
        """
        Set and select the value for 'User Entry' row (10~12)
        """
        if option['keep DFM patterns'] in [True, 1]: return
        self.COMObject.SetUserRatios(dev_period, row_index, value)
        self.COMObject.SetSelectedRatios(dev_period, row_index)


    def include_all_ratios(self):
        if option['keep DFM patterns'] in [True, 1]: return
        a = range(1, self.COM.OriginCount+1)
        for i, j in itertools.product(a, a): self.COM.SetExcludedRatios(i, j, 0)


    def clear_notes(self):
        if option['keep DFM patterns'] in [True, 1]: return
        self.COMObject.Notes = ''
        self.notes = ''


    def ex_hi(self, dev_period, count=1, reason='', add_notes=True):
        if option['keep DFM patterns'] in [True, 1]: return
        self.COMObject.Notes = self.notes 

        if global_dict['DFM Ratio']['Old'] == None:
            global_dict['DFM Ratio']['Dev Period'] = dev_period
            global_dict['DFM Ratio']['Old'] = self.selected_ratio(dev_period)
            
        _ex_hi(self.COMObject, dev_period, count, reason, add_notes)

        if dev_period == global_dict['DFM Ratio']['Dev Period']:
            global_dict['DFM Ratio']['New'] = self.selected_ratio(dev_period)

        self.notes = self.COMObject.Notes   


    def ex_lo(self, dev_period, count=1, reason='', add_notes=True):
        if option['keep DFM patterns'] in [True, 1]: return
        self.COMObject.Notes = self.notes 

        if global_dict['DFM Ratio']['Old'] == None:
            global_dict['DFM Ratio']['Dev Period'] = dev_period
            global_dict['DFM Ratio']['Old'] = self.selected_ratio(dev_period)

        _ex_lo(self.COMObject, dev_period, count, reason, add_notes)

        if dev_period == global_dict['DFM Ratio']['Dev Period']:
            global_dict['DFM Ratio']['New'] = self.selected_ratio(dev_period)

        self.notes = self.COMObject.Notes


    def select_high(self, dev_period=1, count=1, reason='', add_notes=True):
        if option['keep DFM patterns'] in [True, 1]: return

        self.COMObject.Notes = self.notes 

        if global_dict['DFM Ratio']['Old'] == None:
            global_dict['DFM Ratio']['Dev Period'] = dev_period
            global_dict['DFM Ratio']['Old'] = self.selected_ratio(dev_period)

        _select_high(self.COMObject, dev_period, count, reason, add_notes)

        if dev_period == global_dict['DFM Ratio']['Dev Period']:
            global_dict['DFM Ratio']['New'] = self.selected_ratio(dev_period)

        self.notes = self.COMObject.Notes


    def select_low(self, dev_period=1, count=1, reason='', add_notes=True):
        if option['keep DFM patterns'] in [True, 1]: return
        self.COMObject.Notes = self.notes 

        if global_dict['DFM Ratio']['Old'] == None:
            global_dict['DFM Ratio']['Dev Period'] = dev_period
            global_dict['DFM Ratio']['Old'] = self.selected_ratio(dev_period)

        _select_low(self.COMObject, dev_period, count, reason, add_notes)

        if dev_period == global_dict['DFM Ratio']['Dev Period']:
            global_dict['DFM Ratio']['New'] = self.selected_ratio(dev_period)

        self.notes = self.COMObject.Notes


    def ex_LDF(self, dev_period, row, reason = '', add_notes=True):
        if option['keep DFM patterns'] in [True, 1]: return
        self.COMObject.Notes = self.notes 
        _ex_LDF(self.COMObject, dev_period, row, reason, add_notes)
        self.notes = self.COMObject.Notes


    def ex_row(self, row, add_notes=True):
        if option['keep DFM patterns'] in [True, 1]: return
        self.COMObject.Notes = self.notes 
        _ex_row(self.COMObject, row, add_notes)
        self.notes = self.COMObject.Notes

    def ex_AY(self, AY, add_notes=True):
        if option['keep DFM patterns'] in [True, 1]: return
        """
        Exclude ratios for an accident year.
        """
        self.COMObject.Notes = self.notes 
        _ex_AY(self.COMObject, AY, add_notes)
        self.notes = self.COMObject.Notes

    def ex_COVID_AY(self, AY:list=[2020,2021], add_notes=True):
        if option['keep DFM patterns'] in [True, 1]: return
        """
        Exclude ratios for accident year 2020 and 2021.
        """
        self.COMObject.Notes = self.notes 
        _ex_COVID_AY(self.COMObject, AY, add_notes)
        self.notes = self.COMObject.Notes

    def ex_diagonal(self, dev_index, start_row="min_row", end_row="max_row", reason='', add_notes=True):
        """
        dev_index: start counting from the right hand side (latest development period)
        """
        if option['keep DFM patterns'] in [True, 1]: return
        self.COMObject.Notes = self.notes 
        _ex_diagonal(self.COMObject, dev_index, start_row, end_row, reason, add_notes)
        self.notes = self.COMObject.Notes


    def set_tail_value(self, dev_period, n_year = "Max", exclude = None, value_list = 'Auto Search', historical_ratio_data = 'Prior DFM', ignore_pending_claims=False):
        """
        value_list: 'Auto Search'/ List / pd.core.series.Series
        n_year: "Max"/integer
        exclude: None/"Low"/"High"
        value_list: if
        historical_ratio_data: Works with the 'value_list' arg. The DFM object which is used for historical data points
        """
        if option['keep DFM patterns'] in [True, 1]: return
        if str(value_list) == 'Auto Search':
            value_list = []

            if historical_ratio_data == 'Prior DFM':
                DFM_Triangle = self
            else:
                DFM_Triangle = DFM(historical_ratio_data)

            # add data from row 2 to (10-dev_period)
            for row in range_(2, 10-dev_period):  # if (10-dev_period) < 2, nothing will be added
                value_list.append(self.ratio(row, dev_period))
                
            if n_year == "Max":
                for i in range(10):
                    try: 
                        if i == 0:
                            value_list.append(self.prior(-1-12*i).ratio(1, dev_period))
                        else:
                            value_list.append(DFM_Triangle.prior(-1-12*i).ratio(1, dev_period))
                    except: 
                        break
                n_year = len(value_list)
            else: 
                for i in range(n_year-1):
                    try: value_list.append(DFM_Triangle.prior(-1-12*i).ratio(1, dev_period))
                    except: 
                        print(f"Invalid input for 'n_year' at development period {DFM_Triangle.dev_period(dev_period)}, the maximum number of year is {len(value_list)}.")
                        n_year = i+1
                        break

        elif type(value_list) == list:
            if type(n_year) == int:
                value_list = value_list[0: n_year]
            pass

        elif type(value_list) == pd.core.series.Series:
            value_list = list(value_list)

            while str(value_list[0]) == 'nan':
                value_list.pop(0)
                
            if type(n_year) == int:
                value_list = value_list[0: n_year]
            
        elif str(type(value_list)) == "<class 'dict_values'>":
            value_list = list(value_list)
            if type(n_year) == int:
                value_list = value_list[0: n_year]

        else:
            print("Invalid Input for arg: 'value_list'")
            return


        for i in range(len(value_list)):
            value_list[i] = round(value_list[i], 4)
            
        excluded = ""
        if exclude == "Low":
            excluded = "ex low "
            removed_value_index = value_list.index(min(value_list))
        elif exclude == "High":
            excluded = "ex high "
            removed_value_index = value_list.index(max(value_list))
            
        new_list = value_list.copy()
        
        if excluded != "":
            new_list.pop(removed_value_index)

        avg = sum(new_list)/len(new_list)

        CCP = Triangle('Claim Counts--Pending', self.project, self.reserving_class)
        CCPL = CCP.diagonal()
        row_index_2 = self.COM.DevelopmentCount(1)-dev_period
        note = ''

        if ignore_pending_claims == False and CCPL['Latest'][CCP.origin_label[row_index_2]] == 0 and self.COM.OutputVector.DatasetType.Category.Name not in ['G ALAE', 'H Severity']:  # 2nd row, dev period (9)
            self.set_user_value(1, dev_period)
            note += f"For development period {self.dev_period(dev_period)}, selected an LDF of 1.0 since {CCP.org_label[row_index_2]} has no open claims;"
        else:
            self.set_user_value(avg, dev_period)
            self.set_selected_estimate(avg_factor = '10: User Entry', dev_period = dev_period)
            note += f"For development period {self.dev_period(dev_period)}, selected a {n_year}-year {excluded}average = ("
            note += f" + ".join(map(str, new_list)) + f")/{len(new_list)} = {round(avg, 4)};"

        self.add_notes(note)
        

    def set_ratio_patterns(self, aDFM = 'prior DFM', row_index = 'all', col_index = 'all', row_offset = 'automatic', col_offset = 0):
        """
        Copy the selected/excluded LDF ratio patterns from another DFM.
        Example: If aDFM starts from 2014 Q1 and current DFM starts from 2015 Q1, the row offset will be set to -4 automatically.
        """
        if option['keep DFM patterns'] in [True, 1]: return

        if aDFM == 'prior DFM':
            aDFM = self.prior()

        org_count = self.COMObject.OriginCount
        
        if row_index == 'all':
            row_index = range_(org_count)
        if col_index == 'all':
            col_index = range_(org_count)

        if row_offset == 'automatic':
            try:
                row_offset = int(aDFM.COMObject.OriginLabel(1)[:4]) - int(self.COMObject.OriginLabel(1)[:4])
                if 'Q' in self.COMObject.OriginLabel(1):
                    row_offset = row_offset*4
            except:
                row_offset = 0
        else:
            row_offset = 0

        for j in col_index: 
            for i in row_index: 
                if i + j > 40: continue
                self.COMObject.SetExcludedRatios(i, j, aDFM.COMObject.ExcludedRatios(i-row_offset, j-col_offset))


    def set_average_formula_patterns(self, aDFM='prior DFM', col_index='all', skip_user_entry_values=True):
        """
        Copy the average factors selection patterns from another DFM
        skip_user_entry_values: if the reference DFM has 'User Entry' as the selected average formula, don't apply that selection to the target DFM
        """
        if option['keep DFM patterns'] in [True, 1]: return

        if aDFM == 'prior DFM':
            aDFM = self.prior()

        dev_counts = self.COMObject.DevelopmentCount(1)

        if col_index == 'all':
            col_index = range_(dev_counts)

        for i in col_index:
            if 'User Entry' in aDFM.COM.AverageFormula(aDFM.COM.SelectedRatios(i)) and skip_user_entry_values:
                continue
            self.COMObject.SetSelectedRatios(i, aDFM.COMObject.SelectedRatios(i))


    def get_average_factors(self):
        formulas = [self.COMObject.AverageFormula(i) for i in range_(self.COMObject.RatioAverageCount)]
        for f in formulas: print(f"'{f}'")


    def apply_adjustments(self, avg_type='Default', other_adjustment=[0]):
        """
        Use this method to apply COVID/Growth Adjustments and perform the calculations automatically.

        Input Values:
        avg_type (str): 'Simple - 3'/'Simple - 5'/'Simple - 12'/'Simple - 20';
        adjustments (dict)

        dev (int): Development Period;

        """
        _final_selection(xMethod=self.COMObject, avg_type=avg_type, other_adjustment = other_adjustment)
        self.notes = self.COMObject.Notes


    def save(self):
        if option['save'] not in [True, 1]: return
        self.COMObject.Notes = self.notes
        # if 'due to' in self.COMObject.Notes and ' due to' not in self.COMObject.Notes:
        #     self.COMObject.Notes = self.COMObject.Notes.replace('due to', ' due to')
        try:
            self.COMObject.Save()
            print(re.sub(r'(?<!\r)\n', '', self.COMObject.Notes))
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(f"\nMethod saved at {current_time}.")
        except Exception as e:
            print(e)


    def view(self, editable=False):
        if option['view'] not in [True, 1]: return
        self.COMObject.Notes = self.notes
        _preview(self.COMObject, editable)


    def prior(self, index = 'automatic'):
        if index == 'automatic':
            index = option['prior quarter index']
        # proj_list=[i.Name for i in ResQApp.Projects() if '-' in i.Name and len(i.Name) <= 26 and 'NJ_Annual_Prod' in i.Name]  
        prior_proj_name = DATE(index).ResQ_Project_Name
        prior_DFM = DFM(self.name, project=prior_proj_name, reserving_class=self.reserving_class)
        return prior_DFM


    def view_prior(self, index=-1):
        if option['view'] not in [True, 1]: return
        self.prior(index).view()


    def view_prior_notes(self, index=-1):
        proj_list=[i.Name for i in ResQApp.Projects() if '-' in i.Name and len(i.Name) <= 26 and 'NJ_Annual_Prod' in i.Name]  
        prior_proj_name = proj_list[index-1]
        prior_DFM = ResQ.Project(prior_proj_name).COM.GetReservingClass(self.reserving_class).GetDFMMethod(self.name)
        print(f"Notes from [{prior_proj_name}]:\n")
        print(prior_DFM.Notes)
        print('\n')


    def set_summary_ratio_basis(self, basis_object, data_type='Vector'):
        if type(basis_object) == str:
            basis_object_name = basis_object
        elif str(type(basis_object)) == "<class 'ResQToolBox2.Vector'>":
            basis_object_name = basis_object.name
        elif str(type(basis_object)) == "<class 'win32com.client.CDispatch'>":
            basis_object_name = basis_object.Name
    
        if data_type=='Vector':
            self.COMObject.SummaryRatioBasis = self.xReservingClass.GetVector(basis_object_name)
        elif data_type=='Triangle':
            self.COMObject.SummaryRatioBasis = self.xReservingClass.GetTriangle(basis_object_name)
    
        self.summary_ratio_basis = self.COMObject.SummaryRatioBasis.Name
        print(f"Summary Ratio Basis changed to [{self.summary_ratio_basis}]")


    def reset_ratio_basis(self):
        '''
        Set the summary ratio basis same as prior quarter's
        '''
        if self.summary_ratio_basis == "Not Selected":
            if self.prior().COMObject.SummaryRatioBasis.DatasetType.DataFormat == 1:
                self.set_summary_ratio_basis(self.prior().summary_ratio_basis, 'Vector')
            else:
                self.set_summary_ratio_basis(self.prior().summary_ratio_basis, 'Triangle')


    def quick_preview(self):
        """
        View the DFM ratio triangles for current project & prior-year project.
        """
        self.plot_diagnostics()
        self.view_prior(-4)
        self.view()


    def plot_diagnostics(self, dev_count=range(3), triangle = 'Default', transpose=False):
        """
        Visualize the diagnostic triangle data related to current DFM method.
        dev_count: range/list/integer
        triangle: Enter the name of the dataset (triangle); If 'Default', lookup value in the mapping table: "E:\\ResQ\\Automations\\library\\ToolBoxResource\\diagnostic_mapping.xlsx"
        """
        dataset_map = {}
        file_location = "E:\\ResQ\\Automations\\library\\ToolBoxResource\\diagnostic_mapping.xlsx"
        wb = load_workbook(file_location)
        ws = wb['Sheet1']

        for row in range(2, ws.max_row+1):
            dataset_map[ws.cell(row, 1).value] = ws.cell(row, 2).value

        if triangle == 'Default':
            try:
                triangle_name = dataset_map[self.name]
            except:
                print(f"Diagnostics Triangle Not Found in Mapping Table!\nYou can add records in {file_location}")
                print(f"u.startfile(r'{file_location}')")
                return
        else:
            triangle_name = triangle

        try:
            _plot_triangle(
                self.xReservingClass.GetTriangle(triangle_name), 
                dev_count, 
                self.xReservingClass.GetDFMMethod(self.name).OriginCount, transpose
            )
        except Exception as e:
            print(e)
            

    def plot_ultimates(self, show_ratios=False, D=None):
        """
        Visualize the ultimate selected (projected) vector for the current DFM method.
        """
        try:
            _plot_ultimates(self.COMObject, show_ratios, D)
        except:
            self.reset_ratio_basis()
            _plot_ultimates(self.COMObject, show_ratios, D)


    def extended_ratio_data(self):
        """
        This function will return a list of dataframe: [historical ratios, excluded ratios]

        """
        return extended_DFM_triangle(self.COMObject)


    def forcast(self, steps = 5):
        '''
        Using ARIMA (Autoregressive Integrated Moving Average)
        To address potential non-stationarity issues in the data and to find the most suitable parameters for the ARIMA model. 
        auto_arima from the pmdarima library automates the process of ARIMA model selection by trying different combinations of p, d, q (and seasonal parameters P, D, Q, s if seasonality is expected) to find the best model based on a given criterion, usually the AIC (Akaike Information Criterion).
        '''
        import plotly.graph_objs as go
        import pmdarima as pm
        
        ult_values = [self.COMObject.Ultimates(value) for value in range_(1, self.COMObject.DevelopmentCount(1))]
    
        for plot_label in ['Incurred', 'Counts', 'Paid', 'Value']:
            if plot_label in self.name:
                break
    
        num_training_points = len(ult_values) - steps
        
        # Generating the date range
        quarters = pd.date_range(start="2013Q1", periods=len(ult_values), freq='Q')
        
        # Creating the dataframe
        df = pd.DataFrame(data=ult_values, index=quarters, columns=[plot_label])
        
        # Splitting the dataset into training and test sets
        train_data = df.iloc[:num_training_points]
        test_data = df.iloc[num_training_points:]
        
        # Using auto_arima to find the best ARIMA model
        auto_model = pm.auto_arima(train_data, seasonal=True, m=4, trace=True,
                                error_action='ignore', suppress_warnings=True,
                                stepwise=True)
        
        print(auto_model.summary())
        
        # Forecasting the next 5 quarters with the best model
        auto_forecast, auto_conf_int = auto_model.predict(n_periods=steps, return_conf_int=True)
        
        # Creating a modified test data for comparison
        test_data_mod = test_data.copy()
        test_data_mod['Forecast'] = auto_forecast
        test_data_mod['Lower_CI'] = auto_conf_int[:, 0]
        test_data_mod['Upper_CI'] = auto_conf_int[:, 1]
        
        
        quarter_labels = [f"{x.year} Q{(x.month-1)//3 + 1}" for x in quarters] # Creating quarter labels
        fig = go.Figure() # Creating the Plotly figure
        
        # Adding the training data plot
        fig.add_trace(go.Scatter(x=quarter_labels[:num_training_points], y=train_data[plot_label],
                                mode='lines+markers', name='Training Data', line_shape='spline'))
        
        # Actuarial selected
        fig.add_trace(go.Scatter(x=quarter_labels[num_training_points:], y=test_data_mod[plot_label],
                                mode='lines+markers', name='Actuarial Selected', line_shape='spline'))
        
        # Forecasted line
        fig.add_trace(go.Scatter(x=quarter_labels[num_training_points:], y=test_data_mod['Forecast'],
                                mode='lines+markers', name='Forecasted Value', line_shape='spline',
                                text=[f'Forecast: {y:.2f}' for y in test_data_mod['Forecast']],
                                hoverinfo='text+x'))
        
        # Confidence interval (as a shaded area)
        fig.add_trace(go.Scatter(x=quarter_labels[num_training_points:] + quarter_labels[num_training_points:][::-1],
                                y=test_data_mod['Upper_CI'].tolist() + test_data_mod['Lower_CI'].tolist()[::-1],
                                fill='toself', fillcolor='rgba(0,100,80,0.2)',
                                line=dict(color='rgba(255,255,255,0)'),
                                hoverinfo="skip",
                                showlegend=False))
        
        # Connect 1
        last_train_point = train_data.iloc[-1]
        first_actual_point = test_data.iloc[0]
        connect_trace = go.Scatter(
            x=[quarter_labels[num_training_points-1], quarter_labels[num_training_points]],  # Indices 32 and 33 correspond to the last and first points
            y=[last_train_point[plot_label], first_actual_point[plot_label]],
            mode='lines',
            line=dict(color='gray', dash='dash'),
            showlegend=False
        )
        
        # Connect 2
        last_train_point = train_data.iloc[-1][plot_label]
        first_forecast_point = test_data_mod.iloc[0]['Forecast']
        connect_forecast_trace = go.Scatter(
            x=[quarter_labels[num_training_points-1], quarter_labels[num_training_points]],  # Indices 32 and 33 for the last training and first forecast points
            y=[last_train_point, first_forecast_point],
            mode='lines',
            line=dict(color='blue', dash='dash'),
            showlegend=False
        )
        
        
        fig.add_trace(connect_forecast_trace) # Adding the connection trace to the figure
        fig.add_trace(connect_trace) # Adding the connection trace to the figure
        
        # Updating the layout to include all quarter labels and adjust the size
        fig.update_layout(title=f'Quarterly {plot_label} with ARIMA Forecast',
                        xaxis=dict(title='Accident Quarter', tickvals=list(range(len(quarter_labels))), ticktext=quarter_labels),
                        yaxis_title=plot_label,
                        hovermode='closest',
                        width=1000,  # Width of the figure in pixels
                        height=500)  # Height of the figure in pixels
        
        # Display the plot
        fig.show()


    def decompose(self):

        import matplotlib.pyplot as plt
        import statsmodels.api as sm
        from statsmodels.tsa.arima.model import ARIMA
        from pandas.plotting import register_matplotlib_converters
        register_matplotlib_converters()
        
        from statsmodels.tsa.seasonal import seasonal_decompose
        
        ult_values = [self.COMObject.Ultimates(value) for value in range_(1, self.COMObject.DevelopmentCount(1))]
        # Generating the date range
        quarters = pd.date_range(start="2013Q1", periods=len(ult_values), freq='Q')
        
        # Creating the dataframe
        df = pd.DataFrame(data=ult_values, index=quarters, columns=['Counts'])
        
        # Decomposing the time series
        decomposition = seasonal_decompose(df, model='additive', period=4)
        
        # Extracting the trend, seasonality, and residuals
        trend = decomposition.trend
        seasonal = decomposition.seasonal
        residual = decomposition.resid
        
        # Plotting the decomposed components
        plt.figure(figsize=(14, 8))
        
        # Trend
        plt.subplot(411)
        plt.plot(trend, label='Trend')
        plt.legend(loc='best')
        plt.xticks(rotation=45)
        
        # Seasonality
        plt.subplot(412)
        plt.plot(seasonal,label='Seasonality')
        plt.legend(loc='best')
        plt.xticks(rotation=45)
        
        # Residuals
        plt.subplot(413)
        plt.plot(residual, label='Residuals')
        plt.legend(loc='best')
        plt.xticks(rotation=45)
        
        # Original Data
        plt.subplot(414)
        plt.plot(df, label='Original')
        plt.legend(loc='best')
        plt.xticks(rotation=45)
        
        plt.tight_layout()
        plt.show()


    def offset(self):
        """
        Calculate offset for dev_period (n) due to adjustments made for dev_period (n+1);
        After applying this method, dev_period (n) will not be affected by selections for dev_period (n+1).

        The following methods will updated the stored LDF values automatically:
        ex_hi/ex_lo/select_high/select_low;
        
        If needs manual update, use global_dict['DFM Ratio']
        global_dict['DFM Ratio']['Old']: selected LDF for dev_period (n+1) before any adjustment
        global_dict['DFM Ratio']['New']: selected LDF for dev_period (n+1) after any adjustment;
        """
        dev_period = global_dict['DFM Ratio']['Dev Period'] - 1
        adj_ratio = self.selected_ratio(dev_period) * global_dict['DFM Ratio']['Old'] / global_dict['DFM Ratio']['New']
        
        note = f"Adjust the selected LDF for {self.dev_period(dev_period)} to offset the low LDF selected for {self.dev_period(dev_period+1)}.\n"
        note += f"{self.selected_ratio(dev_period)} * {global_dict['DFM Ratio']['Old']} / {global_dict['DFM Ratio']['New']} = {round(adj_ratio, 4)}"

        self.add_notes(note)
        
        self.set_user_value(adj_ratio, dev_period)

        global_dict['DFM Ratio']['Old'] = None
        global_dict['DFM Ratio']['New'] = None
        global_dict['DFM Ratio']['Dev Period'] = None


    def set_custom_averages(
            self, 
            avg_index, 
            avg_name, 
            periods_included, 
            weight_type = 'Volume', 
            ex_hi_lo = 0
            ):
        
        cm = self.COM.CustomAverages(avg_index)

        cm.Name = avg_name # 'Simple - xx

        if weight_type=='Simple':
            cm.WeightType = 0

        if weight_type=='Volume':
            cm.WeightType = 1

        if ex_hi_lo == 0:
            cm.ExcludeHighLow = False
            cm.ExcludeHighLow2 = 0
        else:
            cm.ExcludeHighLow = True
            cm.ExcludeHighLow2 = ex_hi_lo # Exclude High/Low Ratios (number of times) 

        cm.PeriodsIncluded = periods_included
        # cm.SaveAsDefault()
        

class CCM():

    """
    Cape Code Method

    """

    def __init__(self, name):
        sProject = proj_info['Project Name']
        sReservingClass = proj_info['Reserving Class']
        self.name = name
        self.xProject = ResQApp.Projects().Item(sProject)
        self.xReservingClass = self.xProject.GetReservingClass(sReservingClass)
        self.COM = self.COMObject = self.xReservingClass.GetCapeCodeMethod(name)
        self.type = 'ResQ Cape Code Method'
    
    def save(self):
        if option['save'] not in [True, 1]: return
        self.COMObject.Save()
        # print(self.COMObject.Notes)
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"\nMethod saved at {current_time}.")


class BF():
    def __init__(self, name, project="Default", reserving_class="Default"):

        xProject = _project2com(project)
        xReservingClass = _reserving_class2com(xProject, reserving_class)

        self.name = name
        self.COM = self.COMObject = xReservingClass.GetBFMethod(name)
        self.reserving_class = xReservingClass.Path
        self.project = xProject.Name

        self.type = 'ResQ BF Method'
        self.notes = self.COMObject.Notes
        self.projected_ultimates = None
        self.update()
    
    def set_prior(self, prior_vector_name):
        self.COM.Prior = Vector(prior_vector_name).COM

    def set_percentage_developed(self, dataset_name, dataset_type=2):
        """
        ResQ.percentage_developed_type = {'Latest/Ultimates': 0, 'Pattern vector': 1, 'DFM dev factors': 2}
        """
        self.COM.PercentageDeveloped = Vector(dataset_name).COM

    def load_settings_from(self, aBF):
        """
        aDFM: enter the reserving class path or ResQToolBox2.DFM object
        """
        if type(aBF) == str:
            aBF = BF(self.name, '', aBF)

        self.set_percentage_developed(aBF.COM.PercentageDeveloped.Name)
        self.set_prior(aBF.COM.Prior.Name)
        print(f"% Developed:  {self.COM.PercentageDeveloped.Name}\nPrior Vector: {self.COM.Prior.Name}")
    
    def update(self):
        pass
        # self.projected_ultimates = [self.COMObject.Ultimates(i) for i in range_(self.count)]

    def clear_notes(self):
        self.COMObject.Notes = ''
        self.notes = self.COMObject.Notes

    def view(self): 
        print(f"% Developed:  {self.COM.PercentageDeveloped.Name}\nPrior Vector: {self.COM.Prior.Name}")
        return Vector(self.name).view()

    def save(self):
        if option['save'] not in [True, 1]: return
        self.COMObject.Notes = _format_notes(self.notes)

        try:
            self.COMObject.Save()
        except Exception as e:
            print(e)

        if self.COMObject.Notes != '':
            print(self.COMObject.Notes)

        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"\nMethod saved at {current_time}.")



# Full Name

class Result_Selection(RS):

    def __init__(self, name):
        super().__init__(name)


class DFM_Method(DFM):

    """
    Use the DFM class to generate an instance
    name: str
    return a modified COM onject
    =========================================
    note:
        this class is the same as class: DFM
    """

    def __init__(self, name):
        super().__init__(name)


class Cape_Code_Method(CCM):

    def __init__(self, name):
        super().__init__(name)


class BF_Method(BF):

    def __init__(self, name):
        super().__init__(name)



def ResQFolder(Folder_name):
    xFolder = ResQApp.ProjectFolders().Item(Folder_name)
    # print(xFolder.Name)
    return xFolder


def _add_project_folder(NewFolderName, ParentFolder=None):  # input == string variables
    xFolder = ResQApp.ProjectFolders().Add()
    xFolder.Name = NewFolderName
    if ParentFolder is not None:
        xFolder.ParentFolder = ResQFolder(ParentFolder)
    xFolder.Save()


def _add_reserving_class_type(name, level = 1, project = proj_info['Project Name']):

    xProject = _project2com(project)
    new_resc_type = xProject.ReservingClassTypes().Add()
    new_resc_type.Name = name
    new_resc_type.Level = level
    new_resc_type.Save()

    print('Reserving Class Type Added!')


def _add_reserving_class_path(project, parent_class, reserving_class_type):
    """
    Add a new reserving Class Path (ResQFolder) in a ResQ Project
    """
    xProject = _project2com(project)
    new_path = xProject.ReservingClasses().Add()
    new_path.Name = reserving_class_type
    new_path.ParentClass = xProject.GetReservingClass(parent_class)
    new_path.ReservingClassType = xProject.GetReservingClassType(reserving_class_type)
    new_path.Save()


def _add_dataset(project, reserving_class, dataset_type: str, name='default', formula='default'):
    project = _project2com(project)
    reserving_class = _reserving_class2com(project, reserving_class)
    
    dataset_type = project.GetDatasetType(dataset_type)
    data_format = dataset_type.DataFormat
    if data_format == 1:
        new_dataset = reserving_class.Vectors().Add()
    elif data_format == 0:
        new_dataset = reserving_class.Triangles().Add()
    
    if name == 'default':
        name = dataset_type.Name
    new_dataset.Name = name
    new_dataset.DatasetType = dataset_type
    
    if formula != 'default':
        new_dataset.Formula = formula
    try:
        new_dataset.Save()
    except Exception as e:
        if 'cannot be more than one' in str(e):
            pass
        else:
            print(e)


def _add_method(name, method_type, project='Default', reserving='Default'):
    """
    see method_type values from ResQ.method_type
    """
    xProject = _project2com(project)
    xReservingClass = _reserving_class2com(reserving)

    new_method = xReservingClass.AddMethod(method_type)
    new_method.Name = name
    new_method.OutputVector.Name = name
    new_method.OutputVector.DatasetType = xProject.DatasetTypes().Item(name)

    new_method.Save()


def _add_BFMethod(name, method_type=ResQ.method_type['BF'], project='Default', xReservingClass='Default'):

    if type(name) != str:
        name = name.Name

    xProject = _project2com(project)

    if xReservingClass == 'Default':
        xReservingClass = proj_info['xReservingClass']
    elif type(xReservingClass) == str:
        xReservingClass = xProject.GetReservingClass(xReservingClass)

    # Create Instance
    new_method = xReservingClass.AddMethod(method_type)
    new_method.Name = name
    new_method.OutputVector.Name = name
    new_method.OutputVector.DatasetType = xProject.DatasetTypes().Item(name)

    # Set Properties (Copy from last project)

    old_proj = projects['NJ_Annual_Prod_2022 Q4-Nov']
    old_BFM = old_proj.GetReservingClass(proj_info['Reserving Class']).GetBFMethod(name)

    new_method.OriginLength = old_BFM.OriginLength

    new_method.LatestType = 0 # ResQ.dataset_type Triangle
    new_method.Latest = proj_info['xReservingClass'].GetTriangle(old_BFM.Latest.Name)

    new_method.PercentageDevelopedType = ResQ.percentage_developed_type['DFM dev factors']
    new_method.PercentageDeveloped = vectors[old_BFM.PercentageDeveloped.Name]

    new_method.Prior = vectors[old_BFM.Prior.Name]
    new_method.PriorType = old_BFM.PriorType  # 0 == ultimates

    new_method.Save()


def _add_CCMethod():
    xP = proj_info['xProject']
    xR = proj_info['xReservingClass']

    method_name = 'D 53 - Cape Cod Gross Loss Incurred'

    CC = xR.AddMethod(3)
    CC.Name = method_name
    CC.OutputVector.Name = method_name
    CC.OutputVector.DatasetType = xP.DatasetTypes().Item(method_name)

    CC.OriginLength = 3
    CC.Exposure = xR.GetVector('Total Earned Exposure')

    CC.LatestType = 0 # ResQ.dataset_type['Triangle']
    CC.Latest = xR.GetTriangle('Gross Loss--Incurred')

    CC.PercentageDevelopedType = ResQ.percentage_developed_type['DFM dev factors']
    CC.PercentageDeveloped = xR.GetVector('D 23 - Incurred DFM w/ Selected LDFs')

    CC.AutoTrendFit = True
    CC.DecayFactor = 0.50
    CC.Save()


def _add_DFM(
    project, 
    reserving_class, 
    method_name: str, 
    input_triangle: str,
    output_vector='default',
    name='default',
    period_length = 12
):
    if reserving_class.Vectors().Item(method_name) is not None:
        reserving_class.Vectors().Item(method_name).Delete()
    
    DFM = reserving_class.AddMethod(ResQ.method_type['DFM'])
    DFM.Name = method_name
    if reserving_class.Triangles().Item(input_triangle) is None:
        _add_dataset(project, reserving_class, input_triangle)
    DFM.InputTriangle = reserving_class.Triangles().Item(input_triangle)
    DFM.OutputVector.Name = DFM.Name
    DFM.OutputVector.DatasetType = project.DatasetTypes().Item(DFM.Name)
    DFM.OriginLength = period_length
    DFM.DevelopmentLength = period_length
    
    DFM.Save()


def _add_dataset_type(xProject, ds_name, ds_cat, aggregate = True, unique = True):
    D = xProject.DatasetTypes().Add()
    D.Name = ds_name
    D.Category = xProject.Categories().Item(ds_cat)
    D.Aggregated = aggregate
    D.Unique = unique
    D.Save()
    return D


def _duplicate_project(old_proj, new_proj, output_folder="current", write_log = False):

    if type(old_proj) == str:  # old.proj -> a project object
        old_proj = ResQApp.Projects().Item(old_proj)

    if old_proj is None:  # Check if old proj exists
        if write_log == False:
            print("The old project does not exist.")
        return
    
    if type(new_proj) != str:
        return
    
    if ResQApp.Projects().Item(new_proj) is not None:
        if write_log == False:
            print("The new project name already exists.")
        return

    if write_log != False:
        log(f'Duplicating project - {old_proj.Name}')

    New_Proj = old_proj.Duplicate()
    New_Proj.Name = new_proj

    if write_log == False:
        print(f"New project [{New_Proj.Name}] created. Project settings unchanged.")
    else:
        log(f"New project [{New_Proj.Name}] created. Project settings unchanged.")

    if output_folder != "current":
        New_Proj.Folder = ResQFolder(output_folder)
    New_Proj.Save()


def get_user_info(limit=15):
    info = {}
    for u in ResQApp.Users(): 
        info[u.Name] = u.DateLastAccessed.strftime("%Y-%m-%d %H:%M:%S")

    sorted_info = sorted(info.items(), key=lambda x:x[1], reverse=1)
    for i in range_(limit):
        if i < 10:
            space = " "
        else:
            space = ""
        print(f"[{i}] {space}", f"{sorted_info[i-1][0]: <26}", sorted_info[i-1][1])


def _get_triangle(
        triangle_name,
        project = 'Default', 
        reserving_class = 'Default', 
        period_length: int = 12, 
        pull_latest_diagonal: bool = False,
        decimal: int = 2,
        print_name: bool = False):
    
    if type(triangle_name) == str:

        xReservingClass = _reserving_class2com(project, reserving_class)
        T = xReservingClass.GetTriangle(triangle_name)
    else:
        T = triangle_name

    if T is None: return None

    if print_name is True:
        try:
            print(xReservingClass.Path + ' -- ' + T.Name)
        except:
            print(xReservingClass)

    T.OriginLength = period_length
    T.DevelopmentLength = period_length
    dev_count = T.DevelopmentCount(T.GetOriginDate(1))

    if pull_latest_diagonal in [0, False]:
        dic_df = {}
        for j in range_(dev_count):
            if 'Counts' in T.Name:
                # df[T.DevelopmentLabel(j)] = [int(T.ValuesByIndex(i, j)) for i in range_(dev_count-(j-1))] + ['' for i in range(j-1)]
                dic_df[T.DevelopmentLabel(j)] = [int(T.ValuesByIndex(i, j)) for i in range_(dev_count-(j-1))] + ['' for k in range(j-1)]    
            else:
                dic_df[T.DevelopmentLabel(j)] = [round(T.ValuesByIndex(i, j), decimal) for i in range_(dev_count-(j-1))] + ['' for k in range(j-1)]  
        
        df = pd.DataFrame(
            dic_df, 
            index = [T.OriginLabel(i) for i in range_(dev_count)])
        
        return df

    elif pull_latest_diagonal in [1, True]:
        latest = []
        for i in range_(dev_count):
            if 'Counts' in T.Name:
                latest.append(
                    int(T.ValuesByIndex(i, T.DevelopmentCount(T.GetOriginDate(i)))))
            else:
                latest.append(
                    round(T.ValuesByIndex(i, T.DevelopmentCount(T.GetOriginDate(i))), decimal))

        df = pd.DataFrame(
            {'Latest':latest}, 
            index = [T.OriginLabel(i) for i in range_(dev_count)])
        
        return df


def export_triangle_to_xl(
        df, 
        wb, 
        ws = 'active worksheet',
        start_row: int = 4, 
        start_col: int = 1, 
        data_type = 'triangle',
        reserving_class = None,
        triangle_name = None
    ):
    
    from openpyxl.styles.borders import Border, Side
    from openpyxl.styles import Font, Alignment
    
    if ws == 'active worksheet':
        ws = wb.active

    if data_type == 'triangle':
        # ws.cell(start_row, start_col).value = 'AY'
        if triangle_name is not None:
            cl = ws.cell(start_row - 2, start_col)
            cl.value = triangle_name
            cl.font = Font(bold=True)

        if reserving_class is not None:
            cl = ws.cell(start_row - 3, start_col)
            cl.value = reserving_class

        for i in range(df.shape[0]): # AY labels
            cl = ws.cell(start_row + i + 1, start_col)
            cl.value = df.index[i]
            cl.font = Font(bold=True)
            cl.border = Border(right=Side(style='thin'))
            cl.alignment = Alignment(horizontal='center')

        for j in range(df.shape[1]): # dev labels
            cl = ws.cell(start_row, start_col + j + 1)
            cl.value = df.columns[j]
            cl.font = Font(bold=True)
            cl.border = Border(bottom=Side(style='thin'))
            cl.alignment = Alignment(horizontal='right')

        for i in range(df.shape[0]): # values
            for j in range(df.shape[1]):
                cl = ws.cell(start_row + 1 + i, start_col + 1 + j)
                cl.value = df.iloc[i, j]
                if cl.value == '-':
                    cl.alignment = Alignment(horizontal='center')
                else:
                    cl.alignment = Alignment(horizontal='right')

    elif data_type == 'latest':

        cl = ws.cell(start_row, start_col)
        cl.value = 'Latest'
        cl.font = Font(bold=True)
        cl.border = Border(bottom=Side(style='thin'))
        cl.alignment = Alignment(horizontal='center')

        for i in range(df.shape[0]): # AY labels
            cl = ws.cell(start_row + i + 1, start_col)
            cl.value = df['Latest'][i]
            cl.alignment = Alignment(horizontal='right')


# Reserve Review Methods Collection

def wait_user_editing():
    global stop_preview
    while True:
        time.sleep(0.3)
        try:
            if Excel['Application'].ActiveWorkbook is None:
                # print("WB closed")
                stop_preview = True
                break
            elif 'DFM' not in Excel['Application'].ActiveWorkbook.Name:
                # print("WB closed")
                stop_preview = True
                break
            if [999,999] != [Excel['Application'].ActiveCell.Row, Excel['Application'].ActiveCell.Column]:
                break
        except:
            pass


def _preview_inner(xMethod, editable=False):
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Color

    global stop_preview
    
    org_count = xMethod.OriginCount
    dev_count = xMethod.DevelopmentCount(1)

    index_off = 3

    if org_count == 10: # AY
        notes_index = [23 + index_off + org_count, index_off]
        tmpl_file = "E:\\ResQ\\Automations\\library\\ToolBoxResource\\AY_DFM_macro.xlsm"   
    else:  # AQ
        notes_index = [23 + index_off + org_count, index_off]
        tmpl_file = "E:\\ResQ\\Automations\\library\\ToolBoxResource\\AQ_DFM_macro.xlsm"

    object_name = xMethod.Name.replace('/', "-")
    current_time = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    out_file = r"E:\ResQ\Automations\library\tmp" + '\\' + f"[{xMethod.Project.Name}] {object_name} ({current_time}).xlsm"


    def get_ratios(xMethod, t='ratio'):
        df = pd.DataFrame()
        df_1 = pd.DataFrame()
        df['AY'] = [xMethod.OriginLabel(i) for i in range_(xMethod.OriginCount)]
        df_1['AY'] = [xMethod.OriginLabel(i) for i in range_(xMethod.OriginCount)]

        dev_count = xMethod.DevelopmentCount(1)

        for j in range_(org_count):
            df[xMethod.DevelopmentLabel(j)] = [xMethod.Ratios(i,j) for i in range_(1,org_count)]
            df_1[xMethod.DevelopmentLabel(j)] = [xMethod.ExcludedRatios(i,j) for i in range_(1,org_count)]

        if org_count == 10:
            k = 2
            for i in range(1, org_count):
                for j in range_(1, k):
                    df.iloc[i, -j] = None
                    df_1.iloc[i, -j] = None
                k += 1

        elif org_count == 40:
            n = xMethod.Project.Name
            if 'Jan' in n or 'Feb' in n or 'Mar' in n: qtr = 1
            elif 'Apr' in n or 'May' in n or 'Jun' in n: qtr = 2
            elif 'Jul' in n or 'Aug' in n or 'Sep' in n: qtr = 3
            else: qtr = 4

            k = 5-int(qtr)
            for i in range(0, org_count):
                for j in range_(1, k): 
                    df.iloc[i, -j] = None
                    df_1.iloc[i, -j] = None
                    if j == org_count: break
                k += 1

        avg_formula_count = 0
        for i in range_(1, 20):
            try:
                xMethod.AverageRatioValues(1, i)
                avg_formula_count = i
            except:
                break

        # Avg Ratios        
        df_2 = pd.DataFrame()
        df_2['avg func'] = [xMethod.AverageFormula(i) for i in range_(1, avg_formula_count)]

        for j in range_(dev_count):
            try:
                df_2[f'dev {j}'] = [xMethod.AverageRatioValues(j, i) for i in range_(1, avg_formula_count)]
            except:
                print('error in .AverageRatioValues')

        # SelectedRatios
        df_3 = pd.DataFrame() 
        for j in range_(dev_count):
            df_3[f'dev {j}'] = [xMethod.SelectedRatios(j)]

        if t == 'ratio':
            return df
        if t == 'ex':
            return df_1
        if t == 'avg':
            return df_2
        if t == 'seleted ratio':
            return df_3

    def load2ws(xMethod, wb_name):
            
        from openpyxl.styles import Font, Color
        df_ratios = get_ratios(xMethod, 'ratio')
        df_ex = get_ratios(xMethod, 'ex')
        df_avg = get_ratios(xMethod, 'avg')
        df_select = get_ratios(xMethod, 'seleted ratio')

        WB = load_workbook(wb_name, data_only=False, keep_vba=True)
        ws = WB['Ratios & Average Selection']

        # Part 0 : reserving class path
        ws.cell(1, 3).value = xMethod.ReservingClass.Path

        # Part 1 : ratios
        for j in range(1+index_off, index_off + dev_count): # dev labels
            ws.cell(index_off, j).value = df_ratios.columns[j-index_off]

        for i in range(1+index_off, 1+index_off + dev_count): # org labels
            try: ws.cell(i, index_off).value = df_ratios.iloc[i-index_off-1,0]
            except: pass

        for i in range(1+index_off, index_off + dev_count):
            for j in range(1+index_off, index_off + dev_count):
                ws.cell(i,j).value = df_ratios.iloc[i-index_off-1,j-index_off]
                if df_ex.iloc[i-index_off-1,j-index_off] == 1: # format excluded values
                    # ws.cell(i,j).font = ws.cell(i,j).font.copy(strike=True, color='bd45c0')
                    ws.cell(i,j).font = Font(strike=True, color='bd45c0')
        
        # Part 2 : Average values
        for i in range(index_off+xMethod.OriginCount+3, index_off+xMethod.OriginCount+3+13):
            for j in range_(index_off, index_off+xMethod.OriginCount):
                try:
                    ws.cell(i, j).value = df_avg.iloc[i-3-index_off-org_count, j-index_off]
                except:
                    ws.cell(i, j).value = '-'

        
        # Part 3 : Selected Ratios
        for j in range(index_off+1, org_count+index_off+1):
            try:  
                ws.cell(org_count+17+index_off, j).value = df_select.iloc[0,j-4]
            except:
                ws.cell(org_count+17+index_off, j).value = 1


        # Part 4: Notes
        WB['Ratios & Average Selection'][f'{int2col(index_off)}{org_count+23+index_off}'].value = re.sub(r'(?<!\r)\n', '', xMethod.Notes)

        # Part 5: Ultimates
        Ult = [xMethod.Ultimates(i) for i in range_(dev_count)]
        for i in range(1+index_off, 1+index_off+dev_count): 
            cell = WB['Ratios & Average Selection'][f'{int2col(index_off+org_count+2)}{i}']
            cell.value = Ult[i-index_off-1]
            cell.font = Font(color='f2f2f2')  # grey
            cell.fill = PatternFill(start_color='a6a6a6', end_color='a6a6a6', fill_type='solid')

        ## find ratios
        try:
            try:
                D = xMethod.ReservingClass.GetVector(xMethod.SummaryRatioBasis.Name)
                D.PeriodLength = xMethod.OriginLength
                v_D = [D.ValuesByIndex(i) for i in range_(dev_count)]
            except:
                D = xMethod.ReservingClass.GetDFMMethod(xMethod.SummaryRatioBasis.Name)
                v_D = [D.Ultimates(i) for i in range_(dev_count)]

            for i in range(1+index_off, 1+index_off+dev_count):
                if org_count < 20:
                    cell = WB['Ratios & Average Selection'][f'{int2col(index_off+org_count+2+12)}{i}']
                else:
                    cell = WB['Ratios & Average Selection'][f'{int2col(index_off+org_count+2+12+5)}{i}']
                cell.value = v_D[i-index_off-1]
                cell.font = Font(color='f2f2f2')  # grey
                cell.fill = PatternFill(start_color='a6a6a6', end_color='a6a6a6', fill_type='solid')
        except:
            pass

        # Save Workbook
        try:
            WB.save(out_file)
        except:
            Excel['Application'].Visible = True
        #import os
        #os.startfile(r"Z:\Actuarial\Automation Tools\ResQ\tmp\AY_DFM_001.xlsm")
        Excel['Application'].Visible = True
        # Excel['Application'].DisplayFormulaBar = True
        Excel['Application'].DisplayAlerts = False
        Excel['Application'].Workbooks.Open(out_file)

        u.active_window(os.path.basename(out_file) + ' - Excel')
    

    def edit_method(xMethod):  
        global stop_preview
        WB = Excel['Application'].ActiveWorkbook
        ws = Excel['Application'].ActiveSheet
        last_saved = str(WB.BuiltinDocumentProperties("Last Save Time"))

        n = xMethod.Project.Name
        if 'Jan' in n or 'Feb' in n or 'Mar' in n: qtr = 1
        elif 'Apr' in n or 'May' in n or 'Jun' in n: qtr = 2
        elif 'Jul' in n or 'Aug' in n or 'Sep' in n: qtr = 3
        else: qtr = 4

        
        if org_count == 10:
            r=range(index_off + 1,index_off + 1 + org_count-1)
            tri_range = [[i, j] for i, j in itertools.product(r,r) if (i+j)< 4 +index_off+org_count]
        else: # 40
            r = range(index_off + 1, index_off + 1 + org_count + qtr - 5)
            tri_range = [[i, j] for i, j in itertools.product(r,r) if (i+j)<=38+1+index_off+qtr]

        check_point = 0
        pending_update = 0

        def COM2ws_ratio_selections():
            ### COM to WS
            # P1: updating rarios selections
            ws.Cells(notes_index[0], notes_index[1]).Formula = xMethod.Notes
            # Blue
            # ws.Range(f"{int2col(cur_col)}{org_count+3+index_off}:{int2col(cur_col)}{org_count+3+index_off+8}").Interior.Color = 15898436  
            for i in range_(org_count+3+index_off, org_count+3+index_off+11):
                ws.Cells(i, cur_col).Value = xMethod.AverageRatioValues(cur_col-index_off, i-index_off-2-org_count)
            # white - finish 
            # ws.Range(f"{int2col(cur_col)}{org_count+3+index_off}:{int2col(cur_col)}{org_count+3+index_off+8}").Interior.Color = 16777215 

            ws.Cells(org_count+17+index_off, cur_col).Value = xMethod.SelectedRatios(cur_col-index_off)

        def COM2ws_ult(): 
            # P2: update ult col in ws
            start_index = 4+index_off-cur_col+dev_count-1
            if org_count == 10 : start_index+=1
            
            # org_color = ws.Cells(1,1).Interior.Color
            # Blue
            # ws.Range(f"{int2col(index_off + org_count + 2)}{start_index}:{int2col(index_off + org_count + 2)}{index_off+dev_count}").Interior.Color = 15898436 
             
            for i in range_(start_index, 1+index_off+dev_count-1): 
                ws.Cells(i, index_off + org_count + 2).Value = xMethod.Ultimates(i-index_off)
    
            # Org color finish 
            # ws.Range(f"{int2col(index_off + org_count + 2)}{start_index}:{int2col(index_off + org_count + 2)}{index_off+dev_count}").Interior.Color = org_color 
        
        def COM2ws_notes(): 
            # P3: Notes
            ws.Cells(notes_index[0],notes_index[1]).Formula = xMethod.Notes

        while True and not stop_preview:
            try:
                check_point += 1
                if check_point > 50:
                    if str(WB.BuiltinDocumentProperties("Last Save Time")) != last_saved:
                        # WB.Save()
                        last_saved = str(WB.BuiltinDocumentProperties("Last Save Time"))
                        print(f"workbook saved! - {last_saved}")

                    if Excel['Application'].ActiveWorkbook is None:
                        break
                    elif 'DFM' not in Excel['Application'].ActiveWorkbook.Name:
                        break
                    check_point = 0

                time.sleep(0.05)
                
                cur_row = Excel['Application'].ActiveCell.Row
                cur_col = Excel['Application'].ActiveCell.Column


                if [cur_row, cur_col] == [1, 1]:
                    continue
                ## F1:
                if [cur_row, cur_col] in tri_range: # make changes to xMethod!
                    ws.Cells(1,1).Activate()

                    if ws.Cells(cur_row, cur_col).Font.Strikethrough == True:
                        ws.Cells(cur_row, cur_col).Font.Strikethrough = False
                        ws.Cells(cur_row, cur_col).Font.Color = -15066598
                        xMethod.SetExcludedRatios(cur_row-index_off, cur_col-index_off, 0)
                        if  xMethod.Notes != '':
                            xMethod.Notes += '\r\n'
                        xMethod.Notes += f"Include LDF for {xMethod.OriginLabel(cur_row-index_off)} at development period {xMethod.DevelopmentLabel(cur_col-index_off)};\r\n"
                
                    else:
                        ws.Cells(cur_row, cur_col).Font.Strikethrough = True
                        ws.Cells(cur_row, cur_col).Font.Color = -2991679 # Pink
                        xMethod.SetExcludedRatios(cur_row-index_off, cur_col-index_off, 1)
                        if  xMethod.Notes != '':
                            xMethod.Notes += '\r\n'
                        xMethod.Notes += f"Exclude LDF for {xMethod.OriginLabel(cur_row-index_off)} at development period {xMethod.DevelopmentLabel(cur_col-index_off)};\r\n"

                    COM2ws_ratio_selections()
                    COM2ws_ult()
                    COM2ws_notes()

                ## F2:
                if cur_row in range(index_off + org_count + 3, index_off + org_count + 3 + 9) and\
                   cur_col in range(index_off + 1 , index_off + 1 + org_count):

                    ws.Cells(1,1).Activate()
                    
                    if ws.Cells(cur_row, cur_col).Interior.Color != 8905428: # if not Green (selected) make change to ratio selection
                        ws.Cells(index_off+org_count+17, cur_col).Formula = cur_row - 2 - index_off - org_count  # change selected ratio index
                        xMethod.SetSelectedRatios(cur_col - index_off, cur_row - 2 - index_off - org_count)

                    COM2ws_ult()
                    COM2ws_notes()

                ## F3    
                # use input ratio area:
                if cur_row in range(index_off + org_count + 3 + 9, index_off + org_count + 3 + 9 + 3) and\
                   cur_col in range(index_off + 1 , index_off + 1 + org_count):
                    init_value = ws.Cells(cur_row, cur_col).Value

                    pending_update = {'item id':'user ratio',
                                      'dev index': cur_col - index_off,
                                      'avg index': cur_row - 2 - index_off - org_count, 
                                      'value': ws.Cells(cur_row, cur_col).Value}

                    ws.Cells(1,1).Activate()

                    if ws.Cells(cur_row, cur_col).Interior.Color != 8905428: # if not Green (selected) make change to ratio selection
                        ws.Cells(index_off+org_count+17, cur_col).Formula = cur_row - 2 - index_off - org_count  # change selected ratio index
                        xMethod.SetSelectedRatios(cur_col - index_off, cur_row - 2 - index_off - org_count)

                    COM2ws_ult()
                    COM2ws_notes()

                if pending_update != 0:
                   xMethod.SetUserRatios(pending_update['dev index'],pending_update['avg index'],pending_update['value']); pending_update = 0
                   COM2ws_ult()
                   COM2ws_notes()

            except (AttributeError, pythoncom.com_error) as e:
                # print('User is editing Cell.-edit_method')
                # print(e)
                wait_user_editing()
                xMethod.Notes = ws.Cells(notes_index[0],notes_index[1]).Formula.replace('\n', '\r\n')

    load2ws(xMethod, tmpl_file)

    if editable is True or editable == 1:
        edit_method(xMethod)
        print("Finish preview, no change has been made to ResQ database.")

        #sync_data()


def _preview(xMethod, editable=False):
    global stop_preview
    stop_preview=False

    org_count = xMethod.OriginCount
    notes_index = [25 + org_count,2]
    
    if editable is False or editable == 0:
        _preview_inner(xMethod, editable)

    elif editable is True or editable == 1:
        try:
            _preview_inner(xMethod, editable)
        except:
            pass


def _ex_diagonal(xMethod, dev_index, start_row="min_row", end_row="max_row", reason='', add_notes=True):  # Exclude Diagonals for Diminished Value Claims
    # add blank space

    label2index = {xMethod.OriginLabel(i): i for i in range_(xMethod.OriginCount)}

    if start_row == "min_row":
        start_row = 1
    elif type(start_row) == str: 
        start_row = label2index[start_row]

    if end_row == "max_row":
        end_row = xMethod.OriginCount
    elif type(end_row) == str: 
        end_row = label2index[end_row]

    if xMethod.OriginCount == 10: 
        dev_count = 9
        if start_row > 2000: start_row = start_row - (new.year - 10)
        if end_row > 2000: end_row = end_row - (new.year - 10)
    elif xMethod.OriginCount == 40: 
        dev_count = 37

    if dev_index == 1: num_count = '1st' 
    elif dev_index == 2: num_count = '2nd'
    elif dev_index == 3: num_count = '3rd' 
    else: num_count = str(dev_index) + 'th'

    col_offset = 0
    if start_row != 1:
        col_offset = 1 - start_row
    
    for row in range_(start_row, end_row):
        xMethod.SetExcludedRatios(row, dev_count + 1 - dev_index + col_offset, 1); dev_index += 1
    
    if add_notes in [True, 1]:
        
        if xMethod.Notes != '': xMethod.Notes += '\r\n' 

        if start_row == 1 and end_row == xMethod.OriginCount:
            xMethod.Notes += f"Excluded LDFs on the {num_count} to last diagonal{reason};"
        else:
            xMethod.Notes += f"Excluded {xMethod.OriginLabel(start_row)}-{xMethod.OriginLabel(end_row)} LDFs on the {num_count} to last diagonal{reason};"
    


def _ex_row(xMethod, row, add_notes=True):
    if type(row) == list:
        for i in row:
            for j in range(xMethod.OriginCount):
                xMethod.SetExcludedRatios(i, j, 1)
    else:
        [xMethod.SetExcludedRatios(row, dev, 1) for dev in range(xMethod.OriginCount)]
    
    if add_notes in [True, 1]:
        if xMethod.Notes != '': xMethod.Notes += '\r\n'    
        xMethod.Notes += f"Exclude LDFs for {xMethod.OriginLabel(row)};\r\n"


def _ex_AY(xMethod, AY, reason=''):
    AY_list = AY
    AY_index = []
    lb_list = []; lb = ''
    if type(AY) == list:
        for k in range(len(AY_list)):
            for i in range_(xMethod.OriginCount):
                if str(AY_list[k]) in xMethod.OriginLabel(i):
                    AY_index.append(i)         
                    if str(AY_list[k]) not in lb_list:
                        lb_list.append(str(AY_list[k]))
                        lb += f' {str(AY_list[k])[0:4]};'

        if xMethod.Notes != '': xMethod.Notes += '\r\n'
        xMethod.Notes += f"Excluded LDFs for AY{lb}\r\n"

    elif type(AY) == int:
        for i in range_(xMethod.OriginCount):
            if str(AY) in xMethod.OriginLabel(i):
                AY_index.append(i)

        if xMethod.Notes != '': xMethod.Notes += '\r\n'
        xMethod.Notes += f"Excluded LDFs for AY {AY}{reason};\r\n"

    for i in AY_index:
        for j in range_(xMethod.DevelopmentCount(i)):
            xMethod.SetExcludedRatios(i, j, 1)


def _ex_COVID_AY(xMethod, AY=[2020,2021], add_notes=True):
    AY_list = AY
    AY_index = []
    if type(AY) == list:
        for k in range(len(AY_list)):
            for i in range_(xMethod.OriginCount):
                if str(AY_list[k]) in xMethod.OriginLabel(i):
                    AY_index.append(i)
    elif type(AY) == int:
        for i in range_(xMethod.OriginCount):
            if str(AY) in xMethod.OriginLabel(i):
                AY_index.append(i)

    for i in AY_index:
        for j in range_(xMethod.DevelopmentCount(i)):
            xMethod.SetExcludedRatios(i, j, 1)

    if add_notes in [True, 1]:
        if xMethod.Notes != '': xMethod.Notes += '\r\n'
        xMethod.Notes += "Excluded 2020, 2021 LDFs since they are distorted by COVID."


def _ex_hi(xMethod, dev_period, count=1, reason='', add_notes=True):

    org_counts = xMethod.OriginCount
    ratios = [xMethod.Ratios(AY, dev_period) for AY in range_(org_counts) if xMethod.ExcludedRatios(AY, dev_period)==0]
    high_values = sorted(ratios, reverse=True)[:count]

    for org_row in range_(org_counts - dev_period):
        if xMethod.Ratios(org_row, dev_period) in high_values:
            xMethod.SetExcludedRatios(org_row, dev_period, 1)


    if add_notes in [True, 1]:
        if 'due to' in reason and ' due to' not in reason:
            reason = reason.replace('due to', ' due to')

        if xMethod.Notes != '': xMethod.Notes += '\r\n\r\n'

        xMethod.Notes += f"Select low LDF for development period {xMethod.DevelopmentLabel(dev_period)}{reason};"

        if "for ..." in reason:
            aDFM = DFM(xMethod.Name)
            total_dev = xMethod.DevelopmentCount(1)
            xMethod.Notes = xMethod.Notes.replace('for ...', f'for {aDFM.COMObject.OriginLabel(total_dev+1-dev_period)} at {aDFM.dev_month(dev_period)} months')


def _ex_lo(xMethod, dev_period=1, count=1, reason='', add_notes=True):

    org_counts = xMethod.OriginCount
    ratios = [xMethod.Ratios(AY, dev_period) for AY in range_(org_counts) if xMethod.ExcludedRatios(AY, dev_period)==0]
    low_values = sorted(ratios, reverse=False)[:count]

    for org_row in range_(org_counts - dev_period):
        if xMethod.Ratios(org_row, dev_period) in low_values:
            xMethod.SetExcludedRatios(org_row, dev_period, 1)

    if add_notes in [True, 1]:
        if 'due to' in reason and ' due to' not in reason:
            reason = reason.replace('due to', ' due to')

        if xMethod.Notes != '': xMethod.Notes += '\r\n\r\n'

        xMethod.Notes += f"Select high LDF for development period {xMethod.DevelopmentLabel(dev_period)}{reason};"

        if "for ..." in reason:
            aDFM = DFM(xMethod.Name)
            total_dev = xMethod.DevelopmentCount(1)
            xMethod.Notes = xMethod.Notes.replace('for ...', f'for {aDFM.COMObject.OriginLabel(total_dev+1-dev_period)} at {aDFM.dev_month(dev_period)} months')


def _select_high(xMethod, dev_period=1, count=1, reason='', add_notes=True):

    org_counts = xMethod.OriginCount
    ratios = [xMethod.Ratios(AY, dev_period) for AY in range_(org_counts) if xMethod.ExcludedRatios(AY, dev_period)==0]
    high_values = sorted(ratios, reverse=True)[:count]

    for org_row in range_(org_counts - dev_period):
        if xMethod.Ratios(org_row, dev_period) not in high_values:
            xMethod.SetExcludedRatios(org_row, dev_period, 1)

    if add_notes in [True, 1]:
        if 'due to' in reason and ' due to' not in reason:
            reason = reason.replace('due to', ' due to')

        if xMethod.Notes != '': xMethod.Notes += '\r\n\r\n'

        xMethod.Notes += f"Select high LDF for development period {xMethod.DevelopmentLabel(dev_period)}{reason};"

        if "for ..." in reason:
            aDFM = DFM(xMethod.Name)
            total_dev = xMethod.DevelopmentCount(1)
            xMethod.Notes = xMethod.Notes.replace('for ...', f'for {aDFM.COMObject.OriginLabel(total_dev+1-dev_period)} at {aDFM.dev_month(dev_period)} months')


def _select_low(xMethod, dev_period=1, count=1, reason='', add_notes=True):

    org_counts = xMethod.OriginCount
    ratios = [xMethod.Ratios(AY, dev_period) for AY in range_(org_counts) if xMethod.ExcludedRatios(AY, dev_period)==0]
    low_values = sorted(ratios, reverse=False)[:count]

    for org_row in range_(org_counts - dev_period):
        if xMethod.Ratios(org_row, dev_period) not in low_values:
            xMethod.SetExcludedRatios(org_row, dev_period, 1)

    if add_notes in [True, 1]:
        if 'due to' in reason and ' due to' not in reason:
            reason = reason.replace('due to', ' due to')

        if xMethod.Notes != '': xMethod.Notes += '\r\n\r\n'
        
        xMethod.Notes += f"Select low LDF for development period {xMethod.DevelopmentLabel(dev_period)}{reason};"
    
        if "for ..." in reason:
            aDFM = DFM(xMethod.Name)
            total_dev = xMethod.DevelopmentCount(1)
            xMethod.Notes = xMethod.Notes.replace('for ...', f'for {aDFM.COMObject.OriginLabel(total_dev+1-dev_period)} at {aDFM.dev_month(dev_period)} months')


def _ex_LDF(xMethod, dev, row,  reason='', add_notes=True):

    label2index = {xMethod.OriginLabel(i): i for i in range_(xMethod.OriginCount)}

    if type(row)==int:
        row = str(row)
    row = label2index[row]

    xMethod.SetExcludedRatios(row, dev, 1)

    if add_notes in [True, 1]:
        if xMethod.Notes != '': xMethod.Notes += '\r\n\r\n'
        xMethod.Notes += f"Exclude LDF for {xMethod.OriginLabel(row)} at development period {xMethod.DevelopmentLabel(dev)}{reason};"


def _view_result_selection(xResult, view_in_excel=False):

    org_count = xResult.OriginCount

    df = pd.DataFrame([xResult.OriginLabel(i) for i in range_(org_count)], 
                      index = list(range_(1, org_count)), 
                      columns = ['AY'])

    # method_list = [xResult.Dataset(i).Name for i in range_(15) if len(xResult.Dataset(i).Name)>1 and 'Total' not in xResult.Dataset(i).Name]
    method_list = [xResult.Dataset(i).Name for i in range_(xResult.DatasetCount)]
    method_index = 0

    for method in method_list:
        method_index += 1

        df[method] = np.nan
        df[f"Weight ({method_index})"] = np.nan

        for row in range(1, org_count+1):
            value = "%.0f" % xResult.DatasetValues(DatasetIndex=method_index, 
                                                   OriginIndex=row, 
                                                   OriginLength=xResult.OriginLength)
            
            weight = xResult.Weights(DatasetIndex=method_index, 
                                     OriginIndex=row)

            df.loc[row, method] = value
            df.loc[row, f"Weight ({method_index})"] = weight

    df["Selected Ultimate"] = [xResult.Ultimates(i, xResult.OriginLength) for i in range_(org_count)]
            

    # return tables
    if view_in_excel in [False, 0]:
        return df

    elif view_in_excel in [True, 1]:
        from openpyxl import load_workbook
        from win32com.client import Dispatch
        import random
        if org_count == 10:
            source_file = "E:\\ResQ\\Automations\\library\\ToolBoxResource\\AY_ResultSelection.xlsm"
        elif org_count == 40:
            source_file = "E:\\ResQ\\Automations\\library\\ToolBoxResource\\AQ_ResultSelection.xlsm"

        WB = load_workbook(source_file, data_only=False, keep_vba=True)
        ws = WB['Method']
         
        for i in range_(org_count):
            ws.cell(i + 2, 2).value = df['AY'][i]

        for j in range(3, 3 + len(method_list)*2):
            for i in range_(3-1, 3-1 + org_count):
                try:
                    if i == 2:
                        ws.cell(i, j).value = df.columns[j-2]
                    else:
                        ws.cell(i, j).value = float(df.iloc[i-3, j-2])
                except:
                    pass

        for i in range(3, 3 + org_count): 
            ws[f'AF{i}'].value = df.loc[i-2, "Selected Ultimate"]

        for Col in range_(5, col('AF')-2):
            if ws.cell(3, Col).value == None:
                ws.column_dimensions[col(Col)].hidden= True

        current_time = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
        new_file = f"E:\\ResQ\\Automations\\library\\tmp\\[{xResult.Project.Name}] -- {xResult.Name} ({current_time}).xlsm"

        WB.save(new_file)

        Excel['Application'] = Dispatch('Excel.Application')
        Excel['Application'].Visible = 1
        # Excel['Application'].DisplayFormulaBar = 1
        Excel['Application'].DisplayAlerts = 0
        WB = Excel['Application'].Workbooks.Open(new_file)

        u.active_window(os.path.basename(new_file) + ' - Excel')


def _plot_triangle(xTriangle, dev_count, org_count, transpose=False):
    
    import matplotlib.pyplot as plt
    from matplotlib.pyplot import figure
    import warnings
    warnings.filterwarnings("ignore")

    figure(figsize=(12, 6), dpi=92)
    plt.style.use('default')
    
    DT = xTriangle

    if org_count == 10: 
        DT.OriginLength = DT.DevelopmentLength = 12
    else: 
        DT.OriginLength = DT.DevelopmentLength = 3

    if type(dev_count) == range:
        if dev_count[0] == 0:
            dev_count = range_(list(dev_count)[-1]+1)
    if type(dev_count) == int:
        dev_count = [dev_count]

    # plot lines
    if transpose is False:
        x = [DT.OriginLabel(i) for i in range_(org_count)]
        for dev in dev_count:
            y = []
            for i in range_(org_count):
                if DT.ValuesByIndex(i, dev)!=0: 
                    y.append(DT.ValuesByIndex(i, dev))
                else: 
                    y.append(None)
                
            plt.plot(x, y, '--s', label = DT.DevelopmentLabel(dev))

        current_values = plt.gca().get_yticks()

        if abs(y[-5]) < 1:
            plt.gca().set_yticklabels([f'{y:.1%}' for y in current_values])
            plt.ylabel("Ratios")
        else:
            plt.gca().set_yticklabels([f'{y:,.0f}' for y in current_values])
            plt.ylabel("Values")

        plt.grid(True, ls='--', color='gainsboro')

        if org_count == 10:
            plt.xlabel("Accident Years")
        else:
            plt.xlabel("Accident Quarters")
            plt.xticks(rotation=90)

    elif transpose is True:
        x = []
        for i in range_(org_count):
            try: 
                x.append(DT.DevelopmentLabel(i))
            except: 
                x.append(None)

        for dev in dev_count:
            y = []
            for i in range_(org_count):
                if DT.ValuesByIndex(i,dev)!=0: 
                    y.append(DT.ValuesByIndex(i, dev))
                else: 
                    y.append(None)
                
            plt.plot(x, y, '--s', label = DT.DevelopmentLabel(dev))

        current_values = plt.gca().get_yticks()

        if current_values[1] <= 1 and current_values[1] >= -1:
            plt.gca().set_yticklabels([f'{y:.1%}' for y in current_values])
        else:
            plt.gca().set_yticklabels([f'{y:,.0f}' for y in current_values])

        plt.grid(True, ls='--', color='gainsboro')

        if org_count == 10:
            plt.xlabel("Development Years")
        else:
            plt.xlabel("Development Quarters")
            plt.xticks(rotation=90)

        plt.ylabel("Counts/Ratios")     

    plt.legend()
    plt.title(DT.Name)
    plt.show()


def _plot_ultimates(xMethod, show_ratios=False , D=None):
    import matplotlib.pyplot as plt
    from matplotlib.pyplot import figure
    import seaborn as sns
    import warnings
    warnings.filterwarnings("ignore")
    #plt.style.use(['fivethirtyeight','seaborn-deep'])
    # sns.set(style="ticks", context="talk")
    #sns.set_theme(style="darkgrid", palette="pastel")
    sns.set(rc={'figure.figsize':(10.85, 4.8)})
    
    DevelopmentCount = xMethod.DevelopmentCount(1)
    v_Ult = [xMethod.Ultimates(i) for i in range_(DevelopmentCount)]

    # Set labels
    if D is None and show_ratios is False: # show counts
        y = np.array(v_Ult)
        ylabel = "Ultimates"
            
    else: # show as ratios
        if D is None:
            try:
                D = xMethod.ReservingClass.GetVector(xMethod.SummaryRatioBasis.Name)
                D.PeriodLength = xMethod.OriginLength
            except:
                D = xMethod.ReservingClass.GetDFMMethod(xMethod.SummaryRatioBasis.Name)
            

        if type(D) != list:
            #D.PeriodLength = 3*
            try:
                v_D = [D.ValuesByIndex(i) for i in range_(DevelopmentCount)]
            except:
                v_D = [D.Ultimates(i) for i in range_(DevelopmentCount)]
        else:
            v_D = D

        y = np.array([v_Ult[i]/v_D[i] for i in range(len(v_Ult))])
        ylabel = "Ultimates (ratio)"


    if DevelopmentCount == 10:
        xlabel = "Accident Year"
    else:
        xlabel = "Accident Quarter"
        
    # Set data
    v_Ult = [xMethod.Ultimates(i) for i in range_(DevelopmentCount)]

    data = pd.DataFrame({"org_label" : [xMethod.OriginLabel(i) for i in range_(DevelopmentCount)],
                         "x-index" : list(range_(1, DevelopmentCount)),
                         "y-value" : y})

    # plot module
    my_ply = sns.regplot(x="x-index", y="y-value", data = data,
                         scatter_kws={"label": "total_bill"},
                         line_kws={"color": "green"}
                         )

    my_ply.set_xticks(range(1,DevelopmentCount+1,1))
    my_ply.set_xticklabels(data['org_label'], fontsize=10)

    my_ply.set(xlabel = xlabel, ylabel = ylabel, title = f'Projected Utimates - {xMethod.Name}')
    #plt.style.use('dark_background')

    if D == None:
        current_values = plt.gca().get_yticks()
        if current_values[1] <= 1 and current_values[1] >= -1:
            plt.gca().set_yticklabels([f'{y:.1%}' for y in current_values])
        else:
            plt.gca().set_yticklabels([f'{y:,.0f}' for y in current_values])
    else:
        current_values = plt.gca().get_yticks()
        plt.gca().set_yticklabels(['{:.2%}'.format(y) for y in current_values])

    if DevelopmentCount > 10:
        plt.xticks(rotation=90)

    plt.show()


def set_adjustment(input_dict) -> dict:
    """
    Keys = 'counts', 'paid loss', 'incurred loss', 'accounting cutoff', 'other'

    Values = list of adjustment values
        For example: [adjustment value for 1st development period,
                      adjustment value for 2nd development period,
                      adjustment value for 3rd development period, 
                      ...]
    """
    if adjustment == {
            'counts': [0], 'paid loss': [0], 'incurred loss': [0], 
            'accounting cutoff': [0], 'other': [0]
        }:
        for key, value in input_dict.items():
            adjustment[key] = value

        for key, value in adjustment.items():
            l = []
            for i in value:
                if type(i) == float:
                    i = round(i, 4)
                if i is None:
                    continue
                l.append(str(round(i*100, 4)) + '%')
            adjustment[key] = l



def _F_list_1(input_list, output_type='Value'):
    '''
    Calculate formula part 1 (left side)
    '''
    a1 = input_list
    a3 = [];
    a3_formula = []

    for i in range(len(a1)):
        a3.append(1+a1[i])
        if a1[i] > 0:
            a3_formula.append(f"(1+{round(a1[i]*100, 3)}%)")
        elif a1[i] < 0:
            a3_formula.append(f"(1-{round(-a1[i]*100, 3)}%)")
        else:
            a3_formula.append("1")
        j = i+1

        if i != len(a1)-1:
            if a1[j] == 0: continue
            a3[i] = a3[i]/(1+a1[j])

            if a1[j] > 0:
                a3_formula[i] = f"{a3_formula[i]}/(1+{round(a1[j]*100, 2)}%)"
            elif a1[j] < 0:
                a3_formula[i] = f"{a3_formula[i]}/(1-{round(-a1[j]*100, 2)}%)"

    if output_type == 'Value':
        return a3
    else:
        return a3_formula
    

def _F_list_2(input_list, output_type='Value'):
    '''
    Calculate formula part 2 (right side)
    '''
    a1 = input_list
    a3 = []
    a3_formula = []

    for i in range(len(a1)):
        if i != len(a1) - 1:
            a3.append( (1+a1[i])/(1+a1[i+1]) )
            a3_formula.append( f"{1+a1[i]}/{1+a1[i+1]}" )
        else:
            a3.append(1+a1[i])
            a3_formula.append( f"{round(1+a1[i], 4)}" )

    for i in a3_formula:
        if i[-2:] == '/1':
            a3_formula[a3_formula.index(i)] = i[:-2]

    if output_type == 'Value':
        return a3
    else:
        return a3_formula
        

def _remove_lines(note_string: str, keyword_list: list) -> str:
    lines = note_string.splitlines()

    filtered_lines = [
        line for line in lines 
        if not any(keyword in line for keyword in keyword_list)
    ]
    
    return "\r\n".join(filtered_lines) if "\r\n" in note_string else "\n".join(filtered_lines)


adjustment = {
    'counts': [0], 'paid loss': [0], 'incurred loss': [0], 
    'accounting cutoff': [0], 'other': [0]
}

def _final_selection(xMethod, avg_type='Default', adjustments=adjustment, other_adjustment=[0]):

    for k in adjustments.keys():
        for ct in range(len(adjustments[k])):
            if type(adjustments[k][ct]) == str:
                adjustments[k][ct] = round(float(adjustments[k][ct].replace('%', ''))/100, 4)
                if abs(adjustments[k][ct]) < 0.0001:
                    adjustments[k][ct] = 0

    
    len_list = [len(i) for i in adjustment.values()]
    max_len = 0
    for i in len_list:
        if i > max_len:
            max_len = i

    dev_period = [i for i in range_(max_len)]

    if len(other_adjustment) < len(dev_period):
        other_adjustment = other_adjustment + [0]*(len(dev_period)-len(other_adjustment))

    for i in range(len(other_adjustment)):
        if type(other_adjustment[i]) == str:
            other_adjustment[i] = float(other_adjustment[i].replace('%', ''))/100

    index2formula = {i: xMethod.AverageFormula(i) for i in range_(xMethod.RatioAverageCount)}
    formula2index = {xMethod.AverageFormula(i): i for i in range_(xMethod.RatioAverageCount)}


    if type(avg_type) == str:
        avg_type = [avg_type for i in range(len(dev_period))]

    for item in adjustment.keys():
        if len(adjustment[item]) != len(dev_period):
            zero_count = len(dev_period) - len(adjustment[item])
            adjustment[item] = adjustment[item] + [0 for _ in range(zero_count)]

    # Clear old notes
    for kwd in [
        'For development period (', 
        'Apply growth adjustments of ', 
        'Apply accounting cutoff ',
        'Selected average factor: ',
        'Selected LDF after adjustments: '
    ]:
        if kwd in xMethod.Notes:
            xMethod.Notes = _remove_lines(xMethod.Notes, [kwd])

    for order in range(len(dev_period)): 
        # order = 0 --> 1st dev period
        # order = 1 --> 2nd dev period
        if xMethod.Notes != '':
            note_list = ['\r\n\r\n']
        else: 
            note_list = []

        note_list.append(f"For development period {xMethod.DevelopmentLabel(dev_period[order])}:\r\n")
        
        # Get Avg Index & Value
        AvgIndex = 'Not Selected'

        if type(avg_type[order]) == int:
            AvgIndex = avg_type[order]
            selc_avg_factor = f'  ◦ Selected average factor: "{index2formula[avg_type[order]].split(": ")[1]}"'
            
        else: # user input == string
            if avg_type[order] == 'Default':
                AvgIndex = xMethod.SelectedRatios(dev_period[order])

            for item in index2formula.values():
                if avg_type[order] == item:
                    AvgIndex = formula2index[item]
                    break

            if AvgIndex == 'Not Selected': # If not found in exact match, search again
                for item in index2formula.values():
                    if avg_type[order] in item:
                        AvgIndex = formula2index[item]
                        break

            if AvgIndex == 'Not Selected':
                print('invalid input for key argument: avg_type, this function will use the default formula for calculations')
                # AvgIndex = 1
                AvgIndex = xMethod.SelectedRatios(dev_period[order])

            selc_avg_factor = f'  ◦ Selected average factor: "{index2formula[AvgIndex].split(": ")[1]}"'

        avg_ratio_value = xMethod.AverageRatioValues(dev_period[order], AvgIndex)

        # Set Category
        category = xMethod.OutputVector.DatasetType.Category.Name
        dataset_type = xMethod.OutputVector.DatasetType.Name

        if '52' in dataset_type:
            print("Don't apply adjustment to this method.")
            return

        elif category == 'C Claim Count': 
            
            FP1 = _F_list_1(adjustment['counts'], 'Formula')[order]
            FP2 = _F_list_2(adjustments['counts'], 'Formula')[order]
            adj = _F_list_2(adjustments['counts'], 'Value')[order]
            
        elif category == 'H Severity': 
            
            if _F_list_1(adjustment['counts'], 'Formula')[order].count('(') <= 1:
                FP1 = _F_list_1(adjustment['incurred loss'], 'Formula')[order] + '/' + _F_list_1(adjustment['counts'], 'Formula')[order]
            else:
                FP1 = _F_list_1(adjustment['incurred loss'], 'Formula')[order] + '/(' + _F_list_1(adjustment['counts'], 'Formula')[order] + ')'
                
            FP2 = f"({_F_list_2(adjustments['incurred loss'], 'Formula')[order]})/({_F_list_2(adjustments['counts'], 'Formula')[order]})"
            adj = _F_list_2(adjustments['incurred loss'], 'Value')[order] / _F_list_2(adjustments['counts'], 'Value')[order]

        elif   'Paid' in dataset_type \
            or 'Salv DFM' in dataset_type \
            or 'Subr DFM' in dataset_type:
            
            FP1 = _F_list_1(adjustment['paid loss'], 'Formula')[order]
            FP2 = _F_list_2(adjustments['paid loss'], 'Formula')[order]
            adj = _F_list_2(adjustments['paid loss'], 'Value')[order]

        elif 'Incurred' in dataset_type: 
            
            FP1 = _F_list_1(adjustment['incurred loss'], 'Formula')[order]
            FP2 = _F_list_2(adjustments['incurred loss'], 'Formula')[order]
            adj = _F_list_2(adjustments['incurred loss'], 'Value')[order]

        if FP1.count('(') == 1 and r'/' not in FP1:
            FP1 = FP1.replace('(', '').replace(')', '')

        accounting_cutoff = 1 + adjustments['accounting cutoff'][order]

        if category == 'H Severity':
            accounting_cutoff = 1

        other_adj = 1 + other_adjustment[order]

        # Calculate
        final_selected_value = avg_ratio_value * adj * accounting_cutoff * other_adj
        if final_selected_value != avg_ratio_value:
            # print(f"Selected ratio for devlopment period {dev_period[order]}:", "%.4f" % final_selected_value) 
            xMethod.SetUserRatios(dev_period[order], 10, final_selected_value)
            xMethod.SetSelectedRatios(dev_period[order], 10)

        final_formula = f"{selc_avg_factor} ({avg_ratio_value:.4f})\r\n" + f"  ◦ Selected LDF after adjustments: {avg_ratio_value:.4f}"

        if adj != 1: 
            # note_list.append(f"  • Apply growth adjustments of {FP1} = {FP2} for development period {xMethod.DevelopmentLabel(dev_period[order])};\r\n")
            note_list.append(f"  ◦ Apply growth adjustments of {FP1} = {FP2};\r\n")
            final_formula += f" * {FP2}"

        if accounting_cutoff != 1: 
            # note_list.append(f"  • Apply accounting cutoff {accounting_cutoff-1:.2%} for development period {xMethod.DevelopmentLabel(dev_period[order])};\r\n")
            note_list.append(f"  ◦ Apply accounting cutoff 1+{accounting_cutoff-1:.2%} = {round(accounting_cutoff, 4)};\r\n")
            final_formula += f" * {accounting_cutoff:.4f}"

        if other_adjustment[order] != 0:
            final_formula += ' * ' + str(round(other_adj, 3))

        final_formula += f" = {final_selected_value:.4f}"

        if adj == 1 and accounting_cutoff == 1 and other_adj == 1:
            pass
        else:
            note_list.append(final_formula)

        if len(note_list) not in [1, 2]:
            for i in note_list: 
                xMethod.Notes += i

    # print('\n')


def select_project(name=''):
    if name == '':
        search(projects)
    else:
        set_project(name)


def select_reserving_class(path=''):

    if path == '':
        for item in tq(proj_info['xProject'].ReservingClasses(), 'Loading Reserving Classes...'):
            if item.Path.count('\\') == 4:
            # if [i for i in item.ChildClasses()] == []:
                reserving_classes[item.Path] = item

        search(reserving_classes)
    else:
        set_reserving_class(path)


def jupyter_settings():
    
    import sys
    import jupyterlab
    import jupyter_core.paths
    import getpass

    user_name = getpass.getuser()

    # Python interpreter path
    python_path = sys.executable
    print("Python Interpreter Path:", python_path)
    
    # JupyterLab extension paths
    # extension_paths = jupyterlab.paths.LAB_EXTENSIONS_PATH
    # print("JupyterLab Extension Paths:")
    # for path in extension_paths:
    #     print("-", path)
    
    # Jupyter app paths
    app_paths = jupyter_core.paths.jupyter_path()
    print("\nJupyter App Paths:")
    for path in app_paths:
        print("-", path)

    print("\nLabextensions Paths:")
    print(f"C:\\Users\\{user_name}\\AppData\\Roaming\\Python\\share\\jupyter\\labextensions\\jupyterlab-execute-time\\schemas\\jupyterlab-execute-time\\settings.json")



def extended_DFM_triangle(xMethod):

    c_year = DATE().year
    dev_labels = [xMethod.DevelopmentLabel(i) for i in range_(xMethod.OriginCount)]
    index_range = range(c_year - 15, c_year)  # Range for the index
    # Create an empty DataFrame
    df1 = pd.DataFrame(index=index_range, columns=dev_labels)  # ratio
    df2 = pd.DataFrame(index=index_range, columns=dev_labels)  # excluded

    # Pull current year int from project name
    current_proj = xMethod.Project.Name

    for current_year in range(2000, 2077):
        if str(current_year) in current_proj:
            break

    # Get Ratios from current project
    for dev_label in dev_labels:
        for row in range(current_year - 9, current_year):
            ratio_value = xMethod.Ratios(row - current_year + 10, dev_labels.index(dev_label) + 1)
            excluded = xMethod.ExcludedRatios(row - current_year + 10, dev_labels.index(dev_label) + 1)
            df1.loc[row, dev_label] = round(ratio_value, 4)
            df2.loc[row, dev_label] = excluded

    # Get Ratios from old projects
    for year in range(current_year - 7, current_year):

        xProject_old = ResQ.Project(current_proj.replace(str(current_year), str(year))).COM
        if xProject_old is None: 
            continue

        xReservingClass_old = xProject_old.GetReservingClass(proj_info['Reserving Class'])
        if xReservingClass_old is None: 
            continue

        DFM_old = xReservingClass_old.GetDFMMethod(xMethod.Name)

        for dev_label in dev_labels:
            ratio_value = DFM_old.Ratios(1, dev_labels.index(dev_label) + 1)
            excluded = DFM_old.ExcludedRatios(1, dev_labels.index(dev_label) + 1)
            df1.loc[year-8, dev_label] = round(ratio_value, 4)
            df2.loc[year-8, dev_label] = excluded
            
    return (df1, df2)


def _format_notes(note: str):
    try:
        if note.index('\n') == 0:
            note = note[1:]
    except:
        pass
    if '\n' in note and '\r\n' not in note:
        note = note.replace('\n', '\r\n')
    return note
            




