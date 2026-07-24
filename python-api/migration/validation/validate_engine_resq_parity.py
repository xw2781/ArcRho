"""Validate live ResQ generated datasets against fresh ArcRho engine output.

The validation scope is deliberately fixed to the final-validation project.  The
ResQ connection defaults and reserving-class list are imported from
``resq_data_migration`` so this tool cannot drift from the migration scope.

Run with Python 3.10 from the repository root:

    py -3.10 python-api/migration/validation/validate_engine_resq_parity.py

The process needs a Windows session that can authenticate to ResQ.  It writes
per-project Markdown and Excel findings below ``validation/results``. Temporary
engine CSVs are removed on exit, and project dataset caches or sidecars are
never updated.
"""

from __future__ import annotations

import argparse
import csv
import math
import os
import subprocess
import sys
import tempfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable, Sequence


_VALIDATION_DIR = Path(__file__).resolve().parent
_MIGRATION_DIR = _VALIDATION_DIR.parent
_REPO_ROOT = _MIGRATION_DIR.parents[1]
if str(_MIGRATION_DIR) not in sys.path:
    sys.path.insert(0, str(_MIGRATION_DIR))

import resq_data_migration as migration  # noqa: E402
from resq_migration.core import (  # noqa: E402
    METHOD_TYPE_NONE_CODE,
    _dataset_cache_csv_file_name,
    _vector_cache_csv_file_name,
)
from resq_migration.engine import EngineGenerationError, generate_engine_csv  # noqa: E402
from resq_migration.extractors import export_triangle, export_vector  # noqa: E402


# TARGET_PROJECT_NAME = "NJ_Annual_Prod_202605_Fake"
TARGET_PROJECT_NAME = "NJ_Annual_Prod_2026 Q2-May"
DEFAULT_RESULT_ROOT = _VALIDATION_DIR / "results"
DEFAULT_REPORT_DIR = DEFAULT_RESULT_ROOT / TARGET_PROJECT_NAME
DEFAULT_TEMP_ROOT = _REPO_ROOT / "python-api" / "logs" / "tmp"
DEFAULT_ABSOLUTE_TOLERANCE = 1e-9
DEFAULT_RELATIVE_TOLERANCE = 1e-12
MARKDOWN_START_MARKER = "<!-- BEGIN GENERATED FINDINGS -->"
MARKDOWN_END_MARKER = "<!-- END GENERATED FINDINGS -->"
ProgressCallback = Callable[[str], None]


@dataclass
class ComparisonResult:
    """The reviewable outcome for one dataset or reserving-class failure."""

    rc_path: str
    kind: str
    resq_formula: str = ""
    dataset_name: str = ""
    dataset_type: str = ""
    status: str = "match"
    categories: tuple[str, ...] = ()
    origin_length: int | None = None
    development_length: int | None = None
    resq_shape: tuple[int, int] | None = None
    engine_shape: tuple[int, int] | None = None
    mismatch_count: int = 0
    max_absolute_delta: float | None = None
    max_relative_delta: float | None = None
    first_mismatch_cell: str = ""
    resq_value: object = None
    engine_value: object = None
    error: str = ""

    @property
    def is_issue(self) -> bool:
        return self.status != "match"


@dataclass
class ValidationSummary:
    """Aggregate run state used by both report formats."""

    started_at: datetime
    rc_paths: list[str]
    results: list[ComparisonResult] = field(default_factory=list)
    skipped_by_reason: dict[str, int] = field(default_factory=dict)
    finished_at: datetime | None = None

    @property
    def eligible_count(self) -> int:
        return len(self.results)

    @property
    def match_count(self) -> int:
        return sum(not item.is_issue for item in self.results)

    @property
    def issue_results(self) -> list[ComparisonResult]:
        return [item for item in self.results if item.is_issue]

    @property
    def issue_count(self) -> int:
        return len(self.issue_results)

    @property
    def skipped_count(self) -> int:
        return sum(self.skipped_by_reason.values())


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _safe_text(value: object) -> str:
    return str(value or "").strip()


def _emit_progress(progress_callback: ProgressCallback | None, message: str) -> None:
    """Send non-critical diagnostic progress without interrupting validation."""

    if progress_callback is None:
        return
    try:
        progress_callback(message)
    except Exception:
        pass


def _get_method_type_code(dataset: object) -> int:
    """Return a ResQ method code, treating an absent property as non-method data."""

    try:
        value = getattr(dataset, "MethodType")
    except Exception:
        return METHOD_TYPE_NONE_CODE
    try:
        return int(value)
    except (TypeError, ValueError):
        return METHOD_TYPE_NONE_CODE


def _minimal_payload(dataset: object) -> dict[str, str]:
    """Read only the fields required by the migration's generated-instance gate."""

    name = _safe_text(getattr(dataset, "Name", ""))
    try:
        dataset_type = _safe_text(getattr(getattr(dataset, "DatasetType"), "Name", ""))
    except Exception:
        dataset_type = ""
    return {"name": name, "dataset_type": dataset_type}


def _eligible_dataset(dataset: object) -> tuple[bool, str]:
    """Apply the same generated-instance gate used by the ResQ migration."""

    if _get_method_type_code(dataset) != METHOD_TYPE_NONE_CODE:
        return False, "method_output"
    try:
        eligible = migration._is_engine_generated_instance(_minimal_payload(dataset))
    except Exception:
        eligible = False
    return (True, "") if eligible else (False, "not_generated_instance")


def _is_missing(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    try:
        return math.isnan(float(value))
    except (TypeError, ValueError):
        return False


def _as_float(value: object) -> float:
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Expected a numeric dataset value, received {value!r}.") from exc


def read_engine_csv(path: Path) -> list[list[float | None]]:
    """Read the data-engine's headerless CSV while preserving blank cells."""

    rows: list[list[float | None]] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        for row_number, raw_row in enumerate(csv.reader(handle), start=1):
            row: list[float | None] = []
            for column_number, raw_value in enumerate(raw_row, start=1):
                text = raw_value.strip()
                if not text:
                    row.append(None)
                    continue
                try:
                    value = float(text)
                except ValueError as exc:
                    raise ValueError(
                        "Engine output contains a non-numeric value at "
                        f"row {row_number}, column {column_number}: {raw_value!r}."
                    ) from exc
                row.append(None if math.isnan(value) else value)
            rows.append(row)
    if not rows:
        raise ValueError("Engine output CSV was empty.")
    return rows


def _matrix_shape(matrix: Sequence[Sequence[object]]) -> tuple[int, int]:
    return len(matrix), max((len(row) for row in matrix), default=0)


def _matrix_value(matrix: Sequence[Sequence[object]], row: int, column: int) -> object:
    if row >= len(matrix) or column >= len(matrix[row]):
        return None
    return matrix[row][column]


def _format_cell(row: int, column: int) -> str:
    return f"origin {row + 1}, development {column + 1}"


def compare_matrices(
    resq_values: Sequence[Sequence[object]],
    engine_values: Sequence[Sequence[object]],
    *,
    absolute_tolerance: float = DEFAULT_ABSOLUTE_TOLERANCE,
    relative_tolerance: float = DEFAULT_RELATIVE_TOLERANCE,
) -> dict[str, object]:
    """Compare shape, missingness, and numeric cells without coercing blanks to zero."""

    resq_shape = _matrix_shape(resq_values)
    engine_shape = _matrix_shape(engine_values)
    categories: list[str] = []
    mismatch_count = 0
    missingness_mismatch_count = 0
    numeric_mismatch_count = 0
    max_absolute_delta: float | None = None
    max_relative_delta: float | None = None
    first_mismatch: tuple[str, object, object] | None = None

    if resq_shape != engine_shape:
        categories.append("shape")
        mismatch_count += 1
        first_mismatch = ("shape", resq_shape, engine_shape)

    row_limit = min(resq_shape[0], engine_shape[0])
    column_limit = min(resq_shape[1], engine_shape[1])
    for row in range(row_limit):
        for column in range(column_limit):
            resq_value = _matrix_value(resq_values, row, column)
            engine_value = _matrix_value(engine_values, row, column)
            resq_missing = _is_missing(resq_value)
            engine_missing = _is_missing(engine_value)
            if resq_missing or engine_missing:
                if resq_missing != engine_missing:
                    missingness_mismatch_count += 1
                    mismatch_count += 1
                    if first_mismatch is None:
                        first_mismatch = (_format_cell(row, column), resq_value, engine_value)
                continue

            try:
                resq_number = _as_float(resq_value)
                engine_number = _as_float(engine_value)
            except ValueError:
                numeric_mismatch_count += 1
                mismatch_count += 1
                if first_mismatch is None:
                    first_mismatch = (_format_cell(row, column), resq_value, engine_value)
                continue

            if not (math.isfinite(resq_number) and math.isfinite(engine_number)):
                equal = resq_number == engine_number
            else:
                equal = math.isclose(
                    resq_number,
                    engine_number,
                    abs_tol=absolute_tolerance,
                    rel_tol=relative_tolerance,
                )
            if equal:
                continue

            absolute_delta = abs(resq_number - engine_number)
            scale = max(abs(resq_number), abs(engine_number))
            relative_delta = 0.0 if scale == 0 else absolute_delta / scale
            max_absolute_delta = max(max_absolute_delta or 0.0, absolute_delta)
            max_relative_delta = max(max_relative_delta or 0.0, relative_delta)
            numeric_mismatch_count += 1
            mismatch_count += 1
            if first_mismatch is None:
                first_mismatch = (_format_cell(row, column), resq_value, engine_value)

    if missingness_mismatch_count:
        categories.append("missingness")
    if numeric_mismatch_count:
        categories.append("numeric")
    first_cell, first_resq, first_engine = first_mismatch or ("", None, None)
    return {
        "matches": not categories,
        "categories": tuple(categories),
        "resq_shape": resq_shape,
        "engine_shape": engine_shape,
        "mismatch_count": mismatch_count,
        "max_absolute_delta": max_absolute_delta,
        "max_relative_delta": max_relative_delta,
        "first_mismatch_cell": first_cell,
        "resq_value": first_resq,
        "engine_value": first_engine,
    }


def _result_from_payload(
    rc_path: str,
    kind: str,
    payload: dict[str, object],
    comparison: dict[str, object],
) -> ComparisonResult:
    return ComparisonResult(
        rc_path=rc_path,
        kind=kind,
        resq_formula=_safe_text(payload.get("resq_formula")),
        dataset_name=_safe_text(payload.get("name")),
        dataset_type=_safe_text(payload.get("dataset_type")),
        status="match" if comparison["matches"] else "mismatch",
        categories=tuple(comparison["categories"]),
        origin_length=int(payload.get("origin_length") or 0) or None,
        development_length=int(payload.get("development_length") or 0) or None,
        resq_shape=comparison["resq_shape"],
        engine_shape=comparison["engine_shape"],
        mismatch_count=int(comparison["mismatch_count"]),
        max_absolute_delta=comparison["max_absolute_delta"],
        max_relative_delta=comparison["max_relative_delta"],
        first_mismatch_cell=_safe_text(comparison["first_mismatch_cell"]),
        resq_value=comparison["resq_value"],
        engine_value=comparison["engine_value"],
    )


def validate_payload(
    *,
    rc_path: str,
    kind: str,
    payload: dict[str, object],
    temp_dir: Path,
    engine_generator: Callable[..., None] = generate_engine_csv,
    absolute_tolerance: float = DEFAULT_ABSOLUTE_TOLERANCE,
    relative_tolerance: float = DEFAULT_RELATIVE_TOLERANCE,
    progress_callback: ProgressCallback | None = None,
) -> ComparisonResult:
    """Generate one isolated engine CSV and compare it with an extracted ResQ payload."""

    name = _safe_text(payload.get("name"))
    dataset_type = _safe_text(payload.get("dataset_type")) or name
    origin_length = int(payload.get("origin_length") or 0)
    development_length = int(payload.get("development_length") or 0)
    is_vector = kind == "vector"
    if is_vector:
        output_name = _vector_cache_csv_file_name(name, origin_length)
    else:
        output_name = _dataset_cache_csv_file_name(name, origin_length, development_length)
    output_path = temp_dir / output_name

    try:
        _emit_progress(
            progress_callback,
            f"ENGINE START {kind} | {name} | {origin_length}x{development_length}",
        )
        engine_generator(
            project_name=TARGET_PROJECT_NAME,
            rc_path=rc_path,
            dataset_type=dataset_type,
            data_path=output_path,
            origin_length=origin_length,
            development_length=development_length,
            is_vector=is_vector,
            server_root=migration.SERVER_ROOT,
            cumulative=True,
            calendar=False,
        )
        engine_values = read_engine_csv(output_path)
    except (EngineGenerationError, OSError, ValueError) as exc:
        return ComparisonResult(
            rc_path=rc_path,
            kind=kind,
            resq_formula=_safe_text(payload.get("resq_formula")),
            dataset_name=name,
            dataset_type=dataset_type,
            status="engine_generation_error",
            origin_length=origin_length or None,
            development_length=development_length or None,
            error=str(exc),
        )
    except Exception as exc:  # Preserve the individual dataset and continue the run.
        return ComparisonResult(
            rc_path=rc_path,
            kind=kind,
            resq_formula=_safe_text(payload.get("resq_formula")),
            dataset_name=name,
            dataset_type=dataset_type,
            status="engine_generation_error",
            origin_length=origin_length or None,
            development_length=development_length or None,
            error=f"{type(exc).__name__}: {exc}",
        )

    comparison = compare_matrices(
        payload.get("values") if isinstance(payload.get("values"), list) else [],
        engine_values,
        absolute_tolerance=absolute_tolerance,
        relative_tolerance=relative_tolerance,
    )
    return _result_from_payload(rc_path, kind, payload, comparison)


def _record_skip(summary: ValidationSummary, reason: str) -> None:
    summary.skipped_by_reason[reason] = summary.skipped_by_reason.get(reason, 0) + 1


def _resq_formula(dataset: object) -> str:
    """Read a dataset's ResQ Formula property without blocking validation."""

    try:
        return _safe_text(dataset.Formula)
    except Exception:
        return ""


def _read_collection(
    *,
    summary: ValidationSummary,
    rc_path: str,
    kind: str,
    datasets: Iterable[object],
    exporter: Callable[[object], dict[str, object]],
    temp_dir: Path,
    engine_generator: Callable[..., None],
    absolute_tolerance: float,
    relative_tolerance: float,
    progress_callback: ProgressCallback | None,
) -> None:
    for dataset in datasets:
        minimal = _minimal_payload(dataset)
        formula = _resq_formula(dataset)
        eligible, skip_reason = _eligible_dataset(dataset)
        if not eligible:
            _record_skip(summary, skip_reason)
            _emit_progress(
                progress_callback,
                f"SKIP {kind} | {minimal['name'] or '<unnamed>'} | {skip_reason}",
            )
            continue
        try:
            payload = exporter(dataset)
        except Exception as exc:
            result = ComparisonResult(
                rc_path=rc_path,
                kind=kind,
                resq_formula=formula,
                dataset_name=minimal["name"],
                dataset_type=minimal["dataset_type"],
                status="resq_read_error",
                error=f"{type(exc).__name__}: {exc}",
            )
            summary.results.append(result)
            _emit_progress(
                progress_callback,
                f"RESQ READ ERROR {kind} | {result.dataset_name} | {result.error}",
            )
            continue
        payload["resq_formula"] = formula
        result = validate_payload(
            rc_path=rc_path,
            kind=kind,
            payload=payload,
            temp_dir=temp_dir,
            engine_generator=engine_generator,
            absolute_tolerance=absolute_tolerance,
            relative_tolerance=relative_tolerance,
            progress_callback=progress_callback,
        )
        summary.results.append(result)
        detail = f"mismatches={result.mismatch_count}" if result.status == "mismatch" else result.error
        _emit_progress(
            progress_callback,
            f"{result.status.upper()} {kind} | {result.dataset_name} | {detail or 'ok'}",
        )


def _connect_resq(app_factory: Callable[[], object] | None = None) -> object:
    if app_factory is not None:
        app = app_factory()
    else:
        try:
            import win32com.client
        except ImportError as exc:  # pragma: no cover - environment-dependent
            raise RuntimeError("pywin32 is required to connect to ResQ.") from exc
        app = win32com.client.Dispatch("ResQ3Automation.ResQApplication")
    app.ConnectByName(migration.CONNECTION_NAME, migration.USER_NAME, migration.PASSWORD)
    return app


def run_validation(
    *,
    absolute_tolerance: float = DEFAULT_ABSOLUTE_TOLERANCE,
    relative_tolerance: float = DEFAULT_RELATIVE_TOLERANCE,
    app_factory: Callable[[], object] | None = None,
    engine_generator: Callable[..., None] = generate_engine_csv,
    temp_root: Path = DEFAULT_TEMP_ROOT,
    progress_callback: ProgressCallback | None = None,
) -> ValidationSummary:
    """Run the fixed-project live parity validation and return all outcomes."""

    if absolute_tolerance < 0 or relative_tolerance < 0:
        raise ValueError("Comparison tolerances must be non-negative.")

    rc_paths = migration._configured_rc_paths(migration.RC_PATH)
    summary = ValidationSummary(started_at=_utc_now(), rc_paths=rc_paths)
    previous_scope = migration._apply_runtime_scope(TARGET_PROJECT_NAME, migration.SERVER_ROOT)
    app: object | None = None
    temp_root.mkdir(parents=True, exist_ok=True)
    try:
        try:
            _emit_progress(progress_callback, f"CONNECT ResQ | project={TARGET_PROJECT_NAME}")
            app = _connect_resq(app_factory)
            project = app.Projects().Item(TARGET_PROJECT_NAME)
            _emit_progress(progress_callback, f"CONNECTED ResQ | project={TARGET_PROJECT_NAME}")
        except Exception as exc:
            summary.results.append(
                ComparisonResult(
                    rc_path="",
                    kind="connection" if app is None else "project",
                    status="resq_connection_error" if app is None else "project_unavailable",
                    error=f"{type(exc).__name__}: {exc}",
                )
            )
            _emit_progress(progress_callback, f"CONNECTION ERROR | {type(exc).__name__}: {exc}")
            return summary
        with tempfile.TemporaryDirectory(
            prefix="final_resq_engine_validation_", dir=temp_root
        ) as temporary_directory:
            temp_dir = Path(temporary_directory)
            for rc_index, rc_path in enumerate(rc_paths, start=1):
                print(f"RC {rc_index}/{len(rc_paths)}: {rc_path}", flush=True)
                _emit_progress(progress_callback, f"RC START {rc_index}/{len(rc_paths)} | {rc_path}")
                try:
                    reserving_class = project.ReservingClasses().Item(rc_path)
                except Exception as exc:
                    summary.results.append(
                        ComparisonResult(
                            rc_path=rc_path,
                            kind="reserving_class",
                            status="rc_unavailable",
                            error=f"{type(exc).__name__}: {exc}",
                        )
                    )
                    _emit_progress(progress_callback, f"RC UNAVAILABLE | {rc_path} | {type(exc).__name__}: {exc}")
                    continue

                for kind, collection_name, exporter in (
                    ("triangle", "Triangles", export_triangle),
                    ("vector", "Vectors", export_vector),
                ):
                    try:
                        _emit_progress(progress_callback, f"COLLECTION START {kind} | {rc_path}")
                        collection = getattr(reserving_class, collection_name)()
                        _read_collection(
                            summary=summary,
                            rc_path=rc_path,
                            kind=kind,
                            datasets=collection,
                            exporter=exporter,
                            temp_dir=temp_dir,
                            engine_generator=engine_generator,
                            absolute_tolerance=absolute_tolerance,
                            relative_tolerance=relative_tolerance,
                            progress_callback=progress_callback,
                        )
                        _emit_progress(progress_callback, f"COLLECTION COMPLETE {kind} | {rc_path}")
                    except Exception as exc:
                        summary.results.append(
                            ComparisonResult(
                                rc_path=rc_path,
                                kind=kind,
                                status="resq_read_error",
                                error=f"{type(exc).__name__}: {exc}",
                            )
                        )
                        _emit_progress(progress_callback, f"COLLECTION ERROR {kind} | {rc_path} | {type(exc).__name__}: {exc}")
    finally:
        if app is not None:
            try:
                app.Disconnect()
            except Exception:
                pass
        migration._restore_runtime_scope(previous_scope)
        summary.finished_at = _utc_now()
    return summary


def _format_number(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return format(value, ".12g")
    return str(value)


def _markdown_escape(value: object) -> str:
    return _format_number(value).replace("|", "\\|").replace("\n", " ")


def render_markdown_findings(summary: ValidationSummary) -> str:
    """Render the replaceable findings section of the tracked Markdown report."""

    finished = summary.finished_at or _utc_now()
    lines = [
        "## Latest generated findings",
        "",
        f"- Run started (UTC): `{summary.started_at.isoformat(timespec='seconds')}`",
        f"- Run finished (UTC): `{finished.isoformat(timespec='seconds')}`",
        f"- Reserving classes in scope: {len(summary.rc_paths)}",
        f"- Eligible datasets compared: {summary.eligible_count}",
        f"- Matches: {summary.match_count}",
        f"- Issues: {summary.issue_count}",
        f"- Skipped datasets: {summary.skipped_count}",
    ]
    if summary.skipped_by_reason:
        skipped = ", ".join(
            f"{reason}={count}" for reason, count in sorted(summary.skipped_by_reason.items())
        )
        lines.append(f"- Skip reasons: {skipped}")
    lines.extend(["", "### Issues", ""])
    issues = summary.issue_results
    if not issues:
        lines.append("No parity issues were found.")
        return "\n".join(lines)

    lines.extend([
        "| RC path | Kind | Dataset | Status | Categories | Mismatches | First difference |",
        "| --- | --- | --- | --- | --- | ---: | --- |",
    ])
    for item in issues:
        lines.append(
            "| "
            + " | ".join(
                [
                    _markdown_escape(item.rc_path),
                    _markdown_escape(item.kind),
                    _markdown_escape(item.dataset_name),
                    _markdown_escape(item.status),
                    _markdown_escape(", ".join(item.categories) or item.error),
                    str(item.mismatch_count),
                    _markdown_escape(item.first_mismatch_cell or item.error),
                ]
            )
            + " |"
        )
    return "\n".join(lines)


def replace_generated_findings(document: str, findings: str) -> str:
    """Replace only the generated section, preserving hand-authored report guidance."""

    start = document.find(MARKDOWN_START_MARKER)
    end = document.find(MARKDOWN_END_MARKER)
    if start < 0 or end < 0 or end <= start:
        raise ValueError("The Markdown report does not contain valid generated-findings markers.")
    before = document[: start + len(MARKDOWN_START_MARKER)]
    after = document[end:]
    return f"{before}\n\n{findings}\n\n{after}"


def update_markdown_report(path: Path, summary: ValidationSummary) -> None:
    document = path.read_text(encoding="utf-8")
    updated = replace_generated_findings(document, render_markdown_findings(summary))
    path.write_text(updated, encoding="utf-8", newline="\n")


def ensure_markdown_report(path: Path) -> None:
    """Create the project-scoped report template on the first validation run."""

    if path.is_file():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(
            [
                "# ResQ/Data-Engine Parity Validation",
                "",
                f"Project: `{TARGET_PROJECT_NAME}`",
                "",
                "This report is generated by `validate_engine_resq_parity.py`. "
                "It compares live ResQ generated datasets with fresh isolated "
                "ArcRho data-engine output.",
                "",
                "<!-- BEGIN GENERATED FINDINGS -->",
                "",
                "## Latest generated findings",
                "",
                "Not yet run.",
                "",
                "<!-- END GENERATED FINDINGS -->",
                "",
            ]
        ),
        encoding="utf-8",
        newline="\n",
    )


def issues_workbook_path(report_dir: Path) -> Path:
    """Return the project-specific Excel review workbook path."""

    return report_dir / f"final_validation_issues_{TARGET_PROJECT_NAME}.xlsx"


def write_issues_workbook(path: Path, summary: ValidationSummary) -> None:
    """Write an atomically replaced, issues-only Excel report for user review."""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:  # pragma: no cover - dependency/environment dependent
        raise RuntimeError(
            "openpyxl is required to write the Excel issues workbook. "
            "Install it in the Python 3.10 environment and run the validator again."
        ) from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    for label, value in (
        ("Project", TARGET_PROJECT_NAME),
        ("Run started (UTC)", summary.started_at.isoformat(timespec="seconds")),
        ("Run finished (UTC)", (summary.finished_at or _utc_now()).isoformat(timespec="seconds")),
        ("Reserving classes", len(summary.rc_paths)),
        ("Eligible datasets", summary.eligible_count),
        ("Matches", summary.match_count),
        ("Issues", summary.issue_count),
        ("Skipped", summary.skipped_count),
    ):
        summary_sheet.append([label, value])

    issues_sheet = workbook.create_sheet("Issues")
    headers = [
        "RC Path",
        "Dataset",
        "ResQ Formula",
        "Data Format",
        "Dataset Type",
        "Status",
        "Categories",
        "Origin Length",
        "Development Length",
        "ResQ Shape",
        "ArcRho Shape",
        "Mismatch Count",
        "Max Absolute Delta",
        "Max Relative Delta",
        "First Difference",
        "ResQ Value",
        "ArcRho Value",
        "Error",
    ]
    issues_sheet.append(headers)
    for item in summary.issue_results:
        issues_sheet.append(
            [
                item.rc_path,
                item.dataset_name,
                item.resq_formula,
                item.kind,
                item.dataset_type,
                item.status,
                ", ".join(item.categories),
                item.origin_length,
                item.development_length,
                "x".join(map(str, item.resq_shape or ())),
                "x".join(map(str, item.engine_shape or ())),
                item.mismatch_count,
                item.max_absolute_delta,
                item.max_relative_delta,
                item.first_mismatch_cell,
                _format_number(item.resq_value),
                _format_number(item.engine_value),
                item.error,
            ]
        )
        # Formula text commonly starts with "=". Force a string cell so Excel
        # shows the ResQ expression instead of evaluating it as an Excel formula.
        issues_sheet.cell(row=issues_sheet.max_row, column=3).data_type = "s"
    if not summary.issue_results:
        issues_sheet.append(["No issues found."])

    for worksheet in (summary_sheet, issues_sheet):
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        worksheet.freeze_panes = "A2"
        for column_cells in worksheet.columns:
            width = max(len(_format_number(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 60)

    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.stem}-", suffix=".xlsx", dir=path.parent
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, path)
    finally:
        workbook.close()
        if temporary_path.exists():
            temporary_path.unlink()


def _parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare live ResQ generated datasets to fresh ArcRho data-engine output."
    )
    parser.add_argument(
        "--report-dir",
        type=Path,
        default=DEFAULT_REPORT_DIR,
        help=(
            "Directory containing this project's final_validation.md and "
            "project-named Excel issues workbook."
        ),
    )
    parser.add_argument(
        "--atol",
        type=float,
        default=DEFAULT_ABSOLUTE_TOLERANCE,
        help=f"Absolute numerical tolerance (default: {DEFAULT_ABSOLUTE_TOLERANCE:g}).",
    )
    parser.add_argument(
        "--rtol",
        type=float,
        default=DEFAULT_RELATIVE_TOLERANCE,
        help=f"Relative numerical tolerance (default: {DEFAULT_RELATIVE_TOLERANCE:g}).",
    )
    parser.add_argument(
        "--no-progress-window",
        dest="progress_window",
        action="store_false",
        default=True,
        help="Do not open the visible PowerShell window that tails the detailed progress log.",
    )
    return parser.parse_args(argv)


def open_progress_window(progress_path: Path) -> bool:
    """Open a visible Windows terminal that follows the current validation log."""

    if os.name != "nt":
        return False
    escaped_path = str(progress_path).replace("'", "''")
    title = f"ArcRho Validation Progress - {TARGET_PROJECT_NAME}".replace("'", "''")
    command = (
        f"$Host.UI.RawUI.WindowTitle = '{title}'; "
        f"Write-Host 'Following: {escaped_path}'; "
        f"Get-Content -LiteralPath '{escaped_path}' -Wait"
    )
    try:
        subprocess.Popen(
            ["powershell.exe", "-NoProfile", "-NoExit", "-Command", command],
            creationflags=getattr(subprocess, "CREATE_NEW_CONSOLE", 0),
        )
    except OSError:
        return False
    return True


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    if args.atol < 0 or args.rtol < 0:
        raise SystemExit("--atol and --rtol must be non-negative.")
    report_dir = args.report_dir.resolve()
    markdown_path = report_dir / "final_validation.md"
    workbook_path = issues_workbook_path(report_dir)
    progress_path = report_dir / "validation_progress.tmp.log"
    ensure_markdown_report(markdown_path)
    report_dir.mkdir(parents=True, exist_ok=True)
    progress_path.write_text("", encoding="utf-8", newline="\n")
    print(f"Detailed progress log: {progress_path}", flush=True)
    if args.progress_window:
        if open_progress_window(progress_path):
            print("Opened PowerShell progress window.", flush=True)
        else:
            print("Could not open a progress window; follow the progress log manually.", flush=True)

    with progress_path.open("w", encoding="utf-8", newline="\n") as progress_file:
        def log_progress(message: str) -> None:
            timestamp = _utc_now().isoformat(timespec="seconds")
            progress_file.write(f"{timestamp} {message}\n")
            progress_file.flush()

        log_progress(f"VALIDATION START | project={TARGET_PROJECT_NAME}")
        try:
            summary = run_validation(
                absolute_tolerance=args.atol,
                relative_tolerance=args.rtol,
                progress_callback=log_progress,
            )
        except Exception as exc:
            log_progress(f"VALIDATION FATAL | {type(exc).__name__}: {exc}")
            raise SystemExit(f"Validation could not start: {type(exc).__name__}: {exc}") from exc

        log_progress("WRITING REPORTS")
        update_markdown_report(markdown_path, summary)
        write_issues_workbook(workbook_path, summary)
        log_progress(
            "VALIDATION COMPLETE | "
            f"eligible={summary.eligible_count} matches={summary.match_count} "
            f"issues={summary.issue_count} skipped={summary.skipped_count}"
        )
    print(
        "Validation complete: "
        f"eligible={summary.eligible_count}, matches={summary.match_count}, "
        f"issues={summary.issue_count}, skipped={summary.skipped_count}"
    )
    print(f"Markdown findings: {markdown_path}")
    print(f"Excel issues: {workbook_path}")
    return 0 if not summary.issue_count else 1


if __name__ == "__main__":  # pragma: no cover - command-line entry point
    raise SystemExit(main())
