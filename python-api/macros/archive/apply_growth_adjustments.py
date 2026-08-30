# <arcrho-macro>
# Title: Apply Growth Adjustments
# Version: 1.0.0
# Release Note: Select Q1-Q4 workbooks using the March, June, September, and December review windows.
# Description: Read growth adjustments from the latest available reserve-review workbook for the
#   active DFM's segment (all 12 paths), applies them to the active DFM method, and adds
#   adjustment notes.  Annual segments use C20 accounting cutoff; quarterly segments
#   (Penn&CT BI/PIP, NY BI/PIP) use C21.
# </arcrho-macro>

from __future__ import annotations

import datetime
import re
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover - runtime user dependency
    load_workbook = None
    _LOAD_WORKBOOK_ERROR = exc
else:
    _LOAD_WORKBOOK_ERROR = None

try:
    from arcrho_api.exceptions import DfmDataError
except Exception:  # pragma: no cover - script can still show useful errors
    DfmDataError = ValueError


def _current_review_quarter() -> str:
    """Return the review quarter for the current three-month workbook window.

    Review workbooks roll forward in March, June, September, and December.  Thus,
    March-May use Q1, June-August use Q2, September-November use Q3, and
    December-February use Q4.  The January-February Q4 workbook belongs to the
    previous calendar year.
    """
    today = datetime.date.today()
    review_quarter = ((today.month - 3) % 12) // 3 + 1
    review_year = today.year - 1 if today.month in (1, 2) else today.year
    return f"{review_year}Q{review_quarter}"


_REVIEW_QUARTER = _current_review_quarter()
GROWTH_ADJUSTMENT_WORKBOOK = (
    rf"E:\ResQ\Automations\Reserve Review\{_REVIEW_QUARTER}\Growth Adjustment {_REVIEW_QUARTER}.xlsx"
)
GROWTH_ADJUSTMENT_SHEET = "Summary"
RESQ_PATH_WORKBOOK = r"E:\ResQ\Automations\Reserve Review\ResQ Path.xlsx"
RESQ_PATH_SHEET = "Sheet1"

# Row numbers (1-indexed, as used by openpyxl cell notation)
HEADER_ROW = 6
ANNUAL_INCURRED_ROWS = [7, 8]
ANNUAL_PAID_ROWS = [10, 11]
ANNUAL_COUNTS_ROWS = [13, 14]
QUARTERLY_INCURRED_ROWS = [7, 8, 9]
QUARTERLY_PAID_ROWS = [10, 11, 12]
QUARTERLY_COUNTS_ROWS = [13, 14, 15]
ANNUAL_CUTOFF_CELL = "C20"
QUARTERLY_CUTOFF_CELL = "C21"
ANNUAL_PANEL_LABEL = "Annual Adjustment"
QUARTERLY_PANEL_LABEL = "Quarterly Adjustment"

PRE_ADJUSTMENT_CELL_NOTE = "Selected before adjustments."


# ---------------------------------------------------------------------------
# Workbook loading helpers
# ---------------------------------------------------------------------------

def _load_segment_column_map(sheet: Any) -> dict[str, tuple[str, bool]]:
    """Scan the header row to build {segment_name: (col_letter, is_quarterly)}.

    Columns belong to the quarterly panel once the 'Quarterly Adjustment' section
    header has been encountered; all others are annual.
    """
    mapping: dict[str, tuple[str, bool]] = {}
    is_quarterly = False
    for cell in sheet[HEADER_ROW]:
        val = _clean_text(cell.value)
        if not val:
            continue
        val_lower = val.lower()
        if val_lower == ANNUAL_PANEL_LABEL.lower():
            is_quarterly = False
        elif val_lower == QUARTERLY_PANEL_LABEL.lower():
            is_quarterly = True
        else:
            mapping[val] = (cell.column_letter, is_quarterly)
    return mapping


def load_segment_path_mapping(resq_path_file: str = RESQ_PATH_WORKBOOK) -> dict[str, str]:
    """Return {reserving_class_path: segment_name} from ResQ Path.xlsx."""
    if load_workbook is None:
        raise DfmDataError(f"openpyxl is required: {_LOAD_WORKBOOK_ERROR}")
    resq_path = Path(resq_path_file)
    if not resq_path.exists():
        raise DfmDataError(f"ResQ Path workbook not found: {resq_path}")
    wb = load_workbook(str(resq_path), data_only=True, read_only=True)
    ws = wb[RESQ_PATH_SHEET]
    path_to_segment: dict[str, str] = {}
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        if not row or row[0] is None:
            continue
        segment_name = _clean_text(row[0])
        segment_path = _clean_text(row[1]) if len(row) > 1 and row[1] is not None else ""
        if segment_name and segment_path:
            path_to_segment[segment_path] = segment_name
    wb.close()
    return path_to_segment


def _resolve_segment_name(reserving_class: str, resq_path_file: str) -> str | None:
    """Return the segment name for a reserving class path, or None if not found."""
    try:
        path_to_segment = load_segment_path_mapping(resq_path_file)
    except Exception:
        return None
    if reserving_class in path_to_segment:
        return path_to_segment[reserving_class]
    rc_lower = reserving_class.lower()
    for path, segment in path_to_segment.items():
        if path.lower() == rc_lower:
            return segment
    return None


def _read_segment_adjustments(sheet: Any, col_letter: str, is_quarterly: bool) -> dict[str, list[Any]]:
    if is_quarterly:
        incurred_rows = QUARTERLY_INCURRED_ROWS
        paid_rows = QUARTERLY_PAID_ROWS
        counts_rows = QUARTERLY_COUNTS_ROWS
        cutoff_cell = QUARTERLY_CUTOFF_CELL
    else:
        incurred_rows = ANNUAL_INCURRED_ROWS
        paid_rows = ANNUAL_PAID_ROWS
        counts_rows = ANNUAL_COUNTS_ROWS
        cutoff_cell = ANNUAL_CUTOFF_CELL
    return {
        "incurred loss": [sheet[f"{col_letter}{r}"].value for r in incurred_rows],
        "paid loss": [sheet[f"{col_letter}{r}"].value for r in paid_rows],
        "counts": [sheet[f"{col_letter}{r}"].value for r in counts_rows],
        "accounting cutoff": [sheet[cutoff_cell].value],
        "other": [0],
    }


def load_adjustments_from_workbook(
    file_path: str = GROWTH_ADJUSTMENT_WORKBOOK,
    *,
    reserving_class: str | None = None,
    resq_path_file: str = RESQ_PATH_WORKBOOK,
) -> dict[str, list[Any]]:
    """Load growth adjustments for the given reserving class from the workbook.

    When *reserving_class* is provided the function looks up the matching segment
    in ResQ Path.xlsx, finds its column in the Growth Adjustment workbook, and reads
    annual or quarterly rows accordingly.  When omitted it falls back to the original
    hard-coded Coll / annual column for backwards compatibility.
    """
    if load_workbook is None:
        raise DfmDataError(f"openpyxl is required to load adjustment workbooks: {_LOAD_WORKBOOK_ERROR}")
    workbook_path = Path(file_path)
    if not workbook_path.exists():
        raise DfmDataError(f"Growth adjustment workbook was not found: {workbook_path}")
    sheet = load_workbook(str(workbook_path), data_only=True)[GROWTH_ADJUSTMENT_SHEET]

    if reserving_class is None:
        # Legacy fallback: original hard-coded Coll column (F), annual adjustment
        return _read_segment_adjustments(sheet, "F", is_quarterly=False)

    col_map = _load_segment_column_map(sheet)

    segment_name = _resolve_segment_name(reserving_class, resq_path_file)
    if segment_name is None:
        raise DfmDataError(
            f"Could not find a segment for reserving class {reserving_class!r}. "
            f"Check that the path exists in {resq_path_file}"
        )
    if segment_name not in col_map:
        raise DfmDataError(
            f"Segment {segment_name!r} was not found in the Growth Adjustment workbook "
            f"(available: {sorted(col_map)})"
        )

    col_letter, is_quarterly = col_map[segment_name]
    return _read_segment_adjustments(sheet, col_letter, is_quarterly)


# ---------------------------------------------------------------------------
# Internal utilities (unchanged)
# ---------------------------------------------------------------------------

def _clean_text(value: Any) -> str:
    return " ".join(str(value if value is not None else "").split()).strip()


def _normalize_label(value: Any) -> str:
    return _clean_text(value)


def _label_key(value: Any) -> str:
    label = _normalize_label(value)
    if ":" in label:
        prefix, rest = label.split(":", 1)
        if prefix.strip().isdigit():
            label = rest.strip()
    return label.lower()


def _number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number != number or number in (float("inf"), float("-inf")):
        return None
    return number


def _adjustment_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            if text.endswith("%"):
                return round(float(text[:-1].strip()) / 100, 4)
            return float(text)
        except ValueError as err:
            raise DfmDataError(f"Invalid adjustment value: {value!r}") from err
    number = _number(value)
    if number is None:
        return None
    return round(number, 4)


def _ensure_matrix(container: dict[str, Any], key: str, rows: int, cols: int, fill: Any = 0) -> list[list[Any]]:
    existing = container.get(key)
    if not isinstance(existing, list):
        existing = []
    existing = [row if isinstance(row, list) else [] for row in existing]
    while len(existing) < rows:
        existing.append([])
    for row in existing:
        while len(row) < cols:
            row.append(fill)
    existing = existing[:rows]
    for index, row in enumerate(existing):
        existing[index] = row[:cols]
    container[key] = existing
    return existing


def _normalize_adjustment_list(values: Any) -> list[float]:
    if values is None:
        return []
    if isinstance(values, (str, bytes)):
        source = [values]
    else:
        try:
            source = list(values)
        except TypeError:
            source = [values]
    out: list[float] = []
    for value in source:
        number = _adjustment_number(value)
        if number is not None:
            out.append(number)
    return out


def _normalize_adjustments(adjustments: Any) -> dict[str, list[float]]:
    if not isinstance(adjustments, dict):
        raise DfmDataError("apply_adjustments requires adjustments to be a dictionary.")
    aliases = {
        "count": "counts",
        "claim count": "counts",
        "claim counts": "counts",
        "paid": "paid loss",
        "paid losses": "paid loss",
        "incurred": "incurred loss",
        "incurred losses": "incurred loss",
        "accounting": "accounting cutoff",
        "cutoff": "accounting cutoff",
    }
    out: dict[str, list[float]] = {
        "counts": [],
        "paid loss": [],
        "incurred loss": [],
        "accounting cutoff": [],
        "other": [],
    }
    for key, values in adjustments.items():
        normalized_key = aliases.get(_clean_text(key).lower(), _clean_text(key).lower())
        if normalized_key in out:
            out[normalized_key] = _normalize_adjustment_list(values)
    return out


def _adjustment_method_kind(dfm: Any) -> str:
    try:
        info = dfm.output_vector_dataset_type()
    except Exception:
        info = None
    if info is not None:
        dataset_type = _clean_text(getattr(info, "name", ""))
        category = _clean_text(getattr(info, "category", ""))
        dataset_type_lower = dataset_type.lower()
        category_lower = category.lower()
        if "52" in dataset_type:
            return "skip"
        if category_lower == "c claim count":
            return "counts"
        if category_lower == "h severity":
            return "severity"
        if "paid" in dataset_type_lower or "salv dfm" in dataset_type_lower or "subr dfm" in dataset_type_lower:
            return "paid"
        if "incurred" in dataset_type_lower:
            return "incurred"

    text = f"{dfm.output_vector} {dfm.name}".lower()
    if re.search(r"(^|\D)52(\D|$)", text):
        return "skip"
    if re.search(r"(^|\W)h\s*\d+", text) or "severity" in text:
        return "severity"
    if "paid" in text or "salv" in text or "subr" in text:
        return "paid"
    if "incurred" in text:
        return "incurred"
    if re.search(r"(^|\W)c\s*\d+", text) or "claim count" in text or "counts" in text:
        return "counts"
    if "reported" in text or "cwp" in text or "cwop" in text:
        return "counts"
    return "incurred"


def _selected_average_row(selected: list[list[Any]], col: int) -> int | None:
    for row, row_values in enumerate(selected):
        if col < len(row_values) and bool(row_values[col]):
            return row
    return None


def _format_adjustment_percent(value: float) -> str:
    return f"{round(abs(value) * 100, 2):g}%"


def _factor_note_part(value: float) -> str:
    if value > 0:
        return f"1+{_format_adjustment_percent(value)}"
    if value < 0:
        return f"1-{_format_adjustment_percent(value)}"
    return "1"


def _compound_adjustment_part(values: list[float], col: int, *, formula_style: str) -> dict[str, Any]:
    current = values[col] if col < len(values) else 0.0
    next_value = values[col + 1] if col + 1 < len(values) else 0.0
    current_factor = 1.0 + current
    next_factor = 1.0 + next_value
    if col + 1 < len(values):
        value = current_factor / next_factor if next_factor else current_factor
    else:
        value = current_factor
    if formula_style == "left":
        formula = _factor_note_part(current)
        if col + 1 < len(values) and next_value != 0:
            formula = f"{formula}/({_factor_note_part(next_value)})"
        if formula.count("(") == 1 and "/" not in formula:
            formula = formula.replace("(", "").replace(")", "")
    else:
        if col + 1 < len(values) and next_factor != 1:
            formula = f"{round(current_factor, 4):g}/{round(next_factor, 4):g}"
        else:
            formula = f"{round(current_factor, 4):g}"
    return {"value": value, "formula": formula}


def _adjustment_for_kind(kind: str, adjustments: dict[str, list[float]], col: int) -> dict[str, Any]:
    counts_left = _compound_adjustment_part(adjustments["counts"], col, formula_style="left")
    counts_right = _compound_adjustment_part(adjustments["counts"], col, formula_style="right")
    incurred_left = _compound_adjustment_part(adjustments["incurred loss"], col, formula_style="left")
    incurred_right = _compound_adjustment_part(adjustments["incurred loss"], col, formula_style="right")
    paid_left = _compound_adjustment_part(adjustments["paid loss"], col, formula_style="left")
    paid_right = _compound_adjustment_part(adjustments["paid loss"], col, formula_style="right")
    if kind == "counts":
        return {"factor": counts_right["value"], "left": counts_left["formula"], "right": counts_right["formula"]}
    if kind == "paid":
        return {"factor": paid_right["value"], "left": paid_left["formula"], "right": paid_right["formula"]}
    if kind == "severity":
        factor = incurred_right["value"] / counts_right["value"] if counts_right["value"] else 1.0
        left = f"{incurred_left['formula']}/{counts_left['formula']}"
        right = f"({incurred_right['formula']})/({counts_right['formula']})"
        return {"factor": factor, "left": left, "right": right}
    return {"factor": incurred_right["value"], "left": incurred_left["formula"], "right": incurred_right["formula"]}


def _has_meaningful_adjustment(growth: float, accounting_cutoff: float, other_factor: float) -> bool:
    return any(abs(value - 1.0) > 0.0000001 for value in (growth, accounting_cutoff, other_factor))


def _display_average_label(label: str) -> str:
    text = _normalize_label(label)
    if ":" in text:
        _prefix, rest = text.split(":", 1)
        if rest.strip():
            return rest.strip()
    return text


def _formula_number(value: float) -> str:
    return f"{float(value):.10g}"


def _compact_formula_number_text(value: Any) -> str:
    text = str(value)

    def repl(match: re.Match[str]) -> str:
        try:
            number = float(match.group(0))
        except ValueError:
            return match.group(0)
        return f"{number:.4f}".rstrip("0").rstrip(".")

    return re.sub(r"(?<![A-Za-z0-9_])[-+]?\d+\.\d+", repl, text)


def _format_note_multiplier(value: float) -> str:
    return f"{float(value):.4f}".rstrip("0").rstrip(".")


def _quote_formula_label(label: str) -> str:
    return '"' + _display_average_label(label).replace('"', "'") + '"'


def _build_user_entry_formula(
    label: str,
    adjustment: dict[str, Any],
    accounting_cutoff: float,
    other_factor: float,
) -> str:
    parts = [_quote_formula_label(label)]
    if abs(float(adjustment["factor"]) - 1.0) > 0.0000001:
        parts.append(f"({adjustment['right']})")
    if abs(accounting_cutoff - 1.0) > 0.0000001:
        parts.append(_formula_number(accounting_cutoff))
    if abs(other_factor - 1.0) > 0.0000001:
        parts.append(_formula_number(other_factor))
    return "= " + " * ".join(parts)


def _mark_selected_before_adjustment(dfm: Any, col: int, label: str) -> None:
    dfm.clear_cell_notes_for_development(col + 1)
    dfm.set_cell_note(_display_average_label(label), col + 1, PRE_ADJUSTMENT_CELL_NOTE)


def _format_adjustment_note(
    dfm: Any,
    col: int,
    label: str,
    average_value: float,
    final_value: float,
    adjustment: dict[str, Any],
    accounting_cutoff: float,
    other_factor: float,
) -> str:
    lines = [f"For development period {dfm.dev_period(col + 1)}:"]
    formula_parts = [f"{average_value:.4f}"]
    if abs(float(adjustment["factor"]) - 1.0) > 0.0000001:
        adjustment_right = _compact_formula_number_text(adjustment["right"])
        lines.append(f"  - Apply growth adjustments of {adjustment['left']} = {adjustment_right};")
        formula_parts.append(adjustment_right)
    if abs(accounting_cutoff - 1.0) > 0.0000001:
        cutoff_text = _format_note_multiplier(accounting_cutoff)
        lines.append(f"  - Apply accounting cutoff 1+{accounting_cutoff - 1.0:.2%} = {cutoff_text};")
        formula_parts.append(cutoff_text)
    if abs(other_factor - 1.0) > 0.0000001:
        formula_parts.append(_format_note_multiplier(other_factor))
    display_label = _display_average_label(label)
    lines.append(f"  - Selected average factor: \"{display_label}\" ({average_value:.4f})")
    lines.append(f"  - Selected LDF after adjustments: {' * '.join(formula_parts)} = {final_value:.4f}")
    return "\n".join(lines)


def _clear_adjustment_notes(dfm: Any) -> None:
    keywords = (
        "For development period ",
        "Apply growth adjustments of ",
        "Apply accounting cutoff ",
        "Selected average factor: ",
        "Selected LDF after adjustments: ",
        "Skipped: no actuary notes",
    )
    lines = [line for line in dfm.notes.splitlines() if not any(keyword in line for keyword in keywords)]
    dfm.update_notes("\n".join(lines).strip())


def apply_adjustments(
    dfm: Any,
    selection: str | None = None,
    *,
    adjustments: dict[str, Any],
    other_adjustment: Any = None,
    add_notes: bool = True,
    clear_prior_notes: bool = True,
) -> Any:
    if selection:
        dfm.set_selected_average(selection)

    normalized = _normalize_adjustments(adjustments)
    other_values = _normalize_adjustment_list(
        other_adjustment if other_adjustment is not None else normalized.get("other", [0])
    )
    dev_count = max(
        [len(values) for values in normalized.values()] + [len(other_values), dfm._average_col_count()],
        default=0,
    )
    if dev_count <= 0:
        return dfm
    for key in ("counts", "paid loss", "incurred loss", "accounting cutoff", "other"):
        values = normalized.setdefault(key, [])
        if len(values) < dev_count:
            values.extend([0.0] * (dev_count - len(values)))
    if len(other_values) < dev_count:
        other_values.extend([0.0] * (dev_count - len(other_values)))

    method_kind = _adjustment_method_kind(dfm)
    if method_kind == "skip":
        return dfm

    labels = dfm._average_labels()
    selected = _ensure_matrix(dfm.average_formulas, "selected", len(labels), dfm._average_col_count(), 0)
    values = _ensure_matrix(dfm.average_formulas, "values", len(labels), dfm._average_col_count(), None)
    if clear_prior_notes:
        _clear_adjustment_notes(dfm)

    changed = False
    note_blocks: list[str] = []
    for col in range(min(dev_count, dfm._average_col_count())):
        selected_row = _selected_average_row(selected, col)
        if selected_row is None:
            if selection:
                selected_row = dfm._ensure_average_label(selection)
                labels = dfm._average_labels()
                selected = _ensure_matrix(dfm.average_formulas, "selected", len(labels), dfm._average_col_count(), 0)
                values = _ensure_matrix(dfm.average_formulas, "values", len(labels), dfm._average_col_count(), None)
            else:
                continue
        if selected_row >= len(labels):
            continue
        average_value = _number(values[selected_row][col] if col < len(values[selected_row]) else None)
        if average_value is None:
            continue
        adjustment = _adjustment_for_kind(method_kind, normalized, col)
        accounting_cutoff = 1.0 if method_kind == "severity" else 1.0 + normalized["accounting cutoff"][col]
        other_factor = 1.0 + other_values[col]
        final_value = average_value * adjustment["factor"] * accounting_cutoff * other_factor
        if not _has_meaningful_adjustment(adjustment["factor"], accounting_cutoff, other_factor):
            continue
        _mark_selected_before_adjustment(dfm, col, labels[selected_row])
        final_user_value = round(final_value, 6)
        formula = _build_user_entry_formula(labels[selected_row], adjustment, accounting_cutoff, other_factor)
        if hasattr(dfm, "set_user_formula"):
            dfm.set_user_formula(formula, final_user_value, col + 1)
        else:
            dfm.set_user_ratio(final_user_value, col + 1)
            user_row = dfm._ensure_average_label("User Entry")
            inputs = _ensure_matrix(dfm.average_formulas, "inputs", len(dfm._average_labels()), dfm._average_col_count(), "")
            inputs[user_row][col] = formula
        changed = True
        if add_notes:
            note_blocks.append(_format_adjustment_note(
                dfm,
                col,
                labels[selected_row],
                average_value,
                final_user_value,
                adjustment,
                accounting_cutoff,
                other_factor,
            ))

    if add_notes and note_blocks:
        existing = dfm.notes
        suffix = "\n\n".join(note_blocks)
        dfm.update_notes(f"{existing}\n\n{suffix}" if existing else suffix)
    if not changed and add_notes:
        dfm.add_notes("No growth/accounting cutoff adjustments were needed for this method.")
    return dfm


def run_macro(active_dfm, active_context=None):
    reserving_class = getattr(active_dfm, "reserving_class", None)
    adjustment = load_adjustments_from_workbook(reserving_class=reserving_class)
    apply_adjustments(active_dfm, adjustments=adjustment)
    return active_dfm
