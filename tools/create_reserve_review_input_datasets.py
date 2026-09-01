"""Create the Inputs.xlsx-linked input vectors for every reserving class of one project.

Each quarter's reserve review starts from one workbook, ``Inputs.xlsx``, that
carries the accounting cutoff for every segment and a per-segment block of
growth adjustment factors. Those four series belong in ArcRho as ordinary
manual-input vectors whose cells are ArcRho in-cell formulas pointing back at
the workbook, so a later Refresh re-reads the sheet rather than asking anyone
to retype it:

    ='E:\\Actuarial\\Reserve Review\\2026Q2\\[Inputs.xlsx]NJ BI'!G5:G14

``ResQ Path.xlsx`` is the list of reserving classes to visit and says whether
each one is annual or quarterly, which decides both the vector's period length
and which accounting cutoff sheet it reads. The growth adjustment factors come
from the "Adjustment Factors" block of the segment's own sheet - columns G, H
and I - not from the raw growth block on the left of the same sheet.

The work runs through the running ArcRho desktop app, so every write takes the
same route as the equivalent click in the app: the dataset is created, its
sidecar saved with the formula links and the values they resolve to, and the
reserving-class index and dependent propagation are left consistent. A dataset
that already exists is left exactly as it is.

Examples:

    python tools/create_reserve_review_input_datasets.py --dry-run
    python tools/create_reserve_review_input_datasets.py --project "NJ_Annual_Prod_2026 Q2-May"
    python tools/create_reserve_review_input_datasets.py --quarter 2026Q2
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "frontend"))
sys.path.insert(0, str(REPO_ROOT / "python-api" / "src"))

from arcrho_api.ui import ArcRhoUI  # noqa: E402
from app_server.services.dataset_number_format_service import (  # noqa: E402
    number_format_decimal_places,
)


# ── Change these before a run ────────────────────────────────────────────────

# The ArcRho project the datasets are created in.
TARGET_PROJECT = "NJ_Annual_Prod_2026 Q2-May Test"

# Blank derives the source folder from today's month: 1-3 gives Q1, 4-6 Q2,
# 7-9 Q3, 10-12 Q4. Set it to a folder name such as "2026Q2" to read a quarter
# other than the current one.
SOURCE_QUARTER_FOLDER = ""

# How the created vectors display their numbers. Growth adjustment factors sit
# just above 1, so the app-wide default of one decimal would show every one of
# them as 1.0. Blank falls back to whatever default the app holds for the type.
NUMBER_FORMAT = "0,000.0000"


# ── Fixed inputs ─────────────────────────────────────────────────────────────

RESERVING_CLASS_WORKBOOK = r"E:\ResQ\Automations\Reserve Review\ResQ Path.xlsx"
SOURCE_ROOT = r"E:\Actuarial\Reserve Review"
SOURCE_WORKBOOK_NAME = "Inputs.xlsx"

# ResQ Path.xlsx names two New Jersey bodily injury classes; the workbook has
# one "NJ BI" segment sheet that both of them read.
SEGMENT_ALIASES = {"NJ BIR51": "NJ BI", "NJ BIx51": "NJ BI"}

CUTOFF_SHEETS = {12: "Annual Accounting Cutoff", 3: "Quarterly Accounting Cutoff"}
CUTOFF_DATASET = "Accounting Cutoff"
CUTOFF_LABEL_COLUMN = "A"
CUTOFF_VALUE_COLUMN = "B"

# The "Adjustment Factors" block of a segment sheet: period labels in F, and
# the factors a triangle is multiplied by in G, H and I.
GROWTH_LABEL_COLUMN = "F"
GROWTH_DATASETS = (
    ("Growth Adjustment--Incurred", "G"),
    ("Growth Adjustment--Counts", "H"),
    ("Growth Adjustment--Paid", "I"),
)

PERIOD_LENGTHS = {"annual": 12, "quarterly": 3}

# A save leaves a propagation walk holding its reserving class, and the next
# write in that class is refused while the hold stands.
HOLD_POLL_SECONDS = 2.0
HOLD_TIMEOUT_SECONDS = 600.0

REQUEST_TIMEOUT_SECONDS = 300.0


class RunError(RuntimeError):
    """Raised when the run cannot continue without guessing."""

    def __init__(self, message: str, *, http_status: int | None = None) -> None:
        super().__init__(message)
        self.http_status = http_status


# ── The running ArcRho app ───────────────────────────────────────────────────

class App:
    """The routes this script drives on the local ArcRho app server."""

    def __init__(self, base_url: str) -> None:
        self.base_url = base_url.rstrip("/")

    def _call(
        self,
        path: str,
        *,
        method: str,
        params: Dict[str, Any] | None = None,
        payload: Dict[str, Any] | None = None,
    ) -> Dict[str, Any]:
        url = f"{self.base_url}{path}"
        if params:
            url = f"{url}?{urllib.parse.urlencode(params)}"
        data = json.dumps(payload).encode("utf-8") if payload is not None else None
        request = urllib.request.Request(
            url,
            data=data,
            headers={"Content-Type": "application/json"},
            method=method,
        )
        try:
            with urllib.request.urlopen(request, timeout=REQUEST_TIMEOUT_SECONDS) as response:
                body = response.read().decode("utf-8")
        except urllib.error.HTTPError as err:
            raise RunError(
                f"{method} {path} failed ({err.code}): {_error_detail(err)}",
                http_status=err.code,
            ) from err
        except urllib.error.URLError as err:
            raise RunError(f"ArcRho is not reachable at {self.base_url}: {err.reason}") from err
        parsed = json.loads(body) if body else {}
        return parsed if isinstance(parsed, dict) else {"result": parsed}

    def get(self, path: str, params: Dict[str, Any] | None = None) -> Dict[str, Any]:
        return self._call(path, method="GET", params=params)

    def post(self, path: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        return self._call(path, method="POST", payload=payload)


def _error_detail(err: urllib.error.HTTPError) -> str:
    body = err.read().decode("utf-8", errors="replace")
    try:
        parsed = json.loads(body)
    except json.JSONDecodeError:
        return body[:300]
    detail = parsed.get("detail") if isinstance(parsed, dict) else None
    return str(detail or body)[:300]


def wait_until_writable(app: App, project: str, reserving_class: str) -> None:
    deadline = time.monotonic() + HOLD_TIMEOUT_SECONDS
    while True:
        busy = app.get(
            "/dependent_propagation/reserving_class_busy",
            {"project_name": project, "reserving_class": reserving_class},
        )
        if not busy.get("busy"):
            return
        if time.monotonic() >= deadline:
            raise RunError(
                f"{reserving_class} is still held by a propagation walk after "
                f"{int(HOLD_TIMEOUT_SECONDS)}s ({busy.get('reason') or 'no reason given'})."
            )
        time.sleep(HOLD_POLL_SECONDS)


# ── The workbooks ────────────────────────────────────────────────────────────

def quarter_folder_for(today: dt.date) -> str:
    return f"{today.year}Q{(today.month - 1) // 3 + 1}"


def label_text(value: Any) -> str:
    """The comparable text of one period label cell.

    Annual sheets hold the year as a number, so 2017.0 and 2017 must both read
    as the "2017" an ArcRho origin label uses.
    """

    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def read_reserving_classes(path: str) -> List[Tuple[str, str, int]]:
    """Return ``(segment, reserving class path, period length)`` for every listed class."""

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(min_row=2, max_col=3, values_only=True))
    finally:
        workbook.close()

    classes: List[Tuple[str, str, int]] = []
    for segment, class_path, period in rows:
        segment = str(segment or "").strip()
        class_path = str(class_path or "").strip()
        period_key = str(period or "").strip().casefold()
        if not segment or not class_path:
            continue
        if period_key not in PERIOD_LENGTHS:
            raise RunError(
                f"{segment}: '{period}' is not a period this script knows; "
                f"expected one of {', '.join(sorted(PERIOD_LENGTHS))}."
            )
        classes.append((segment, class_path, PERIOD_LENGTHS[period_key]))
    if not classes:
        raise RunError(f"{path} lists no reserving classes.")
    return classes


class SourceWorkbook:
    """The period labels of ``Inputs.xlsx``, one map of label to row per sheet."""

    def __init__(self, folder: Path) -> None:
        self.folder = folder
        self.path = folder / SOURCE_WORKBOOK_NAME
        if not self.path.exists():
            raise RunError(
                f"{self.path} does not exist. Set SOURCE_QUARTER_FOLDER at the top of this "
                "script, or pass --quarter, to read a different quarter's folder."
            )
        workbook = openpyxl.load_workbook(str(self.path), data_only=True, read_only=True)
        try:
            self._sheet_by_segment: Dict[str, str] = {}
            self._labels: Dict[Tuple[str, str], Dict[str, int]] = {}
            for sheet in workbook.worksheets:
                title = sheet.title
                rows = list(sheet.iter_rows(min_row=1, max_col=9, values_only=True))
                if title in CUTOFF_SHEETS.values():
                    self._labels[(title, CUTOFF_LABEL_COLUMN)] = _label_rows(rows, CUTOFF_LABEL_COLUMN)
                    continue
                heading = _cell(rows, 0, 0)
                segment = _cell(rows, 1, 0)
                if heading != "Growth Adjustment" or not segment:
                    continue
                self._sheet_by_segment[segment] = title
                self._labels[(title, GROWTH_LABEL_COLUMN)] = _label_rows(rows, GROWTH_LABEL_COLUMN)
        finally:
            workbook.close()

        missing = [name for name in CUTOFF_SHEETS.values() if (name, CUTOFF_LABEL_COLUMN) not in self._labels]
        if missing:
            raise RunError(f"{self.path} is missing the sheet(s): {', '.join(missing)}.")

    def segment_sheet(self, segment: str) -> str:
        name = SEGMENT_ALIASES.get(segment, segment)
        sheet = self._sheet_by_segment.get(name)
        if not sheet:
            known = ", ".join(sorted(self._sheet_by_segment)) or "none"
            raise RunError(f"no segment sheet for '{name}' in {self.path.name}; the workbook has {known}.")
        return sheet

    def row_span(self, sheet: str, column: str, origin_labels: Sequence[str]) -> Tuple[int, int]:
        """The first and last worksheet row holding this vector's origin labels.

        The labels must sit on consecutive rows in the dataset's own order, or
        one Excel range could not stand for the whole vector.
        """

        rows = self._labels.get((sheet, column))
        if rows is None:
            raise RunError(f"'{sheet}' has no period labels in column {column}.")
        first = rows.get(label_text(origin_labels[0]))
        if first is None:
            raise RunError(f"'{sheet}' column {column} has no row for period {origin_labels[0]}.")
        for offset, origin in enumerate(origin_labels):
            expected = first + offset
            found = rows.get(label_text(origin))
            if found != expected:
                raise RunError(
                    f"'{sheet}' column {column} does not list {origin} on row {expected}; "
                    "the periods are missing or out of order."
                )
        return first, first + len(origin_labels) - 1


def _cell(rows: Sequence[Sequence[Any]], row_index: int, column_index: int) -> str:
    if row_index >= len(rows):
        return ""
    row = rows[row_index]
    return label_text(row[column_index]) if column_index < len(row) else ""


def _label_rows(rows: Sequence[Sequence[Any]], column: str) -> Dict[str, int]:
    index = ord(column.upper()) - ord("A")
    labels: Dict[str, int] = {}
    for offset in range(len(rows)):
        text = _cell(rows, offset, index)
        if text:
            labels.setdefault(text, offset + 1)
    return labels


def excel_reference(folder: Path, sheet: str, column: str, first_row: int, last_row: int) -> str:
    source = f"{folder}\\[{SOURCE_WORKBOOK_NAME}]{sheet}".replace("'", "''")
    address = (
        f"{column}{first_row}"
        if first_row == last_row
        else f"{column}{first_row}:{column}{last_row}"
    )
    return f"='{source}'!{address}"


# ── One dataset ──────────────────────────────────────────────────────────────

def dataset_plan(segment_sheet: str, period_length: int) -> List[Tuple[str, str, str]]:
    """``(dataset type, worksheet, value column)`` for the four datasets of one class."""

    return [
        (CUTOFF_DATASET, CUTOFF_SHEETS[period_length], CUTOFF_VALUE_COLUMN),
        *((name, segment_sheet, column) for name, column in GROWTH_DATASETS),
    ]


def label_column_for(sheet: str) -> str:
    return CUTOFF_LABEL_COLUMN if sheet in CUTOFF_SHEETS.values() else GROWTH_LABEL_COLUMN


def read_column_values(app: App, book_path: Path, sheet: str, column: str, first_row: int, last_row: int) -> List[Any]:
    """The stored value of every linked cell, read by the app's own workbook reader."""

    items = [
        {"book_path": str(book_path), "sheet": sheet, "cell": f"{column}{row}"}
        for row in range(first_row, last_row + 1)
    ]
    results = app.post("/excel/read_cells_batch", {"items": items}).get("results") or []
    values: List[Any] = []
    for item, result in zip(items, results):
        if not isinstance(result, dict) or not result.get("ok"):
            reason = (result or {}).get("error") or "the cell could not be read"
            raise RunError(f"'{sheet}'!{item['cell']}: {reason}")
        values.append(result.get("value"))
    if len(values) != len(items):
        raise RunError(f"'{sheet}' returned {len(values)} of {len(items)} linked cells.")
    return values


def usable_prefix_count(values: Sequence[Any]) -> int:
    """How many leading values are populated before the first blank.

    ArcRho's Excel Link reader checks a link's whole target list at once: one
    blank cell inside the range (a future period the workbook has not reached
    yet) makes every cell in that link report as unreadable, not just the
    blank one. Linking only the populated prefix avoids that.
    """

    count = 0
    for value in values:
        if value is None:
            break
        count += 1
    return count


def delete_dataset(app: App, project: str, reserving_class: str, dataset_type: str) -> None:
    """Delete one dataset so ``--overwrite`` can recreate it from scratch.

    Raises the same ``RunError`` a failed create would, carrying the HTTP
    status so a 409 (something still depends on it) can be told apart from a
    real failure.
    """

    wait_until_writable(app, project, reserving_class)
    app.post(
        "/datasets/cached/delete",
        {"project_name": project, "reserving_class": reserving_class, "dataset_names": [dataset_type]},
    )


def create_dataset(
    app: App,
    project: str,
    reserving_class: str,
    dataset_type: str,
    period_length: int,
    number_format: Dict[str, Any],
    source: SourceWorkbook,
    sheet: str,
    column: str,
) -> str:
    """Create one empty vector, link every cell to the workbook, and save it."""

    wait_until_writable(app, project, reserving_class)
    app.post(
        "/datasets/cached/empty",
        {
            "project_name": project,
            "reserving_class": reserving_class,
            "dataset_type": dataset_type,
            "instance_name": dataset_type,
            "data_format": "Vector",
            "origin_length": period_length,
            "development_length": 12,
        },
    )

    # The sidecar read carries the metadata; the cached read carries the
    # geometry, and only it knows which periods the new vector holds.
    loaded = app.post(
        "/dataset/sidecar/load",
        {"project_name": project, "reserving_class": reserving_class, "dataset_name": dataset_type},
    )
    cached = app.post(
        "/dataset/cache/load",
        {
            "project_name": project,
            "reserving_class": reserving_class,
            "dataset_name": dataset_type,
            "csv_file": str(loaded.get("csv_file") or ""),
            "origin_length": period_length,
        },
    )
    origin_labels = [str(label) for label in cached.get("origin_labels") or []]
    if not origin_labels:
        raise RunError(f"{dataset_type} was created without origin labels.")

    first_row, last_row = source.row_span(sheet, label_column_for(sheet), origin_labels)
    values = read_column_values(app, source.path, sheet, column, first_row, last_row)
    linked_count = usable_prefix_count(values)
    external_links = []
    if linked_count:
        formula = excel_reference(source.folder, sheet, column, first_row, first_row + linked_count - 1)
        external_links = [
            {
                "reference": formula,
                "target_cells": [
                    {"row": index, "column": 0, "source_cell": f"{column}{first_row + index}"}
                    for index in range(linked_count)
                ],
            }
        ]
    else:
        formula = f"(no data yet in '{sheet}' column {column}; created with blank values only)"

    wait_until_writable(app, project, reserving_class)
    app.post(
        "/dataset/sidecar/save",
        {
            "project_name": project,
            "reserving_class": reserving_class,
            "dataset_name": dataset_type,
            "dataset_type": dataset_type,
            "source_kind": "input",
            "data_format": "Vector",
            "origin_length": period_length,
            "development_length": int(loaded.get("development_length") or 12),
            "transposed": bool(loaded.get("transposed")),
            "show_subtotal": bool(loaded.get("show_subtotal")),
            "number_format": str(number_format.get("number_format") or ""),
            "decimal_places": int(number_format.get("decimal_places") or 1),
            "origin_labels": origin_labels,
            "csv_file": str(loaded.get("csv_file") or ""),
            "method_type": "None",
            "external_links": external_links,
            "values": [[value] for value in values],
            "mask": [[value is not None] for value in values],
        },
    )
    return formula


# ── One dataset that is already there ────────────────────────────────────────

def _number_text(value: Any) -> str:
    return "blank" if value is None else repr(value)


def _same_number(left: Any, right: Any) -> bool:
    if left is None or right is None:
        return left is None and right is None
    return math.isclose(float(left), float(right), rel_tol=1e-12, abs_tol=0.0)


def validate_dataset(
    app: App,
    project: str,
    reserving_class: str,
    dataset_type: str,
    period_length: int,
    source: SourceWorkbook,
    sheet: str,
    column: str,
) -> List[str]:
    """Compare one dataset that already exists against the workbook.

    Reads only. Returns one line per disagreement between what the dataset
    holds and what this script would have written, so an instance somebody
    built by hand or linked in an earlier quarter can be judged without
    touching it.
    """

    findings: List[str] = []
    loaded = app.post(
        "/dataset/sidecar/load",
        {"project_name": project, "reserving_class": reserving_class, "dataset_name": dataset_type},
    )

    data_format = str(loaded.get("data_format") or "")
    if data_format.casefold() != "vector":
        findings.append(f"data format is {data_format or 'unset'}, expected Vector")
    stored_period = int(loaded.get("origin_length") or 0)
    if stored_period != period_length:
        findings.append(f"period length is {stored_period} months, expected {period_length}")

    cached = app.post(
        "/dataset/cache/load",
        {
            "project_name": project,
            "reserving_class": reserving_class,
            "dataset_name": dataset_type,
            "csv_file": str(loaded.get("csv_file") or ""),
            "origin_length": stored_period or period_length,
        },
    )
    origin_labels = [str(label) for label in cached.get("origin_labels") or []]
    if not origin_labels:
        return findings + ["the dataset holds no periods"]

    try:
        first_row, last_row = source.row_span(sheet, label_column_for(sheet), origin_labels)
    except RunError as err:
        return findings + [f"cannot be checked against the workbook: {err}"]

    expected_values = read_column_values(app, source.path, sheet, column, first_row, last_row)
    linked_count = usable_prefix_count(expected_values)
    if linked_count:
        expected_reference = excel_reference(source.folder, sheet, column, first_row, first_row + linked_count - 1)
        findings.extend(_link_findings(loaded, expected_reference, linked_count))

    stored_rows = cached.get("values") or []
    stored_values = [row[0] if isinstance(row, list) and row else None for row in stored_rows]
    if len(stored_values) != len(expected_values):
        findings.append(
            f"holds {len(stored_values)} periods, the workbook range covers {len(expected_values)}"
        )
    else:
        differences = [
            f"{origin_labels[index]} {_number_text(stored)} vs {_number_text(expected)}"
            for index, (stored, expected) in enumerate(zip(stored_values, expected_values))
            if not _same_number(stored, expected)
        ]
        if differences:
            shown = ", ".join(differences[:4])
            more = f", and {len(differences) - 4} more" if len(differences) > 4 else ""
            findings.append(
                f"{len(differences)} of {len(stored_values)} values differ "
                f"(dataset vs workbook): {shown}{more}"
            )
    return findings


def _link_findings(loaded: Dict[str, Any], expected_reference: str, period_count: int) -> List[str]:
    """What the dataset's links say, measured against the one link it should have."""

    formula_links = [link for link in loaded.get("formula_links") or [] if isinstance(link, dict)]
    external_links = [link for link in loaded.get("external_links") or [] if isinstance(link, dict)]
    internal_links = [link for link in loaded.get("internal_links") or [] if isinstance(link, dict)]

    findings: List[str] = []
    if internal_links:
        findings.append(f"has {len(internal_links)} link(s) to another ArcRho dataset")
    if not formula_links:
        if len(external_links) == 1 and str(external_links[0].get("reference") or "") == expected_reference:
            targets = external_links[0].get("target_cells") or []
            if len(targets) == period_count:
                # Same workbook range, just entered as an Excel Link instead of
                # an in-cell formula. That is not a discrepancy worth flagging.
                return findings
        if external_links:
            references = "; ".join(str(link.get("reference") or "") for link in external_links)
            findings.append(
                "uses an Excel Link rather than an in-cell formula: " + references
            )
            findings.append(f"expected the in-cell formula {expected_reference}")
        else:
            findings.append(f"has no link; expected the in-cell formula {expected_reference}")
        return findings

    if len(formula_links) > 1:
        findings.append(f"has {len(formula_links)} in-cell formulas where one covers the vector")
    link = formula_links[0]
    formula = str(link.get("formula") or "")
    if formula != expected_reference:
        findings.append(f"is linked to {formula}")
        findings.append(f"expected {expected_reference}")
    targets = link.get("target_cells") or []
    if len(targets) != period_count:
        findings.append(f"the formula fills {len(targets)} cells of {period_count}")
    return findings


# ── The run ──────────────────────────────────────────────────────────────────

def number_formats(app: App, dataset_types: Sequence[str]) -> Dict[str, Dict[str, Any]]:
    formats: Dict[str, Dict[str, Any]] = {}
    for dataset_type in dataset_types:
        answer = app.get("/dataset/number-format-defaults", {"dataset_type_name": dataset_type})
        number_format = NUMBER_FORMAT or str(answer.get("resolved_number_format") or "")
        formats[dataset_type] = {
            "number_format": number_format,
            "decimal_places": number_format_decimal_places(number_format),
        }
    return formats


def require_dataset_types(app: App, project: str, dataset_types: Sequence[str]) -> None:
    data = app.get("/dataset_types", {"project_name": project}).get("data") or {}
    defined = {str(row[0]).strip() for row in data.get("rows") or [] if row}
    missing = [name for name in dataset_types if name not in defined]
    if missing:
        raise RunError(
            f"'{project}' has no Dataset Type named {', '.join(missing)}. Add the type(s) in "
            "Project Settings first; creating one rewrites every instance in the project, so "
            "this script will not do it."
        )


def select_classes(
    classes: List[Tuple[str, str, int]],
    wanted_classes: Sequence[str],
) -> List[Tuple[str, str, int]]:
    """Narrow the run to named reserving classes, or keep all of them."""

    if not wanted_classes:
        return classes
    by_path = {path.casefold(): (segment, path, length) for segment, path, length in classes}
    chosen: List[Tuple[str, str, int]] = []
    for wanted in wanted_classes:
        found = by_path.get(wanted.strip().casefold())
        if found is None:
            listed = "\n  ".join(path for _segment, path, _length in classes)
            raise RunError(
                f"'{wanted}' is not one of the reserving classes in "
                f"{RESERVING_CLASS_WORKBOOK}:\n  {listed}"
            )
        chosen.append(found)
    return chosen


def run(project: str, folder: Path, wanted_classes: Sequence[str], dry_run: bool, overwrite: bool) -> int:
    app = App(ArcRhoUI().base_url)
    health = app.get("/app/health")
    if not health.get("ok"):
        raise RunError(f"ArcRho at {app.base_url} did not report a healthy app server.")

    source = SourceWorkbook(folder)
    classes = select_classes(read_reserving_classes(RESERVING_CLASS_WORKBOOK), wanted_classes)
    wanted = [CUTOFF_DATASET, *(name for name, _column in GROWTH_DATASETS)]
    require_dataset_types(app, project, wanted)
    formats = number_formats(app, wanted)

    print(f"Project          {project}")
    print(f"Source workbook  {source.path}")
    print(f"Reserving classes {len(classes)} from {RESERVING_CLASS_WORKBOOK}")
    print("")

    created = 0
    checked = 0
    flagged = 0
    overwritten = 0
    skipped = 0
    failures: List[str] = []
    for segment, reserving_class, period_length in classes:
        print(f"{segment} - {reserving_class}")
        try:
            segment_sheet = source.segment_sheet(segment)
            existing = app.get(
                "/datasets/cached",
                {"project_name": project, "reserving_class": reserving_class},
            )
            # Creating a dataset overwrites one of the same name without
            # asking, and a sidecar filename is case-insensitive on Windows,
            # so a name that differs only in case must still count as present.
            present = {
                str(item.get("name") or "").strip().casefold()
                for item in existing.get("files") or []
            }
        except RunError as err:
            failures.append(f"{segment}: {err}")
            print(f"  ! {err}")
            continue

        for dataset_type, sheet, column in dataset_plan(segment_sheet, period_length):
            if dataset_type.casefold() in present and not overwrite:
                # Never overwritten. Checked instead, and reported.
                checked += 1
                try:
                    findings = validate_dataset(
                        app,
                        project,
                        reserving_class,
                        dataset_type,
                        period_length,
                        source,
                        sheet,
                        column,
                    )
                except RunError as err:
                    failures.append(f"{segment} / {dataset_type}: {err}")
                    print(f"  ! {dataset_type} could not be checked: {err}")
                    continue
                if findings:
                    flagged += 1
                    print(f"  ~ {dataset_type} already there, {len(findings)} finding(s)")
                    for finding in findings:
                        print(f"      {finding}")
                else:
                    print(f"  = {dataset_type} already there and matches the workbook")
                continue
            if dataset_type.casefold() in present and overwrite:
                if dry_run:
                    print(f"  ~ {dataset_type} would be deleted and recreated, linked to '{sheet}' column {column}")
                    overwritten += 1
                    continue
                try:
                    delete_dataset(app, project, reserving_class, dataset_type)
                except RunError as err:
                    if err.http_status == 409:
                        skipped += 1
                        print(f"  ~ {dataset_type} left as-is, something still depends on it: {err}")
                    else:
                        failures.append(f"{segment} / {dataset_type}: {err}")
                        print(f"  ! {dataset_type} could not be deleted: {err}")
                    continue
                try:
                    formula = create_dataset(
                        app,
                        project,
                        reserving_class,
                        dataset_type,
                        period_length,
                        formats[dataset_type],
                        source,
                        sheet,
                        column,
                    )
                    print(f"  ~ {dataset_type} recreated, linked to {formula}")
                    overwritten += 1
                except RunError as err:
                    failures.append(f"{segment} / {dataset_type}: {err}")
                    print(f"  ! {dataset_type} deleted but could not be recreated: {err}")
                continue
            try:
                if dry_run:
                    print(f"  + {dataset_type} would link to '{sheet}' column {column}")
                else:
                    formula = create_dataset(
                        app,
                        project,
                        reserving_class,
                        dataset_type,
                        period_length,
                        formats[dataset_type],
                        source,
                        sheet,
                        column,
                    )
                    print(f"  + {dataset_type} linked to {formula}")
                created += 1
            except RunError as err:
                failures.append(f"{segment} / {dataset_type}: {err}")
                print(f"  ! {dataset_type}: {err}")

    print("")
    verb = "would create" if dry_run else "created"
    overwrite_verb = "would overwrite" if dry_run else "overwrote"
    print(
        f"{verb} {created}, checked {checked} already there ({flagged} with findings), "
        f"{overwrite_verb} {overwritten}, skipped {skipped} with dependents, {len(failures)} failed"
    )
    for failure in failures:
        print(f"  ! {failure}")
    return 1 if failures else 0


def main(argv: List[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--project", default=TARGET_PROJECT, help="ArcRho project to create the datasets in.")
    parser.add_argument(
        "--quarter",
        default=SOURCE_QUARTER_FOLDER,
        help="Source folder under the reserve review root, such as 2026Q2. Defaults to this quarter.",
    )
    parser.add_argument(
        "--reserving-class",
        action="append",
        default=[],
        metavar="PATH",
        help="Visit only this reserving class, as ResQ Path.xlsx writes it. Repeatable.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Report what would be created and write nothing.")
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help=(
            "Delete and recreate a dataset that already exists instead of only checking it. "
            "A dataset something else still depends on is reported and left alone."
        ),
    )
    args = parser.parse_args(argv)

    quarter = str(args.quarter or "").strip() or quarter_folder_for(dt.date.today())
    try:
        return run(
            str(args.project).strip(),
            Path(SOURCE_ROOT) / quarter,
            list(args.reserving_class),
            args.dry_run,
            args.overwrite,
        )
    except RunError as err:
        print(f"Stopped: {err}")
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
