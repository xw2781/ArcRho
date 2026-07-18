"""Convert legacy Reserving Class EEX Formula values into processing rules.

This is an explicit, one-time project migration. It intentionally refuses to
merge with or overwrite an existing ``data_processing_rules.json`` file.

Examples:

    python tools/migrate_eex_formulas.py --project-path "E:\\ArcRho Server\\projects\\Example" --dry-run
    python tools/migrate_eex_formulas.py --project-path "E:\\ArcRho Server\\projects\\Example" --apply
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Mapping, MutableMapping, Sequence, Tuple

import openpyxl


LEGACY_COLUMNS: Tuple[str, ...] = (
    "Name",
    "Level",
    "Formula",
    "EEX Formula",
    "Source",
)
MIGRATED_COLUMNS: Tuple[str, ...] = (
    "Name",
    "Level",
    "Formula",
    "Source",
)
RESERVING_CLASS_SHEET = "Reserving Class Types"
RULES_FORMAT = "arcrho-data-processing-rules-v1"
RULES_FILENAME = "data_processing_rules.json"
RESERVING_CLASS_JSON_FILENAME = "reserving_class_types.json"
RESERVING_CLASS_XLSX_FILENAME = "reserving_class_types.xlsx"
FIELD_MAPPING_FILENAME = "field_mapping.json"
TARGET_SOURCE_MEASURE = "Earned_Exposure"
MIGRATION_USER = "eex-formula-migration"


class MigrationError(RuntimeError):
    """Raised when a project cannot be converted without ambiguity or data loss."""


@dataclass(frozen=True)
class LegacyReservingClassRow:
    name: str
    level: int
    formula: str
    eex_formula: str
    source: str

    def migrated_values(self) -> List[str]:
        return [self.name, str(self.level), self.formula, self.source]


@dataclass(frozen=True)
class ParsedExpression:
    coefficients: Mapping[str, int]
    spellings: Mapping[str, str]
    reference_order: Sequence[str]


@dataclass(frozen=True)
class MigrationPlan:
    project_path: Path
    reserving_class_json_path: Path
    reserving_class_xlsx_path: Path
    field_mapping_path: Path
    rules_path: Path
    original_json_bytes: bytes
    original_xlsx_bytes: bytes
    migrated_reserving_class_payload: Mapping[str, Any]
    rules_document: Mapping[str, Any]
    legacy_rows: Sequence[LegacyReservingClassRow]
    level_fields: Mapping[int, str]

    @property
    def migrated_rule_count(self) -> int:
        rules = self.rules_document.get("rules", [])
        return len(rules) if isinstance(rules, list) else 0

    @property
    def non_empty_eex_count(self) -> int:
        return sum(1 for row in self.legacy_rows if row.eex_formula)


@dataclass(frozen=True)
class _Token:
    kind: str
    value: str
    position: int


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def _canonical_name(value: Any) -> str:
    return " ".join(str(value if value is not None else "").strip().split()).casefold()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_level(value: Any, *, label: str) -> int:
    text = _cell_text(value)
    try:
        number = int(text)
    except (TypeError, ValueError):
        try:
            as_float = float(text)
        except (TypeError, ValueError):
            raise MigrationError(f"{label} must be an integer level; found {text!r}.")
        if not as_float.is_integer():
            raise MigrationError(f"{label} must be an integer level; found {text!r}.")
        number = int(as_float)
    if number < 1:
        raise MigrationError(f"{label} must be at least 1; found {number}.")
    return number


def _load_json_object(path: Path, *, label: str) -> Tuple[Dict[str, Any], bytes]:
    try:
        raw_bytes = path.read_bytes()
    except FileNotFoundError:
        raise MigrationError(f"{label} was not found: {path}")
    except OSError as error:
        raise MigrationError(f"Failed to read {label} at {path}: {error}")

    try:
        payload = json.loads(raw_bytes.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as error:
        raise MigrationError(f"{label} is not valid UTF-8 JSON: {error}")
    if not isinstance(payload, dict):
        raise MigrationError(f"{label} must contain a JSON object.")
    return payload, raw_bytes


def _normalize_legacy_rows(payload: Mapping[str, Any]) -> List[LegacyReservingClassRow]:
    columns = payload.get("columns")
    if not isinstance(columns, list):
        raise MigrationError("reserving_class_types.json is missing a columns list.")
    normalized_columns = [_cell_text(value) for value in columns]
    if normalized_columns != list(LEGACY_COLUMNS):
        if "EEX Formula" not in normalized_columns:
            raise MigrationError(
                "reserving_class_types.json is not a legacy EEX file: "
                f"expected columns {list(LEGACY_COLUMNS)}, found {normalized_columns}."
            )
        raise MigrationError(
            "Unsupported reserving_class_types.json column order. "
            f"Expected {list(LEGACY_COLUMNS)}, found {normalized_columns}."
        )

    raw_rows = payload.get("rows")
    if not isinstance(raw_rows, list):
        raise MigrationError("reserving_class_types.json is missing a rows list.")

    rows: List[LegacyReservingClassRow] = []
    seen: Dict[Tuple[int, str], str] = {}
    for row_number, raw_row in enumerate(raw_rows, start=2):
        if not isinstance(raw_row, list):
            raise MigrationError(
                f"reserving_class_types.json row {row_number} must be an array."
            )
        if len(raw_row) != len(LEGACY_COLUMNS):
            raise MigrationError(
                f"reserving_class_types.json row {row_number} has {len(raw_row)} cells; "
                f"expected {len(LEGACY_COLUMNS)}."
            )
        values = [_cell_text(value) for value in raw_row]
        if not any(values):
            raise MigrationError(
                f"reserving_class_types.json row {row_number} is blank; "
                "remove blank rows before migration."
            )
        name = values[0]
        if not name:
            raise MigrationError(
                f"reserving_class_types.json row {row_number} has a blank Name."
            )
        level = _parse_level(values[1], label=f"Reserving Class Type {name!r} Level")
        key = (level, _canonical_name(name))
        previous = seen.get(key)
        if previous is not None:
            raise MigrationError(
                f"Ambiguous Reserving Class Type {name!r} at level {level}; "
                f"it duplicates {previous!r}."
            )
        seen[key] = name
        rows.append(
            LegacyReservingClassRow(
                name=name,
                level=level,
                formula=values[2],
                eex_formula=values[3],
                source=values[4],
            )
        )
    return rows


def _read_xlsx_rows(path: Path, *, expected_columns: Sequence[str]) -> List[List[str]]:
    try:
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    except FileNotFoundError:
        raise MigrationError(f"reserving_class_types.xlsx was not found: {path}")
    except Exception as error:
        raise MigrationError(f"Failed to read reserving_class_types.xlsx: {error}")

    try:
        if RESERVING_CLASS_SHEET not in workbook.sheetnames:
            raise MigrationError(
                f"reserving_class_types.xlsx is missing sheet {RESERVING_CLASS_SHEET!r}."
            )
        worksheet = workbook[RESERVING_CLASS_SHEET]
        header = [
            _cell_text(worksheet.cell(row=1, column=index).value)
            for index in range(1, int(worksheet.max_column or 1) + 1)
        ]
        while header and not header[-1]:
            header.pop()
        if header != list(expected_columns):
            raise MigrationError(
                "reserving_class_types.xlsx header does not match the expected contract. "
                f"Expected {list(expected_columns)}, found {header}."
            )

        rows: List[List[str]] = []
        width = len(expected_columns)
        for row_number in range(2, int(worksheet.max_row or 1) + 1):
            values = [
                _cell_text(worksheet.cell(row=row_number, column=column).value)
                for column in range(1, width + 1)
            ]
            if any(values):
                rows.append(values)
        return rows
    finally:
        workbook.close()


def _legacy_row_values(rows: Sequence[LegacyReservingClassRow]) -> List[List[str]]:
    return [
        [
            row.name,
            str(row.level),
            row.formula,
            row.eex_formula,
            row.source,
        ]
        for row in rows
    ]


def _load_level_fields(path: Path) -> Dict[int, str]:
    payload, _raw_bytes = _load_json_object(path, label="field_mapping.json")
    raw_rows = payload.get("rows")
    if not isinstance(raw_rows, list):
        raise MigrationError("field_mapping.json is missing a rows list.")

    level_fields: Dict[int, str] = {}
    dataset_fields: set[str] = set()
    for row_number, raw_row in enumerate(raw_rows, start=1):
        if not isinstance(raw_row, dict):
            raise MigrationError(f"field_mapping.json row {row_number} must be an object.")
        field_name = _cell_text(raw_row.get("field_name"))
        significance = _cell_text(raw_row.get("significance"))
        if significance == "Dataset" and field_name:
            dataset_fields.add(_canonical_name(field_name))
        if significance != "Reserving Class":
            continue
        if not field_name:
            raise MigrationError(
                f"field_mapping.json Reserving Class row {row_number} has a blank field_name."
            )
        level = _parse_level(
            raw_row.get("level"),
            label=f"Field Mapping row {row_number} level",
        )
        previous = level_fields.get(level)
        if previous is not None:
            raise MigrationError(
                f"Field Mapping level {level} is ambiguous: {previous!r} and {field_name!r}."
            )
        level_fields[level] = field_name

    if _canonical_name(TARGET_SOURCE_MEASURE) not in dataset_fields:
        raise MigrationError(
            f"field_mapping.json does not map {TARGET_SOURCE_MEASURE!r} as a Dataset field; "
            "the generated rules would fail target validation."
        )
    return level_fields


def _tokenize_additive_expression(expression: str, *, label: str) -> List[_Token]:
    text = str(expression or "").strip()
    if not text:
        raise MigrationError(f"{label} is blank.")

    tokens: List[_Token] = []
    index = 0
    while index < len(text):
        char = text[index]
        if char.isspace():
            index += 1
            continue
        if char in "+-()":
            kind = {
                "+": "PLUS",
                "-": "MINUS",
                "(": "LPAREN",
                ")": "RPAREN",
            }[char]
            tokens.append(_Token(kind=kind, value=char, position=index))
            index += 1
            continue
        if char in "*/":
            raise MigrationError(
                f"{label} uses unsupported operator {char!r} at position {index + 1}; "
                "only additive membership expressions can be migrated."
            )
        if char == '"':
            start = index
            index += 1
            value_chars: List[str] = []
            while index < len(text):
                current = text[index]
                if current == "\\" and index + 1 < len(text):
                    escaped = text[index + 1]
                    if escaped in {'"', "\\"}:
                        value_chars.append(escaped)
                        index += 2
                        continue
                if current == '"':
                    index += 1
                    break
                value_chars.append(current)
                index += 1
            else:
                raise MigrationError(
                    f"{label} has an unterminated quoted name at position {start + 1}."
                )
            value = "".join(value_chars).strip()
            if not value:
                raise MigrationError(f"{label} contains an empty quoted member.")
            tokens.append(_Token(kind="NAME", value=value, position=start))
            continue

        start = index
        while index < len(text) and text[index] not in '+-*/()"':
            index += 1
        value = text[start:index].strip()
        if not value:
            if index < len(text) and text[index] == '"':
                raise MigrationError(
                    f"{label} has an unexpected quote at position {index + 1}."
                )
            raise MigrationError(f"{label} contains an invalid token at position {start + 1}.")
        tokens.append(_Token(kind="NAME", value=value, position=start))

    return tokens


class _AdditiveExpressionParser:
    def __init__(self, tokens: Sequence[_Token], *, label: str):
        self.tokens = list(tokens)
        self.label = label
        self.index = 0
        self.spellings: Dict[str, str] = {}
        self.reference_order: List[str] = []

    def parse(self) -> ParsedExpression:
        coefficients = self._parse_expression()
        if self.index != len(self.tokens):
            token = self.tokens[self.index]
            raise MigrationError(
                f"{self.label} has an unexpected {token.value!r} at position "
                f"{token.position + 1}."
            )
        nonzero = {
            key: coefficient
            for key, coefficient in coefficients.items()
            if coefficient != 0
        }
        return ParsedExpression(
            coefficients=nonzero,
            spellings=dict(self.spellings),
            reference_order=tuple(self.reference_order),
        )

    def _parse_expression(self) -> Dict[str, int]:
        result = self._parse_signed_primary()
        while self.index < len(self.tokens):
            token = self.tokens[self.index]
            if token.kind not in {"PLUS", "MINUS"}:
                break
            self.index += 1
            right = self._parse_signed_primary()
            multiplier = 1 if token.kind == "PLUS" else -1
            self._merge(result, right, multiplier)
        return result

    def _parse_signed_primary(self) -> Dict[str, int]:
        sign = 1
        while self.index < len(self.tokens) and self.tokens[self.index].kind in {
            "PLUS",
            "MINUS",
        }:
            if self.tokens[self.index].kind == "MINUS":
                sign *= -1
            self.index += 1

        if self.index >= len(self.tokens):
            raise MigrationError(f"{self.label} ends with an operator.")
        token = self.tokens[self.index]
        if token.kind == "NAME":
            self.index += 1
            key = _canonical_name(token.value)
            if not key:
                raise MigrationError(
                    f"{self.label} contains a blank member at position {token.position + 1}."
                )
            self.spellings.setdefault(key, token.value)
            self.reference_order.append(key)
            return {key: sign}
        if token.kind == "LPAREN":
            self.index += 1
            nested = self._parse_expression()
            if self.index >= len(self.tokens) or self.tokens[self.index].kind != "RPAREN":
                raise MigrationError(
                    f"{self.label} has an unmatched '(' at position {token.position + 1}."
                )
            self.index += 1
            if sign == -1:
                nested = {key: -value for key, value in nested.items()}
            return nested
        raise MigrationError(
            f"{self.label} has an unexpected {token.value!r} at position "
            f"{token.position + 1}."
        )

    @staticmethod
    def _merge(target: MutableMapping[str, int], source: Mapping[str, int], sign: int) -> None:
        for key, coefficient in source.items():
            target[key] = target.get(key, 0) + sign * coefficient


def _parse_additive_expression(expression: str, *, label: str) -> ParsedExpression:
    tokens = _tokenize_additive_expression(expression, label=label)
    return _AdditiveExpressionParser(tokens, label=label).parse()


def _validate_coefficients(expression: ParsedExpression, *, label: str) -> None:
    unsupported = [
        (
            expression.spellings.get(key, key),
            coefficient,
        )
        for key, coefficient in expression.coefficients.items()
        if coefficient not in {-1, 1}
    ]
    if unsupported:
        rendered = ", ".join(f"{name}={coefficient}" for name, coefficient in unsupported)
        raise MigrationError(
            f"{label} resolves to unsupported repeated coefficients ({rendered}); "
            "resolve duplicate or cancelling members before migration."
        )


def _row_source_expression(row: LegacyReservingClassRow) -> ParsedExpression:
    expression = row.source or f'"{row.name}"'
    parsed = _parse_additive_expression(
        expression,
        label=f"Source for {row.name!r} at level {row.level}",
    )
    _validate_coefficients(
        parsed,
        label=f"Source for {row.name!r} at level {row.level}",
    )
    if not parsed.coefficients:
        raise MigrationError(
            f"Source for {row.name!r} at level {row.level} resolves to no members."
        )
    return parsed


def _slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_value.casefold()).strip("-")
    return slug or "type"


def _allocate_rule_id(row: LegacyReservingClassRow, used_ids: set[str]) -> str:
    base = f"migrated-eex-l{row.level}-{_slug(row.name)}"
    if len(base) > 100:
        digest = hashlib.sha256(base.encode("utf-8")).hexdigest()[:10]
        base = f"{base[:89].rstrip('-')}-{digest}"
    candidate = base
    counter = 2
    while candidate in used_ids:
        candidate = f"{base}-{counter}"
        counter += 1
    used_ids.add(candidate)
    return candidate


def _expand_eex_members(
    row: LegacyReservingClassRow,
    *,
    rows_by_level_name: Mapping[Tuple[int, str], LegacyReservingClassRow],
) -> List[str]:
    normal_source = _row_source_expression(row)
    normal_spellings = {
        key: normal_source.spellings.get(key, key)
        for key in normal_source.coefficients
    }
    eex = _parse_additive_expression(
        row.eex_formula,
        label=f"EEX Formula for {row.name!r} at level {row.level}",
    )

    members: List[str] = []
    seen_members: set[str] = set()
    for reference_key in eex.reference_order:
        referenced_row = rows_by_level_name.get((row.level, reference_key))
        if referenced_row is not None:
            referenced_source = _row_source_expression(referenced_row)
            expanded_keys = [
                key
                for key in referenced_source.reference_order
                if key in referenced_source.coefficients
            ]
        elif reference_key in normal_source.coefficients:
            expanded_keys = [reference_key]
        else:
            spelling = eex.spellings.get(reference_key, reference_key)
            raise MigrationError(
                f"EEX Formula for {row.name!r} at level {row.level} references "
                f"unknown member {spelling!r}."
            )

        for member_key in expanded_keys:
            if member_key not in normal_source.coefficients:
                member_name = (
                    normal_spellings.get(member_key)
                    or eex.spellings.get(member_key)
                    or member_key
                )
                raise MigrationError(
                    f"EEX Formula for {row.name!r} at level {row.level} keeps "
                    f"{member_name!r}, but that member is not in the row's normal Source."
                )
            if member_key in seen_members:
                continue
            seen_members.add(member_key)
            members.append(normal_spellings[member_key])

    if not members:
        raise MigrationError(
            f"EEX Formula for {row.name!r} at level {row.level} resolves to no members."
        )
    return members


def _build_rules(
    rows: Sequence[LegacyReservingClassRow],
    *,
    level_fields: Mapping[int, str],
) -> List[Dict[str, Any]]:
    rows_by_level_name: Dict[Tuple[int, str], LegacyReservingClassRow] = {
        (row.level, _canonical_name(row.name)): row for row in rows
    }
    used_ids: set[str] = set()
    rules: List[Dict[str, Any]] = []

    for row in rows:
        if not row.eex_formula:
            continue
        field_name = level_fields.get(row.level)
        if not field_name:
            raise MigrationError(
                f"EEX Formula for {row.name!r} uses level {row.level}, but "
                "Field Mapping has no Reserving Class field at that level."
            )
        members = _expand_eex_members(
            row,
            rows_by_level_name=rows_by_level_name,
        )
        rules.append(
            {
                "id": _allocate_rule_id(row, used_ids),
                "name": f"Migrated EEX - {row.name}",
                "enabled": True,
                "target": {
                    "source_measure": TARGET_SOURCE_MEASURE,
                },
                "request_conditions": {
                    "all": [
                        {
                            "field": field_name,
                            "level": row.level,
                            "operator": "equals",
                            "value": row.name,
                        }
                    ]
                },
                "row_conditions": {
                    "all": [],
                },
                "action": {
                    "type": "keep_members",
                    "field": field_name,
                    "level": row.level,
                    "members": members,
                },
            }
        )
    return rules


def _json_bytes(payload: Mapping[str, Any]) -> bytes:
    return (
        json.dumps(payload, indent=2, ensure_ascii=False)
        + "\n"
    ).encode("utf-8")


def build_migration_plan(project_path: os.PathLike[str] | str) -> MigrationPlan:
    project = Path(project_path).expanduser().resolve()
    if not project.is_dir():
        raise MigrationError(f"Project path is not a directory: {project}")

    reserving_json_path = project / RESERVING_CLASS_JSON_FILENAME
    reserving_xlsx_path = project / RESERVING_CLASS_XLSX_FILENAME
    field_mapping_path = project / FIELD_MAPPING_FILENAME
    rules_path = project / RULES_FILENAME
    if rules_path.exists():
        raise MigrationError(
            f"Refusing to overwrite existing {RULES_FILENAME}: {rules_path}"
        )

    payload, original_json_bytes = _load_json_object(
        reserving_json_path,
        label=RESERVING_CLASS_JSON_FILENAME,
    )
    rows = _normalize_legacy_rows(payload)
    try:
        original_xlsx_bytes = reserving_xlsx_path.read_bytes()
    except FileNotFoundError:
        raise MigrationError(f"{RESERVING_CLASS_XLSX_FILENAME} was not found: {reserving_xlsx_path}")
    except OSError as error:
        raise MigrationError(f"Failed to read {RESERVING_CLASS_XLSX_FILENAME}: {error}")

    xlsx_rows = _read_xlsx_rows(
        reserving_xlsx_path,
        expected_columns=LEGACY_COLUMNS,
    )
    json_rows = _legacy_row_values(rows)
    if xlsx_rows != json_rows:
        raise MigrationError(
            "The reserving-class JSON/XLSX pair is inconsistent. "
            "Re-save or reconcile the pair before migration."
        )

    level_fields = _load_level_fields(field_mapping_path)
    rules = _build_rules(rows, level_fields=level_fields)
    migrated_at = _utc_now()

    migrated_payload = dict(payload)
    migrated_payload["columns"] = list(MIGRATED_COLUMNS)
    migrated_payload["rows"] = [row.migrated_values() for row in rows]
    migrated_payload["updated_at"] = migrated_at

    rules_document = {
        "json_format": RULES_FORMAT,
        "revision": 1,
        "updated_at": migrated_at,
        "updated_by": MIGRATION_USER,
        "rules": rules,
    }

    return MigrationPlan(
        project_path=project,
        reserving_class_json_path=reserving_json_path,
        reserving_class_xlsx_path=reserving_xlsx_path,
        field_mapping_path=field_mapping_path,
        rules_path=rules_path,
        original_json_bytes=original_json_bytes,
        original_xlsx_bytes=original_xlsx_bytes,
        migrated_reserving_class_payload=migrated_payload,
        rules_document=rules_document,
        legacy_rows=tuple(rows),
        level_fields=dict(level_fields),
    )


def _migration_artifact_paths(plan: MigrationPlan) -> Dict[str, Path]:
    return {
        "json_stage": plan.reserving_class_json_path.with_name(
            plan.reserving_class_json_path.name + ".eex-migration.tmp"
        ),
        "xlsx_stage": plan.reserving_class_xlsx_path.with_name(
            plan.reserving_class_xlsx_path.name + ".eex-migration.tmp"
        ),
        "rules_stage": plan.rules_path.with_name(
            plan.rules_path.name + ".eex-migration.tmp"
        ),
        "json_backup": plan.reserving_class_json_path.with_name(
            plan.reserving_class_json_path.name + ".eex-migration.rollback"
        ),
        "xlsx_backup": plan.reserving_class_xlsx_path.with_name(
            plan.reserving_class_xlsx_path.name + ".eex-migration.rollback"
        ),
        "lock": plan.project_path / ".eex-formula-migration.lock",
    }


def _write_new_file(path: Path, content: bytes) -> None:
    with path.open("xb") as handle:
        handle.write(content)
        handle.flush()
        os.fsync(handle.fileno())


def _build_migrated_xlsx(plan: MigrationPlan, output_path: Path) -> None:
    try:
        workbook = openpyxl.load_workbook(
            plan.reserving_class_xlsx_path,
            data_only=False,
        )
    except Exception as error:
        raise MigrationError(f"Failed to open the reserving-class workbook for migration: {error}")

    try:
        worksheet = workbook[RESERVING_CLASS_SHEET]
        eex_column = LEGACY_COLUMNS.index("EEX Formula") + 1
        worksheet.delete_cols(eex_column, 1)

        for column, header in enumerate(MIGRATED_COLUMNS, start=1):
            worksheet.cell(row=1, column=column, value=header)
        migrated_rows = plan.migrated_reserving_class_payload.get("rows", [])
        for row_number, values in enumerate(migrated_rows, start=2):
            for column, value in enumerate(values, start=1):
                worksheet.cell(row=row_number, column=column, value=_cell_text(value))

        last_output_row = len(migrated_rows) + 1
        for row_number in range(last_output_row + 1, int(worksheet.max_row or 1) + 1):
            for column in range(1, len(MIGRATED_COLUMNS) + 1):
                worksheet.cell(row=row_number, column=column, value=None)

        workbook.save(output_path)
    except Exception as error:
        raise MigrationError(f"Failed to stage the migrated reserving-class workbook: {error}")
    finally:
        workbook.close()


def _sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _assert_originals_unchanged(plan: MigrationPlan) -> None:
    try:
        current_json = plan.reserving_class_json_path.read_bytes()
        current_xlsx = plan.reserving_class_xlsx_path.read_bytes()
    except OSError as error:
        raise MigrationError(f"Failed to recheck source files before migration: {error}")
    if _sha256_bytes(current_json) != _sha256_bytes(plan.original_json_bytes):
        raise MigrationError(
            f"{RESERVING_CLASS_JSON_FILENAME} changed after validation; no files were migrated."
        )
    if _sha256_bytes(current_xlsx) != _sha256_bytes(plan.original_xlsx_bytes):
        raise MigrationError(
            f"{RESERVING_CLASS_XLSX_FILENAME} changed after validation; no files were migrated."
        )
    if plan.rules_path.exists():
        raise MigrationError(
            f"{RULES_FILENAME} appeared after validation; refusing to overwrite it."
        )


def _validate_migrated_outputs(plan: MigrationPlan) -> None:
    payload, _raw_bytes = _load_json_object(
        plan.reserving_class_json_path,
        label=RESERVING_CLASS_JSON_FILENAME,
    )
    columns = [_cell_text(value) for value in payload.get("columns", [])]
    if columns != list(MIGRATED_COLUMNS):
        raise MigrationError(
            f"Migrated {RESERVING_CLASS_JSON_FILENAME} has invalid columns: {columns}."
        )
    expected_rows = plan.migrated_reserving_class_payload.get("rows", [])
    actual_rows = [
        [_cell_text(value) for value in row]
        for row in payload.get("rows", [])
        if isinstance(row, list)
    ]
    if actual_rows != expected_rows:
        raise MigrationError(
            f"Migrated {RESERVING_CLASS_JSON_FILENAME} rows failed verification."
        )
    workbook_rows = _read_xlsx_rows(
        plan.reserving_class_xlsx_path,
        expected_columns=MIGRATED_COLUMNS,
    )
    if workbook_rows != expected_rows:
        raise MigrationError(
            f"Migrated {RESERVING_CLASS_XLSX_FILENAME} rows failed verification."
        )
    rules_payload, _rules_bytes = _load_json_object(
        plan.rules_path,
        label=RULES_FILENAME,
    )
    if rules_payload != plan.rules_document:
        raise MigrationError(f"Migrated {RULES_FILENAME} failed verification.")


def apply_migration(plan: MigrationPlan) -> None:
    artifacts = _migration_artifact_paths(plan)
    conflicts = [path for path in artifacts.values() if path.exists()]
    if conflicts:
        rendered = ", ".join(str(path) for path in conflicts)
        raise MigrationError(
            "Refusing to start while migration staging, rollback, or lock files exist: "
            f"{rendered}"
        )

    rules_created = False
    lock_created = False
    try:
        _write_new_file(artifacts["lock"], b"ArcRho EEX Formula migration in progress.\n")
        lock_created = True
        _assert_originals_unchanged(plan)

        _write_new_file(
            artifacts["json_stage"],
            _json_bytes(plan.migrated_reserving_class_payload),
        )
        _build_migrated_xlsx(plan, artifacts["xlsx_stage"])
        _write_new_file(
            artifacts["rules_stage"],
            _json_bytes(plan.rules_document),
        )
        _write_new_file(artifacts["json_backup"], plan.original_json_bytes)
        _write_new_file(artifacts["xlsx_backup"], plan.original_xlsx_bytes)

        _assert_originals_unchanged(plan)
        os.replace(artifacts["json_stage"], plan.reserving_class_json_path)
        os.replace(artifacts["xlsx_stage"], plan.reserving_class_xlsx_path)
        try:
            os.link(artifacts["rules_stage"], plan.rules_path)
        except FileExistsError:
            raise MigrationError(
                f"{RULES_FILENAME} appeared during migration; refusing to overwrite it."
            )
        except OSError as error:
            raise MigrationError(
                f"Failed to create {RULES_FILENAME} without overwriting: {error}"
            )
        rules_created = True
        artifacts["rules_stage"].unlink()

        _validate_migrated_outputs(plan)
        artifacts["json_backup"].unlink()
        artifacts["xlsx_backup"].unlink()
    except Exception as error:
        rollback_errors: List[str] = []
        try:
            if artifacts["json_backup"].exists():
                os.replace(artifacts["json_backup"], plan.reserving_class_json_path)
        except Exception as rollback_error:
            rollback_errors.append(
                f"failed to restore {RESERVING_CLASS_JSON_FILENAME}: {rollback_error}"
            )
        try:
            if artifacts["xlsx_backup"].exists():
                os.replace(artifacts["xlsx_backup"], plan.reserving_class_xlsx_path)
        except Exception as rollback_error:
            rollback_errors.append(
                f"failed to restore {RESERVING_CLASS_XLSX_FILENAME}: {rollback_error}"
            )
        if rules_created:
            try:
                current_rules = plan.rules_path.read_bytes()
                expected_rules = _json_bytes(plan.rules_document)
                if _sha256_bytes(current_rules) == _sha256_bytes(expected_rules):
                    plan.rules_path.unlink()
                else:
                    rollback_errors.append(
                        f"did not remove {RULES_FILENAME} because it changed after creation"
                    )
            except Exception as rollback_error:
                rollback_errors.append(
                    f"failed to remove newly created {RULES_FILENAME}: {rollback_error}"
                )
        message = str(error)
        if rollback_errors:
            message += " Rollback errors: " + "; ".join(rollback_errors)
        if isinstance(error, MigrationError):
            raise MigrationError(message)
        raise MigrationError(f"Migration failed: {message}")
    finally:
        for key in ("json_stage", "xlsx_stage", "rules_stage"):
            try:
                artifacts[key].unlink(missing_ok=True)
            except OSError:
                pass
        if lock_created:
            try:
                artifacts["lock"].unlink(missing_ok=True)
            except OSError:
                pass


def migration_summary(plan: MigrationPlan, *, mode: str) -> Dict[str, Any]:
    rules = plan.rules_document.get("rules", [])
    rule_summaries = []
    for rule in rules if isinstance(rules, list) else []:
        request_all = rule.get("request_conditions", {}).get("all", [])
        request = request_all[0] if request_all else {}
        rule_summaries.append(
            {
                "id": rule.get("id", ""),
                "reserving_class_type": request.get("value", ""),
                "field": request.get("field", ""),
                "level": request.get("level"),
                "members": rule.get("action", {}).get("members", []),
            }
        )
    return {
        "ok": True,
        "mode": mode,
        "project_path": str(plan.project_path),
        "legacy_eex_rows": plan.non_empty_eex_count,
        "rules_created": plan.migrated_rule_count,
        "outputs": {
            "reserving_class_types_json": str(plan.reserving_class_json_path),
            "reserving_class_types_xlsx": str(plan.reserving_class_xlsx_path),
            "data_processing_rules_json": str(plan.rules_path),
        },
        "rules": rule_summaries,
    }


def _parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Convert a project's legacy EEX Formula column into "
            "data_processing_rules.json."
        )
    )
    parser.add_argument(
        "--project-path",
        required=True,
        help="Explicit path to one ArcRho project directory.",
    )
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate and print the proposed conversion without writing files.",
    )
    mode.add_argument(
        "--apply",
        action="store_true",
        help="Apply the validated migration transaction.",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    try:
        plan = build_migration_plan(args.project_path)
        if args.apply:
            apply_migration(plan)
        summary = migration_summary(
            plan,
            mode="applied" if args.apply else "dry-run",
        )
        print(json.dumps(summary, indent=2, ensure_ascii=False))
        return 0
    except MigrationError as error:
        print(f"Migration refused: {error}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
