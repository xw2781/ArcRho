from __future__ import annotations

import importlib.util
import io
import json
import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from unittest.mock import patch

import openpyxl


REPO_ROOT = Path(__file__).resolve().parents[2]
MODULE_PATH = REPO_ROOT / "tools" / "migrate_eex_formulas.py"
SPEC = importlib.util.spec_from_file_location("migrate_eex_formulas", MODULE_PATH)
assert SPEC is not None and SPEC.loader is not None
migration = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = migration
SPEC.loader.exec_module(migration)


LEGACY_ROWS = [
    ["BI", "5", "", "", '"BI"'],
    ["UMBI", "5", "", "", '"UMBI"'],
    ["BIR51", "5", "", "", '"BIR51"'],
    ["PD", "5", "", "", '"PD"'],
    ["CMP_CAT", "5", "", "", '"CMP_CAT"'],
    [
        "BI Total",
        "5",
        '"BI" + "UMBI" + "BIR51"',
        '"BI"',
        '"BI" + "UMBI" + "BIR51"',
    ],
    [
        "TOTAL PA",
        "5",
        '"BI" + "PD" - "CMP_CAT"',
        '"PD" + "CMP_CAT"',
        '"BI" + "PD" - "CMP_CAT"',
    ],
]


def _write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def _write_legacy_workbook(path: Path, rows: list[list[str]]) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = migration.RESERVING_CLASS_SHEET
    worksheet.append(list(migration.LEGACY_COLUMNS))
    for row in rows:
        worksheet.append(row)
    extra = workbook.create_sheet("Preserved")
    extra["A1"] = "keep me"
    workbook.save(path)
    workbook.close()


def _create_project(root: Path, *, rows: list[list[str]] | None = None) -> Path:
    project = root / "ExampleProject"
    project.mkdir()
    legacy_rows = rows if rows is not None else LEGACY_ROWS
    _write_json(
        project / migration.RESERVING_CLASS_JSON_FILENAME,
        {
            "columns": list(migration.LEGACY_COLUMNS),
            "rows": legacy_rows,
            "updated_at": "2026-01-01T00:00:00Z",
        },
    )
    _write_legacy_workbook(
        project / migration.RESERVING_CLASS_XLSX_FILENAME,
        legacy_rows,
    )
    _write_json(
        project / migration.FIELD_MAPPING_FILENAME,
        {
            "project_name": "ExampleProject",
            "rows": [
                {
                    "field_name": "IBNRCAT",
                    "significance": "Reserving Class",
                    "dataset_type": None,
                    "level": 5,
                },
                {
                    "field_name": "Earned_Exposure",
                    "significance": "Dataset",
                    "dataset_type": "Earned Exposure",
                    "level": None,
                },
            ],
        },
    )
    return project


class EexFormulaMigrationTests(unittest.TestCase):
    def test_main_dry_run_reports_success_without_writing(self) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT) as temp_dir:
            project = _create_project(Path(temp_dir))
            output = io.StringIO()

            with redirect_stdout(output):
                exit_code = migration.main(
                    ["--project-path", str(project), "--dry-run"]
                )

            self.assertEqual(0, exit_code)
            report = json.loads(output.getvalue())
            self.assertEqual("dry-run", report["mode"])
            self.assertEqual(2, report["rules_created"])
            self.assertFalse((project / migration.RULES_FILENAME).exists())

    def test_dry_run_builds_rules_without_writing(self) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT) as temp_dir:
            project = _create_project(Path(temp_dir))
            json_path = project / migration.RESERVING_CLASS_JSON_FILENAME
            xlsx_path = project / migration.RESERVING_CLASS_XLSX_FILENAME
            original_json = json_path.read_bytes()
            original_xlsx = xlsx_path.read_bytes()

            plan = migration.build_migration_plan(project)
            summary = migration.migration_summary(plan, mode="dry-run")

            self.assertEqual(2, plan.migrated_rule_count)
            self.assertEqual(2, summary["legacy_eex_rows"])
            self.assertEqual(
                ["BI"],
                plan.rules_document["rules"][0]["action"]["members"],
            )
            self.assertEqual(
                ["PD", "CMP_CAT"],
                plan.rules_document["rules"][1]["action"]["members"],
            )
            self.assertEqual(original_json, json_path.read_bytes())
            self.assertEqual(original_xlsx, xlsx_path.read_bytes())
            self.assertFalse((project / migration.RULES_FILENAME).exists())

    def test_apply_migrates_json_xlsx_and_creates_rules(self) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT) as temp_dir:
            project = _create_project(Path(temp_dir))
            plan = migration.build_migration_plan(project)

            migration.apply_migration(plan)

            reserving_payload = json.loads(
                (project / migration.RESERVING_CLASS_JSON_FILENAME).read_text(
                    encoding="utf-8"
                )
            )
            self.assertEqual(
                list(migration.MIGRATED_COLUMNS),
                reserving_payload["columns"],
            )
            self.assertTrue(
                all(len(row) == len(migration.MIGRATED_COLUMNS) for row in reserving_payload["rows"])
            )

            workbook = openpyxl.load_workbook(
                project / migration.RESERVING_CLASS_XLSX_FILENAME,
                data_only=False,
            )
            try:
                worksheet = workbook[migration.RESERVING_CLASS_SHEET]
                headers = [
                    worksheet.cell(row=1, column=index).value
                    for index in range(1, len(migration.MIGRATED_COLUMNS) + 1)
                ]
                self.assertEqual(list(migration.MIGRATED_COLUMNS), headers)
                self.assertEqual("keep me", workbook["Preserved"]["A1"].value)
            finally:
                workbook.close()

            rules_payload = json.loads(
                (project / migration.RULES_FILENAME).read_text(encoding="utf-8")
            )
            self.assertEqual(migration.RULES_FORMAT, rules_payload["json_format"])
            self.assertEqual(1, rules_payload["revision"])
            self.assertEqual("IBNRCAT", rules_payload["rules"][0]["action"]["field"])
            self.assertFalse(
                any(path.name.endswith(".eex-migration.rollback") for path in project.iterdir())
            )

    def test_refuses_existing_rules_file(self) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT) as temp_dir:
            project = _create_project(Path(temp_dir))
            existing = project / migration.RULES_FILENAME
            existing.write_text('{"rules":[]}\n', encoding="utf-8")

            with self.assertRaisesRegex(migration.MigrationError, "Refusing to overwrite"):
                migration.build_migration_plan(project)

            self.assertEqual('{"rules":[]}\n', existing.read_text(encoding="utf-8"))

    def test_refuses_inconsistent_json_xlsx_pair(self) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT) as temp_dir:
            project = _create_project(Path(temp_dir))
            workbook_path = project / migration.RESERVING_CLASS_XLSX_FILENAME
            workbook = openpyxl.load_workbook(workbook_path)
            workbook[migration.RESERVING_CLASS_SHEET]["A2"] = "Different"
            workbook.save(workbook_path)
            workbook.close()
            original_json = (
                project / migration.RESERVING_CLASS_JSON_FILENAME
            ).read_bytes()

            with self.assertRaisesRegex(migration.MigrationError, "inconsistent"):
                migration.build_migration_plan(project)

            self.assertEqual(
                original_json,
                (project / migration.RESERVING_CLASS_JSON_FILENAME).read_bytes(),
            )
            self.assertFalse((project / migration.RULES_FILENAME).exists())

    def test_refuses_eex_member_outside_normal_source(self) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT) as temp_dir:
            rows = [list(row) for row in LEGACY_ROWS]
            rows[-1][3] = '"UMBI"'
            project = _create_project(Path(temp_dir), rows=rows)

            with self.assertRaisesRegex(migration.MigrationError, "not in the row's normal Source"):
                migration.build_migration_plan(project)

    def test_refuses_unsupported_multiplication(self) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT) as temp_dir:
            rows = [list(row) for row in LEGACY_ROWS]
            rows[-1][3] = '"PD" * "CMP_CAT"'
            project = _create_project(Path(temp_dir), rows=rows)

            with self.assertRaisesRegex(migration.MigrationError, "unsupported operator"):
                migration.build_migration_plan(project)

    def test_rolls_back_pair_when_rules_creation_fails(self) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT) as temp_dir:
            project = _create_project(Path(temp_dir))
            json_path = project / migration.RESERVING_CLASS_JSON_FILENAME
            xlsx_path = project / migration.RESERVING_CLASS_XLSX_FILENAME
            original_json = json_path.read_bytes()
            original_xlsx = xlsx_path.read_bytes()
            plan = migration.build_migration_plan(project)

            with patch.object(migration.os, "link", side_effect=OSError("simulated failure")):
                with self.assertRaisesRegex(migration.MigrationError, "without overwriting"):
                    migration.apply_migration(plan)

            self.assertEqual(original_json, json_path.read_bytes())
            self.assertEqual(original_xlsx, xlsx_path.read_bytes())
            self.assertFalse((project / migration.RULES_FILENAME).exists())

    def test_refuses_ambiguous_field_mapping_level(self) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT) as temp_dir:
            project = _create_project(Path(temp_dir))
            mapping_path = project / migration.FIELD_MAPPING_FILENAME
            payload = json.loads(mapping_path.read_text(encoding="utf-8"))
            payload["rows"].append(
                {
                    "field_name": "OTHER_RC",
                    "significance": "Reserving Class",
                    "dataset_type": None,
                    "level": 5,
                }
            )
            _write_json(mapping_path, payload)

            with self.assertRaisesRegex(migration.MigrationError, "ambiguous"):
                migration.build_migration_plan(project)


if __name__ == "__main__":
    unittest.main()
