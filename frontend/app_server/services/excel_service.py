"""Workbook reads (openpyxl) and the one COM interop action (open in Excel).

Cell values and file stats are plain file reads of the workbook — no Excel
installation is involved — so they run wherever the workbook is reachable.
Only ``excel_open_workbook`` drives a desktop Excel through win32com.
"""
from __future__ import annotations

import os
import stat
import zipfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any, Dict, List, Mapping
from xml.etree import ElementTree

import openpyxl


EXCEL_BATCH_MAX_WORKERS = 4
# The OOXML package part that carries a workbook's own Created, Modified, and
# Last saved by properties, and the namespaces its elements are written in.
_CORE_PROPERTIES_ENTRY = "docProps/core.xml"
_CORE_PROPERTIES_NS = {
    "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
    "dcterms": "http://purl.org/dc/terms/",
}
# The part holds a handful of short strings; a larger one is a malformed or
# hostile package rather than something worth parsing.
_CORE_PROPERTIES_MAX_BYTES = 64 * 1024


def _item_field(item: Any, name: str) -> Any:
    """Read one field of a batch item, which is a router model or a plain dict."""

    if isinstance(item, Mapping):
        return item.get(name)
    return getattr(item, name, None)


def excel_workbook_readable(book_path: str) -> Dict[str, Any]:
    """Report whether this process can open the workbook as an Excel file.

    Answers for the machine it runs on: on ArcRho Server this is the truth
    the Excel Link Manager reports and the retarget requires, and a path the
    server cannot reach is refused rather than guessed at.
    """

    book = Path(str(book_path or "").strip())
    if not str(book):
        return {"ok": False, "error": "Workbook path is empty."}
    try:
        if not book.is_file():
            return {"ok": False, "error": "The workbook was not found."}
        wb = openpyxl.load_workbook(str(book), data_only=True, read_only=True)
        wb.close()
    except PermissionError:
        return {"ok": False, "error": "The workbook is not readable (permission denied)."}
    except OSError as exc:
        return {"ok": False, "error": f"The workbook could not be opened: {exc.strerror or exc}"}
    except Exception as exc:  # openpyxl raises its own zip/format errors
        return {"ok": False, "error": f"The workbook could not be opened as an Excel file: {exc}"}
    return {"ok": True}


def excel_read_cell(book_path: str, sheet: str, cell: str) -> Dict[str, Any]:
    book = Path(book_path).resolve()
    if not book.exists():
        return {"ok": False, "error": f"File not found: {book_path}"}
    try:
        wb = openpyxl.load_workbook(str(book), data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return {"ok": False, "error": f"Sheet not found: {sheet}"}
        ws = wb[sheet]
        cell_value = ws[cell].value
        wb.close()
        numeric = None
        if cell_value is not None:
            try:
                numeric = float(cell_value)
            except (ValueError, TypeError):
                return {"ok": False, "error": f"Cell value is not numeric: {repr(cell_value)}"}
        return {"ok": True, "value": numeric}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def excel_read_cells_batch(items: list) -> Dict[str, Any]:
    groups: Dict[str, Dict[str, Any]] = {}
    result_keys: List[tuple[str, str, str]] = []
    for item in items:
        resolved = str(Path(str(_item_field(item, "book_path") or "")).resolve())
        book_key = os.path.normcase(resolved)
        sheet = str(_item_field(item, "sheet") or "")
        cell = str(_item_field(item, "cell") or "").upper()
        cell_key = (book_key, sheet, cell)
        result_keys.append(cell_key)
        group = groups.setdefault(book_key, {"path": resolved, "items": {}})
        group["items"].setdefault(cell_key, {"sheet": sheet, "cell": cell})

    def read_workbook(group: Dict[str, Any]) -> Dict[tuple[str, str, str], Dict[str, Any]]:
        book_path = str(group["path"])
        unique_items = group["items"]
        workbook_results: Dict[tuple[str, str, str], Dict[str, Any]] = {}
        p = Path(book_path).resolve()
        if not p.exists():
            return {
                key: {"ok": False, "error": f"File not found: {book_path}"}
                for key in unique_items
            }
        try:
            wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
            try:
                for key, item in unique_items.items():
                    if item["sheet"] not in wb.sheetnames:
                        workbook_results[key] = {"ok": False, "error": f"Sheet not found: {item['sheet']}"}
                        continue
                    val = wb[item["sheet"]][item["cell"]].value
                    try:
                        numeric = float(val) if val is not None else None
                        workbook_results[key] = {"ok": True, "value": numeric}
                    except (ValueError, TypeError):
                        workbook_results[key] = {"ok": False, "error": f"Not numeric: {repr(val)}"}
            finally:
                wb.close()
        except Exception as e:
            for key in unique_items:
                workbook_results.setdefault(key, {"ok": False, "error": str(e)})
        return workbook_results

    by_key: Dict[tuple[str, str, str], Dict[str, Any]] = {}
    if groups:
        with ThreadPoolExecutor(
            max_workers=min(EXCEL_BATCH_MAX_WORKERS, len(groups)),
            thread_name_prefix="arcrho-excel-check",
        ) as executor:
            futures = [executor.submit(read_workbook, group) for group in groups.values()]
            for future in futures:
                by_key.update(future.result())
    results = [dict(by_key[key]) for key in result_keys]
    return {"ok": True, "results": results}


def _run_workbook_path_batch(
    book_paths: list[str],
    read_one: Any,
    thread_name_prefix: str,
) -> Dict[str, Any]:
    """Run ``read_one`` once per distinct workbook path on a bounded pool.

    Callers pass the same path more than once - two datasets reading one
    workbook, a workbook listed under two aliases - so the paths are resolved
    and deduplicated before any I/O and the shared answer is copied back into
    every caller's slot. The pool keeps a listing off a network drive from
    paying one awaited round trip per file.
    """

    resolved_by_key: Dict[str, str] = {}
    result_keys: List[str] = []
    for index, raw_path in enumerate(book_paths):
        if not str(raw_path or "").strip():
            key = f"__invalid_path_{index}"
            result_keys.append(key)
            resolved_by_key[key] = ""
            continue
        resolved = str(Path(str(raw_path or "")).resolve())
        key = os.path.normcase(resolved)
        result_keys.append(key)
        resolved_by_key.setdefault(key, resolved)

    def read_item(item: tuple[str, str]) -> tuple[str, Dict[str, Any]]:
        key, resolved = item
        if not resolved:
            return key, {"ok": False, "path": resolved, "error": "Workbook path is empty."}
        return key, read_one(resolved)

    by_key: Dict[str, Dict[str, Any]] = {}
    if resolved_by_key:
        with ThreadPoolExecutor(
            max_workers=min(EXCEL_BATCH_MAX_WORKERS, len(resolved_by_key)),
            thread_name_prefix=thread_name_prefix,
        ) as executor:
            futures = [executor.submit(read_item, item) for item in resolved_by_key.items()]
            for future in futures:
                key, result = future.result()
                by_key[key] = result
    return {"ok": True, "results": [dict(by_key[key]) for key in result_keys]}


def _stat_workbook(resolved: str) -> Dict[str, Any]:
    try:
        stat_result = os.stat(resolved)
        if not stat.S_ISREG(stat_result.st_mode):
            return {"ok": False, "path": resolved, "error": f"File not found: {resolved}"}
        return {"ok": True, "path": resolved, "mtime": stat_result.st_mtime}
    except OSError as exc:
        return {"ok": False, "path": resolved, "error": str(exc)}


def _workbook_document_properties(resolved: str) -> Dict[str, str]:
    """Read the workbook's own Created/Modified/Last saved by properties.

    These are the document's record of itself - the same kind of answer a
    dataset sidecar gives about a dataset - rather than the file system's,
    so they survive a copy or a move that resets a file's creation time.
    They live in ``docProps/core.xml`` of the OOXML package, so a legacy
    ``.xls``, an encrypted workbook, and anything else that is not a readable
    zip simply have none; that is reported as blank, never as an error, because
    the workbook is still a perfectly good link target.
    """

    try:
        with zipfile.ZipFile(resolved) as package:
            with package.open(_CORE_PROPERTIES_ENTRY) as entry:
                raw = entry.read(_CORE_PROPERTIES_MAX_BYTES)
    except (KeyError, OSError, ValueError, zipfile.BadZipFile):
        return {}
    try:
        root = ElementTree.fromstring(raw)
    except ElementTree.ParseError:
        return {}

    def value(tag: str) -> str:
        found = root.find(tag, _CORE_PROPERTIES_NS)
        return str(found.text or "").strip() if found is not None else ""

    return {
        "created": value("dcterms:created"),
        "modified": value("dcterms:modified"),
        "last_modified_by": value("cp:lastModifiedBy"),
    }


def _stat_and_describe_workbook(resolved: str) -> Dict[str, Any]:
    result = _stat_workbook(resolved)
    if result.get("ok"):
        result.update(_workbook_document_properties(resolved))
    return result


def excel_file_mtimes_batch(book_paths: list[str]) -> Dict[str, Any]:
    return _run_workbook_path_batch(book_paths, _stat_workbook, "arcrho-excel-stat")


def excel_workbook_properties_batch(book_paths: list[str]) -> Dict[str, Any]:
    """``excel_file_mtimes_batch`` plus each workbook's document properties.

    One pass, not two: the stat and the small ``docProps`` read for a path
    happen in the same worker, so a listing over a network drive does not walk
    the same file list twice.
    """

    return _run_workbook_path_batch(
        book_paths, _stat_and_describe_workbook, "arcrho-excel-props"
    )


def excel_open_workbook(book_path: str, sheet: str = "", cell: str = "") -> Dict[str, Any]:
    p = Path(book_path).resolve()
    if not p.exists():
        return {"ok": False, "error": f"File not found: {book_path}"}
    try:
        import pythoncom
        import win32com.client
    except ImportError:
        return {"ok": False, "error": "win32com is not available on this system."}
    try:
        pythoncom.CoInitialize()
        try:
            xl = win32com.client.GetObject(Class="Excel.Application")
        except Exception:
            xl = win32com.client.Dispatch("Excel.Application")
            xl.Visible = True
        full_path = str(p)
        target_wb = None
        already_open = False
        for wb in xl.Workbooks:
            if wb.FullName.lower() == full_path.lower():
                target_wb = wb
                already_open = True
                break
        if target_wb is None:
            target_wb = xl.Workbooks.Open(str(p), ReadOnly=True)
        xl.Visible = True
        try:
            import win32gui
            hwnd = xl.Hwnd
            import win32con
            if win32gui.IsIconic(hwnd):
                win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            win32gui.SetForegroundWindow(hwnd)
        except Exception:
            pass
        target_wb.Activate()
        if sheet and cell:
            try:
                ws = target_wb.Sheets(sheet)
                target = ws.Range(cell)
                scroll_row = max(1, target.Row - 10)
                scroll_col = max(1, target.Column - 10)
                ws.Activate()
                xl.Goto(ws.Cells(scroll_row, scroll_col), True)
                target.Select()
            except Exception:
                pass
        elif sheet:
            try:
                target_wb.Sheets(sheet).Activate()
            except Exception:
                pass
        return {"ok": True, "already_open": already_open}
    except Exception as e:
        return {"ok": False, "error": f"Failed to open workbook: {str(e)}"}
