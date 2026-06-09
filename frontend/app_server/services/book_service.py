"""Workbook/Excel file operations via openpyxl."""
from __future__ import annotations

import os
from pathlib import Path

import openpyxl
from fastapi import HTTPException

from app_server import config


def resolve_allowed_book(path_str: str) -> Path:
    p = Path(path_str).resolve()
    for root in config.ALLOWED_BOOK_DIRS:
        if p == root or str(p).startswith(str(root) + os.sep):
            return p
    raise HTTPException(400, "Workbook path not allowed.")


def get_book_mtime(path: str) -> float:
    return os.stat(path).st_mtime


def read_sheet_matrix(path: str, sheet_name: str, max_rows: int = 200, max_cols: int = 50):
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise HTTPException(404, f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    max_r = min(ws.max_row or 1, max_rows)
    max_c = min(ws.max_column or 1, max_cols)
    rows = []
    for r in range(1, max_r + 1):
        row = []
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            row.append(v)
        rows.append(row)
    return rows
