from __future__ import annotations

import sys
import stat
import tempfile
import unittest
from concurrent.futures import Future
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from unittest import mock


REPO_ROOT = Path(__file__).resolve().parents[2]
FRONTEND_ROOT = REPO_ROOT / "frontend"
if str(FRONTEND_ROOT) not in sys.path:
    sys.path.insert(0, str(FRONTEND_ROOT))

import openpyxl

from app_server.services import excel_service


class _Workbook:
    sheetnames = ["Sheet1"]

    def __init__(self, value: float) -> None:
        self.value = value
        self.closed = False

    def __getitem__(self, key: str):
        if key == "Sheet1":
            return self
        return SimpleNamespace(value=self.value)

    def close(self) -> None:
        self.closed = True


class _RecordingExecutor:
    max_workers = 0

    def __init__(self, *, max_workers: int, thread_name_prefix: str) -> None:
        del thread_name_prefix
        type(self).max_workers = max_workers

    def __enter__(self):
        return self

    def __exit__(self, *_args) -> None:
        return None

    def submit(self, fn, *args):
        future = Future()
        future.set_result(fn(*args))
        return future


class ExcelBatchReadTests(unittest.TestCase):
    def work_dir(self) -> Path:
        """A scratch folder for the tests that need a workbook on disk.

        Created per test rather than in ``setUp`` so the tests that mock the
        file system away keep touching no files at all, and inside the
        repository so a validation run never writes outside it.
        """

        temp_dir = tempfile.TemporaryDirectory(dir=REPO_ROOT)
        self.addCleanup(temp_dir.cleanup)
        return Path(temp_dir.name)

    def item(self, book: str, cell: str = "A1"):
        return SimpleNamespace(book_path=book, sheet="Sheet1", cell=cell)

    def test_deduplicates_cells_opens_each_workbook_once_and_preserves_order(self) -> None:
        first = _Workbook(12.5)
        second = _Workbook(33.0)
        items = [
            self.item("first.xlsx"),
            self.item("second.xlsx"),
            self.item("first.xlsx"),
        ]
        with (
            mock.patch.object(Path, "exists", return_value=True),
            mock.patch.object(
                excel_service.openpyxl,
                "load_workbook",
                side_effect=[first, second],
            ) as load,
        ):
            result = excel_service.excel_read_cells_batch(items)

        self.assertTrue(result["ok"])
        self.assertEqual(result["results"], [
            {"ok": True, "value": 12.5},
            {"ok": True, "value": 33.0},
            {"ok": True, "value": 12.5},
        ])
        self.assertEqual(load.call_count, 2)
        self.assertTrue(first.closed)
        self.assertTrue(second.closed)

    def test_bounds_workbook_concurrency(self) -> None:
        items = [self.item(f"missing-{index}.xlsx") for index in range(8)]
        with mock.patch.object(excel_service, "ThreadPoolExecutor", _RecordingExecutor):
            result = excel_service.excel_read_cells_batch(items)

        self.assertEqual(_RecordingExecutor.max_workers, excel_service.EXCEL_BATCH_MAX_WORKERS)
        self.assertEqual(len(result["results"]), len(items))
        self.assertTrue(all(not item["ok"] for item in result["results"]))

    def test_file_mtimes_deduplicate_stat_calls_and_preserve_order(self) -> None:
        calls = []

        def fake_stat(path: str):
            calls.append(path)
            return SimpleNamespace(
                st_mtime=101.0 if "first" in path else 202.0,
                st_mode=stat.S_IFREG,
            )

        paths = ["first.xlsx", "second.xlsx", "first.xlsx", ""]
        with (
            mock.patch.object(excel_service.os, "stat", side_effect=fake_stat),
            mock.patch.object(excel_service, "ThreadPoolExecutor", _RecordingExecutor),
        ):
            result = excel_service.excel_file_mtimes_batch(paths)

        self.assertTrue(result["ok"])
        self.assertEqual([item.get("mtime") for item in result["results"][:3]], [101.0, 202.0, 101.0])
        self.assertFalse(result["results"][3]["ok"])
        self.assertEqual(len(calls), 2)
        self.assertEqual(_RecordingExecutor.max_workers, 3)

    def test_file_mtime_concurrency_is_bounded(self) -> None:
        paths = [f"workbook-{index}.xlsx" for index in range(8)]
        with (
            mock.patch.object(
                excel_service.os,
                "stat",
                return_value=SimpleNamespace(st_mtime=100.0, st_mode=stat.S_IFREG),
            ),
            mock.patch.object(excel_service, "ThreadPoolExecutor", _RecordingExecutor),
        ):
            result = excel_service.excel_file_mtimes_batch(paths)

        self.assertEqual(_RecordingExecutor.max_workers, excel_service.EXCEL_BATCH_MAX_WORKERS)
        self.assertEqual(len(result["results"]), len(paths))

    def test_workbook_properties_read_the_package_core_properties(self) -> None:
        # Created, Last Modified, and User are the workbook's own record of
        # itself - the workbook-side answer to the dataset table's columns -
        # so they come from docProps/core.xml, not from the file system.
        work = self.work_dir()
        book = work / "Book.xlsx"
        workbook = openpyxl.Workbook()
        workbook.properties.creator = "a.author"
        workbook.properties.lastModifiedBy = "j.tanaka"
        workbook.properties.created = datetime(2024, 3, 2, 9, 15, 41)
        workbook.save(book)

        # Not an OOXML package: a legacy .xls, an encrypted workbook, anything
        # that is not a readable zip. It is still a good link target, so it is
        # reported without properties rather than as an error.
        legacy = work / "Legacy.xls"
        legacy.write_bytes(bytes([0xd0, 0xcf, 0x11, 0xe0]) + b" not a zip")
        missing = work / "Gone.xlsx"

        result = excel_service.excel_workbook_properties_batch(
            [str(book), str(legacy), str(missing), "", str(book)]
        )

        good, plain, absent, blank, repeat = result["results"]
        self.assertTrue(good["ok"])
        self.assertEqual(good["last_modified_by"], "j.tanaka")
        self.assertTrue(good["created"].startswith("2024-03-02T09:15:41"))
        self.assertTrue(good["modified"])
        self.assertIsNotNone(good["mtime"])
        self.assertTrue(plain["ok"])
        self.assertNotIn("last_modified_by", plain)
        self.assertFalse(absent["ok"])
        self.assertFalse(blank["ok"])
        # One read per distinct path; the repeat carries the same answer.
        self.assertEqual(repeat, good)

    def test_file_mtimes_batch_keeps_its_own_result_shape(self) -> None:
        # The properties batch is the richer read; the mtimes batch every
        # freshness check uses must not start paying for a docProps read.
        work = self.work_dir()
        book = work / "Book.xlsx"
        openpyxl.Workbook().save(book)

        result = excel_service.excel_file_mtimes_batch([str(book)])

        self.assertEqual(sorted(result["results"][0].keys()), ["mtime", "ok", "path"])


if __name__ == "__main__":
    unittest.main()
