from __future__ import annotations

import os
from typing import Any, List, Dict, Optional, Tuple

import numpy as np
import pandas as pd
import openpyxl
import json
import threading
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel, Field

from fastapi.staticfiles import StaticFiles
from fastapi.responses import RedirectResponse
from pathlib import Path

import time
import hashlib
from datetime import datetime
from typing import TYPE_CHECKING

try:
    from watchdog.observers import Observer  # type: ignore
    from watchdog.events import FileSystemEventHandler  # type: ignore
except Exception:  # optional dependency
    Observer = None
    FileSystemEventHandler = None

# -----------------------------
# Config - Load from ui_config.json
# -----------------------------
UI_CONFIG_PATH = os.path.join(os.path.dirname(__file__), "ui_config.json")

def load_ui_config() -> Dict[str, Any]:
    """Load configuration from ui_config.json."""
    try:
        with open(UI_CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"root_path": r"E:\ADAS", "paths": {}}

def get_root_path() -> str:
    """Get the root path from config."""
    config = load_ui_config()
    return config.get("root_path", r"E:\ADAS")

def get_path(subpath: str) -> str:
    """Get a full path by joining root_path with subpath."""
    return os.path.join(get_root_path(), subpath)

# Derived paths (loaded dynamically from config)
def _get_data_dir() -> str:
    config = load_ui_config()
    return get_path(config.get("paths", {}).get("web_ui", "Web UI"))

def _get_team_profile_dir() -> str:
    config = load_ui_config()
    return get_path(config.get("paths", {}).get("team_profile", "Team Profile"))

def _get_virtual_projects_dir() -> str:
    config = load_ui_config()
    return get_path(config.get("paths", {}).get("virtual_projects", "Virtual Projects"))

def _get_data_base() -> str:
    config = load_ui_config()
    return get_path(config.get("paths", {}).get("data", "data"))

def _get_requests_dir() -> str:
    config = load_ui_config()
    return get_path(config.get("paths", {}).get("requests", "requests"))

# Legacy compatibility - these will use the dynamic functions
DATA_DIR = os.environ.get("TRI_DATA_DIR") or _get_data_dir()
DATASETS = {
    "paid_demo": os.path.join(_get_data_dir(), "triangle_paid.csv"),
}

PROJECT_BOOK = os.environ.get("ADAS_PROJECT_BOOK") or os.path.join(_get_team_profile_dir(), "Actuarial_NJ.xlsm")

# Project settings JSON files (on shared network drive)
PROJECT_SETTINGS_DIR = os.environ.get("ADAS_PROJECT_SETTINGS_DIR") or _get_team_profile_dir()
PROJECT_SETTINGS_SOURCES = {
    "Actuarial_NJ": "Actuarial_NJ.json",
    # Add more sources here as needed
}

WORKFLOW_DIR = os.environ.get(
    "ADAS_WORKFLOW_DIR",
    str(Path.home() / "Documents" / "ADAS" / "workflows"),
)
WORKFLOW_EXT = ".adaswf"

ALLOWED_BOOK_DIRS = [
    Path(_get_virtual_projects_dir()).resolve(),
    # add more if needed
]

# -----------------------------
# Helpers
# -----------------------------

REQUEST_DIR = os.environ.get("ADAS_REQUEST_DIR") or _get_requests_dir()
DATA_BASE = os.environ.get("ADAS_DATA_BASE") or _get_data_base()


# --- add in Helpers section ---
def _sanitize_folder_name(name: str) -> str:
    invalid = [":", "*", "?", '"', "<", ">", "|"]
    out = name or ""
    for ch in invalid:
        out = out.replace(ch, "_")
    return out

def _sanitize_filename(name: str) -> str:
    invalid = [":", "*", "?", '"', "<", ">", "|", "/", "\\", "\n", "\r", "\t"]
    out = (name or "").strip()
    for ch in invalid:
        out = out.replace(ch, "_")
    return out.strip() or "workflow"

def set_data_path_like_vba(pairs: list[tuple[str, str]]) -> str:
    """
    Re-implement your VBA SetDataPath behavior:
    - ProjectName is used as subfolder (sanitized)
    - fullName is '@'-joined values excluding ProjectName
    - then replace \ and / with ^; replace * with $star$
    """
    proj = ""
    values = []
    for k, v in pairs:
        if (k or "").strip().lower() == "projectname":
            proj = (v or "").strip()
        else:
            values.append((v or "").strip())

    full_name = "@".join(values)
    full_name = full_name.replace("\\", "^").replace("/", "^").replace("*", "$star$")

    base = DATA_BASE.rstrip("\\/") + "\\"
    if proj:
        proj = _sanitize_folder_name(proj)
        return f"{base}{proj}\\{full_name}.csv"
    return f"{base}{full_name}.csv"

def send_request_like_vba(request_info: str) -> str:
    """
    Re-implement your VBA SendRequest:
    - write temp .tmp then atomically publish to .txt
    - filename uses yyyy-mm-dd_hh-mm-ss.000 (ms)
    """
    os.makedirs(REQUEST_DIR, exist_ok=True)

    # VBA: Format(Now, "yyyy-mm-dd_hh-mm-ss") & Format(Timer - Int(Timer), ".000")
    now = datetime.now()
    ms = int(now.microsecond / 1000)
    current_time = now.strftime("%Y-%m-%d_%H-%M-%S") + f".{ms:03d}"

    temp_path = os.path.join(REQUEST_DIR, f"request-{current_time}.tmp")
    final_path = os.path.join(REQUEST_DIR, f"request-{current_time}.txt")

    lines = request_info.split("#")

    with open(temp_path, "w", encoding="utf-8", newline="\n") as f:
        for line in lines:
            f.write(line.rstrip("\r\n") + "\n")
        f.write(f"UserName = {os.environ.get('USERNAME', '')}\n")

    # overwrite protection (same as your VBA logic)
    if os.path.exists(final_path):
        try:
            os.remove(final_path)
        except OSError:
            try:
                os.remove(temp_path)
            except OSError:
                pass
            raise HTTPException(409, "Request file name collision and cannot overwrite.")

    os.replace(temp_path, final_path)
    return final_path

def wait_for_file(path: str, timeout_sec: float) -> bool:
    """
    Wait until file exists.
    Uses watchdog if available; falls back to polling otherwise.
    Assumes producer writes to .tmp then atomically renames.
    """
    # Fast path
    if os.path.exists(path):
        return True

    # Try watchdog first (better for network shares), else fallback to polling.
    try:
        if Observer is None or FileSystemEventHandler is None:
            raise RuntimeError("watchdog not available")
        target = Path(path)
        watch_dir = str(target.parent)
        target_name = target.name

        from threading import Event
        hit = Event()

        class _Handler(FileSystemEventHandler):
            def on_created(self, event) -> None:
                if os.path.basename(event.src_path) == target_name:
                    hit.set()

            def on_moved(self, event) -> None:
                if os.path.basename(event.dest_path) == target_name:
                    hit.set()

        handler = _Handler()
        observer = Observer()
        observer.schedule(handler, watch_dir, recursive=False)
        observer.start()
        try:
            hit.wait(timeout=max(0.0, float(timeout_sec)))
            return os.path.exists(path)
        finally:
            observer.stop()
            observer.join(timeout=1.0)
    except Exception:
        pass

    # Fallback polling
    t0 = time.time()
    while time.time() - t0 <= timeout_sec:
        if os.path.exists(path):
            return True
        time.sleep(0.5)
    return False


def resolve_allowed_book(path_str: str) -> Path:
    p = Path(path_str).resolve()
    for root in ALLOWED_BOOK_DIRS:
        if p == root or str(p).startswith(str(root) + os.sep):
            return p
    raise HTTPException(400, "Workbook path not allowed.")

def get_book_mtime(path: str) -> float:
    return os.stat(path).st_mtime

def read_sheet_matrix(path: str, sheet_name: str, max_rows: int = 200, max_cols: int = 50):
    """
    Return a rectangular matrix (list[list]) trimmed to max_rows/max_cols.
    Reads values only (not styles).
    """
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise HTTPException(404, f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    # find used range (simple heuristic)
    max_r = min(ws.max_row or 1, max_rows)
    max_c = min(ws.max_column or 1, max_cols)

    rows = []
    for r in range(1, max_r + 1):
        row = []
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            row.append(v)
        rows.append(row)
    return rows

class XlsmCellPatch(BaseModel):
    r: int = Field(..., ge=0)   # 0-based
    c: int = Field(..., ge=0)   # 0-based
    value: Any = None

class XlsmPatchRequest(BaseModel):
    sheet: str
    items: List[XlsmCellPatch]
    file_mtime: Optional[float] = None

def atomic_write_csv(df: pd.DataFrame, path: str) -> None:
    tmp = path + ".tmp"
    df.to_csv(tmp, index=False, header=False)
    os.replace(tmp, path)

def make_annual_labels(start_year: int, n_origin: int, n_dev: int) -> Tuple[List[str], List[str]]:
    origin_labels = [str(start_year + i) for i in range(n_origin)]
    dev_labels = [str(12 * (j + 1)) for j in range(n_dev)]
    return origin_labels, dev_labels

def infer_shape(path: str) -> Tuple[int, int]:
    df = pd.read_csv(path, header=None)
    return int(df.shape[0]), int(df.shape[1])

def load_triangle_values(path: str) -> pd.DataFrame:
    return pd.read_csv(path, header=None, dtype="float64")

def triangle_mask(n_origin: int, n_dev: int) -> np.ndarray:
    r = np.arange(n_origin)[:, None]
    c = np.arange(n_dev)[None, :]
    return (c <= r)

def diagonal_indices(n_origin: int, n_dev: int, k: int = 0) -> List[Tuple[int, int]]:
    mask = triangle_mask(n_origin, n_dev)
    out = []
    for r in range(n_origin):
        c = r - k
        if 0 <= c < n_dev and mask[r, c]:
            out.append((r, c))
    return out

# -----------------------------
# API models
# -----------------------------
class PatchItem(BaseModel):
    r: int = Field(..., ge=0)
    c: int = Field(..., ge=0)
    value: Optional[float] = None

class PatchRequest(BaseModel):
    items: List[PatchItem]
    file_mtime: Optional[float] = None

class AnyBookSheetRequest(BaseModel):
    book_path: str
    sheet: str

class AnyBookPatchRequest(BaseModel):
    book_path: str
    sheet: str
    items: List[XlsmCellPatch]
    file_mtime: Optional[float] = None

class WorkflowSaveRequest(BaseModel):
    name: str = ""
    prev_path: Optional[str] = None
    data: Dict[str, Any]

class WorkflowSaveAsRequest(BaseModel):
    path: str
    data: Dict[str, Any]

class WorkflowLoadRequest(BaseModel):
    path: str
# -----------------------------
# App
# -----------------------------
app = FastAPI(title="Triangle Demo API", version="0.1")

BASE_DIR = Path(__file__).resolve().parent
RESTART_FLAG = BASE_DIR / ".restart_app"
SHUTDOWN_FLAG = BASE_DIR / ".shutdown_app"
ELECTRON_RESTART_FLAG = BASE_DIR / ".restart_electron"
ELECTRON_SHUTDOWN_FLAG = BASE_DIR / ".shutdown_electron"

# ---- UI Config API ----

class UIConfigUpdateRequest(BaseModel):
    root_path: str


@app.get("/ui_config")
def get_ui_config() -> Dict[str, Any]:
    """Get the current UI configuration."""
    config = load_ui_config()
    return {"ok": True, "config": config}


@app.post("/ui_config")
def update_ui_config(req: UIConfigUpdateRequest) -> Dict[str, Any]:
    """Update the root path in ui_config.json."""
    config = load_ui_config()
    config["root_path"] = req.root_path
    try:
        with open(UI_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return {"ok": True, "config": config}
    except Exception as e:
        raise HTTPException(500, f"Failed to save config: {str(e)}")


# ---- API routes (keep BEFORE static mount) ----

class AdaTriRequest(BaseModel):
    Path: str
    TriangleName: str
    ProjectName: str
    Cumulative: bool = True
    OriginLength: int = 12
    DevelopmentLength: int = 12
    timeout_sec: float = 6.0


class AdaHeadersRequest(BaseModel):
    periodType: int = 0
    Transposed: bool = False
    PeriodLength: int = 12
    ProjectName: str
    StoredPeriodLength: int = -1
    timeout_sec: float = 6.0


@app.post("/adas/headers")
def adas_headers(req: AdaHeadersRequest) -> Dict[str, Any]:
    # Must match VBA ADASHeaders -> SetDataPath values order (excluding ProjectName)
    pairs = [
        ("Function", "ADASHeaders"),
        ("periodType", str(req.periodType)),
        ("Transposed", str(req.Transposed)),
        ("PeriodLength", str(req.PeriodLength)),
        ("ProjectName", req.ProjectName),
        ("StoredPeriodLength", str(req.StoredPeriodLength)),
    ]

    data_path = set_data_path_like_vba(pairs)
    request_file = None  # <-- add

    if not os.path.exists(data_path):
        request_info = "#".join([f"{k} = {v}" for k, v in pairs] + [f"DataPath = {data_path}"])
        request_file = send_request_like_vba(request_info)

        ok = wait_for_file(data_path, timeout_sec=max(0.1, float(req.timeout_sec)))
        if not ok:
            return {
                "ok": False,
                "status": "timeout",
                "request_file": request_file,
                "data_path": data_path,
            }

    # Read single-row CSV: "2016,2017,..."
    with open(data_path, "r", encoding="utf-8") as f:
        raw = f.read().strip()

    # robust parse: allow commas + newlines
    parts = [x.strip() for x in raw.replace("\n", ",").split(",") if x.strip()]

    # Return vector labels
    return {
        "ok": True,
        "labels": parts,
        "request_file": request_file,
        "data_path": data_path,
    }

# --- add routes near other API routes ---
@app.get("/adas/projects")
def adas_projects() -> Dict[str, Any]:
    if not os.path.exists(PROJECT_BOOK):
        raise HTTPException(404, f"Project workbook not found: {PROJECT_BOOK}")

    wb = openpyxl.load_workbook(PROJECT_BOOK, read_only=True, keep_vba=True, data_only=True)
    first_sheet = wb.sheetnames[0]
    ws = wb[first_sheet]

    vals = []
    for r in range(1, (ws.max_row or 1) + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        vals.append(s)

    # drop header if present
    if vals and vals[0].strip().lower() in ("project name", "projectname"):
        vals = vals[1:]

    # de-dup keep order
    seen = set()
    out = []
    for x in vals:
        if x not in seen:
            out.append(x)
            seen.add(x)

    return {"sheet": first_sheet, "projects": out}

@app.post("/adas/tri")
def adas_tri(req: AdaTriRequest) -> Dict[str, Any]:
    # NOTE:
    # exactly what ADASTri passes into SetDataPath.
    pairs = [
        ("Function", "ADASTri"),
        ("Path", req.Path),
        ("DatasetName", req.TriangleName),
        ("Cumulative", str(req.Cumulative)),
        ("Transposed", str(False)),
        ("Calendar", str(False)),
        ("ProjectName", req.ProjectName),
        ("OriginLength", str(req.OriginLength)),
        ("DevelopmentLength", str(req.DevelopmentLength)),
    ]

    data_path = set_data_path_like_vba(pairs)
    request_file = None  # <-- add

    if not os.path.exists(data_path):
        # Build requestInfo text (what your agent expects)
        request_info = "#".join([f"{k} = {v}" for k, v in pairs] + [f"DataPath = {data_path}"])
        request_file = send_request_like_vba(request_info)

        ok = wait_for_file(data_path, timeout_sec=max(0.1, float(req.timeout_sec)))
        if not ok:
            # Let UI decide whether to keep waiting or show timeout
            return {
                "ok": False,
                "status": "timeout",
                "request_file": request_file,
                "data_path": data_path,
            }

    # Register dataset id deterministically from the datapath
    ds_id = "adastri_" + hashlib.sha1(data_path.encode("utf-8")).hexdigest()[:16]
    DATASETS[ds_id] = data_path  # reuse existing /dataset/{ds_id}

    return {
        "ok": True,
        "ds_id": ds_id,
        "request_file": request_file,
        "data_path": data_path,
    }

@app.post("/book/meta")
def book_meta(req: AnyBookSheetRequest) -> Dict[str, Any]:
    book = resolve_allowed_book(req.book_path)
    if not book.exists():
        raise HTTPException(404, f"Workbook not found: {book}")
    st = os.stat(book)
    wb = openpyxl.load_workbook(str(book), read_only=True, keep_vba=True)
    return {"path": str(book), "mtime": st.st_mtime, "sheets": wb.sheetnames}

@app.post("/book/sheet")
def book_sheet(req: AnyBookSheetRequest) -> Dict[str, Any]:
    book = resolve_allowed_book(req.book_path)
    if not book.exists():
        raise HTTPException(404, f"Workbook not found: {book}")
    st = os.stat(book)
    values = read_sheet_matrix(str(book), sheet_name=req.sheet)
    return {"path": str(book), "sheet": req.sheet, "values": values, "mtime": st.st_mtime}

@app.post("/book/patch")
def book_patch(req: AnyBookPatchRequest) -> Dict[str, Any]:
    book = resolve_allowed_book(req.book_path)
    if not book.exists():
        raise HTTPException(404, f"Workbook not found: {book}")

    st = os.stat(book)
    if req.file_mtime is not None and abs(st.st_mtime - req.file_mtime) > 1e-6:
        raise HTTPException(409, "Workbook changed on disk. Reload and retry.")

    try:
        wb = openpyxl.load_workbook(str(book), data_only=True, keep_vba=True)
        if req.sheet not in wb.sheetnames:
            raise HTTPException(404, f"Sheet not found: {req.sheet}")
        ws = wb[req.sheet]

        for it in req.items:
            ws.cell(row=it.r + 1, column=it.c + 1).value = it.value

        tmp_path = str(book) + ".tmp"
        wb.save(tmp_path)
        os.replace(tmp_path, str(book))

        st2 = os.stat(book)
        return {"ok": True, "mtime": st2.st_mtime}

    except PermissionError:
        raise HTTPException(423, "Workbook is locked (possibly open in Excel). Close it and retry.")

@app.get("/datasets")
def list_datasets() -> List[Dict[str, Any]]:
    out = []
    for ds_id, path in DATASETS.items():
        if not os.path.exists(path):
            continue
        n_origin, n_dev = infer_shape(path)
        st = os.stat(path)
        out.append({
            "id": ds_id,
            "path": path,
            "shape": {"n_origin": n_origin, "n_dev": n_dev},
            "mtime": st.st_mtime,
        })
    return out

@app.get("/dataset/{ds_id}")
def get_dataset(ds_id: str, start_year: int = 2016) -> Dict[str, Any]:
    path = DATASETS.get(ds_id)
    if not path or not os.path.exists(path):
        raise HTTPException(404, f"Unknown dataset: {ds_id}")

    df = pd.read_csv(path, header=None, dtype="float64", keep_default_na=True)
    n_origin, n_dev = df.shape

    origin_labels = [str(start_year + i) for i in range(n_origin)]
    dev_labels = [str(12 * (j + 1)) for j in range(n_dev)]

    values = df.to_numpy()
    mask = ~np.isnan(values)

    st = os.stat(path)
    return {
        "id": ds_id,
        "origin_labels": origin_labels,
        "dev_labels": dev_labels,
        "values": np.where(np.isnan(values), None, values).tolist(),
        "mask": mask.tolist(),
        "mtime": st.st_mtime,
    }

@app.get("/dataset/{ds_id}/diagonal")
def get_diagonal(ds_id: str, k: int = 0, start_year: int = 2016) -> Dict[str, Any]:
    path = DATASETS.get(ds_id)
    if not path or not os.path.exists(path):
        raise HTTPException(404, f"Unknown dataset: {ds_id}")

    df = load_triangle_values(path)
    n_origin, n_dev = df.shape
    origin_labels, dev_labels = make_annual_labels(start_year, n_origin, n_dev)

    idx = diagonal_indices(n_origin, n_dev, k=k)
    items = []
    for r, c in idx:
        v = df.iat[r, c]
        items.append({
            "r": r,
            "c": c,
            "origin": origin_labels[r],
            "dev": dev_labels[c],
            "value": None if pd.isna(v) else float(v),
        })

    return {"id": ds_id, "k": k, "items": items}

@app.post("/dataset/{ds_id}/patch")
def patch_dataset(ds_id: str, req: PatchRequest) -> Dict[str, Any]:
    path = DATASETS.get(ds_id)
    if not path or not os.path.exists(path):
        raise HTTPException(404, f"Unknown dataset: {ds_id}")

    st = os.stat(path)
    if req.file_mtime is not None and abs(st.st_mtime - req.file_mtime) > 1e-6:
        raise HTTPException(409, "File changed on disk. Reload and retry.")

    df = load_triangle_values(path)
    n_origin, n_dev = df.shape
    mask = triangle_mask(n_origin, n_dev)

    applied = 0
    rejected: List[Dict[str, Any]] = []

    for it in req.items:
        r, c = it.r, it.c
        if r >= n_origin or c >= n_dev:
            rejected.append({"r": r, "c": c, "reason": "out_of_range"})
            continue
        if not mask[r, c]:
            rejected.append({"r": r, "c": c, "reason": "outside_triangle"})
            continue

        df.iat[r, c] = np.nan if it.value is None else float(it.value)
        applied += 1

    atomic_write_csv(df, path)
    st2 = os.stat(path)

    return {"ok": True, "applied": applied, "rejected": rejected, "mtime": st2.st_mtime}

# ---- Project Settings JSON API ----

class ProjectSettingsUpdateRequest(BaseModel):
    data: Dict[str, Any]
    file_mtime: Optional[float] = None


class FolderStructureUpdateRequest(BaseModel):
    folders: List[str] = Field(default_factory=list)


FOLDER_STRUCTURE_FILE = "folder_structure.json"


def _folder_structure_path() -> str:
    return os.path.join(PROJECT_SETTINGS_DIR, FOLDER_STRUCTURE_FILE)


@app.get("/project_settings/{source}/folders")
def get_project_folders(source: str) -> Dict[str, Any]:
    """Read folder structure (custom folder paths) from E:\\ADAS\\Team Profile\\folder_structure.json."""
    if source not in PROJECT_SETTINGS_SOURCES:
        raise HTTPException(404, f"Unknown source: {source}")

    filepath = _folder_structure_path()
    folders: List[str] = []
    if os.path.exists(filepath):
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
            folders = data.get(source, [])
            if not isinstance(folders, list):
                folders = []
        except Exception:
            folders = []

    return {"ok": True, "source": source, "folders": folders}


@app.post("/project_settings/{source}/folders")
def update_project_folders(source: str, req: FolderStructureUpdateRequest) -> Dict[str, Any]:
    """Save folder structure for the given source to folder_structure.json."""
    if source not in PROJECT_SETTINGS_SOURCES:
        raise HTTPException(404, f"Unknown source: {source}")

    filepath = _folder_structure_path()
    data: Dict[str, Any] = {}
    if os.path.exists(filepath):
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = {}
    if not isinstance(data, dict):
        data = {}

    data[source] = list(req.folders) if req.folders else []

    try:
        os.makedirs(PROJECT_SETTINGS_DIR, exist_ok=True)
        tmp_path = filepath + ".tmp"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        os.replace(tmp_path, filepath)
        return {"ok": True, "source": source}
    except PermissionError:
        raise HTTPException(423, "File is locked. Another user may have it open.")
    except Exception as e:
        raise HTTPException(500, f"Failed to save folder structure: {str(e)}")


@app.post("/project_settings/{source}/folders/init_from_data")
def init_folder_structure_from_data(source: str) -> Dict[str, Any]:
    """Build folder_structure.json from the project JSON (e.g. Actuarial_NJ.json) by extracting
    all folder paths from the Folder column and adding parent paths so the tree is complete."""
    if source not in PROJECT_SETTINGS_SOURCES:
        raise HTTPException(404, f"Unknown source: {source}")

    filename = PROJECT_SETTINGS_SOURCES[source]
    filepath = os.path.join(PROJECT_SETTINGS_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(404, f"Settings file not found: {filepath}")

    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Find first sheet (skip non-sheet keys like customFolders)
    sheet_name = next((k for k in data if k != "customFolders" and isinstance(data.get(k), dict)), None)
    if not sheet_name:
        raise HTTPException(400, "No sheet data found in project JSON")

    sheet = data[sheet_name]
    headers = sheet.get("headers") or []
    rows = sheet.get("rows") or []
    folder_idx = next((i for i, h in enumerate(headers) if h == "Folder"), -1)
    if folder_idx < 0:
        raise HTTPException(400, "Folder column not found in project JSON")

    # Collect all folder paths that appear in rows
    path_set: set = set()
    for row in rows:
        folder = (row[folder_idx] or "").strip()
        if not folder:
            continue
        path_set.add(folder)
        # Add all parent path prefixes so the tree has full hierarchy
        parts = folder.split("\\")
        for i in range(1, len(parts)):
            path_set.add("\\".join(parts[:i]))

    folders_list = sorted(path_set)
    filepath_folders = _folder_structure_path()
    existing: Dict[str, Any] = {}
    if os.path.exists(filepath_folders):
        try:
            with open(filepath_folders, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            existing = {}
    if not isinstance(existing, dict):
        existing = {}
    existing[source] = folders_list

    try:
        os.makedirs(PROJECT_SETTINGS_DIR, exist_ok=True)
        tmp_path = filepath_folders + ".tmp"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(existing, f, indent=2, ensure_ascii=False)
        os.replace(tmp_path, filepath_folders)
        return {"ok": True, "source": source, "folders_count": len(folders_list), "folders": folders_list}
    except PermissionError:
        raise HTTPException(423, "File is locked. Another user may have it open.")
    except Exception as e:
        raise HTTPException(500, f"Failed to write folder structure: {str(e)}")


@app.get("/project_settings/{source}")
def get_project_settings(source: str) -> Dict[str, Any]:
    """Read project settings JSON from shared network drive."""
    if source not in PROJECT_SETTINGS_SOURCES:
        raise HTTPException(404, f"Unknown source: {source}")

    filename = PROJECT_SETTINGS_SOURCES[source]
    filepath = os.path.join(PROJECT_SETTINGS_DIR, filename)

    if not os.path.exists(filepath):
        raise HTTPException(404, f"Settings file not found: {filepath}")

    st = os.stat(filepath)
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    return {
        "ok": True,
        "source": source,
        "path": filepath,
        "mtime": st.st_mtime,
        "data": data,
    }

@app.post("/project_settings/{source}")
def update_project_settings(source: str, req: ProjectSettingsUpdateRequest) -> Dict[str, Any]:
    """Update project settings JSON on shared network drive with conflict detection."""
    if source not in PROJECT_SETTINGS_SOURCES:
        raise HTTPException(404, f"Unknown source: {source}")

    filename = PROJECT_SETTINGS_SOURCES[source]
    filepath = os.path.join(PROJECT_SETTINGS_DIR, filename)

    # Check for concurrent modification
    if os.path.exists(filepath):
        st = os.stat(filepath)
        if req.file_mtime is not None and abs(st.st_mtime - req.file_mtime) > 0.001:
            raise HTTPException(409, "File was modified by another user. Please refresh and try again.")

    # Atomic write: write to temp file then rename
    try:
        tmp_path = filepath + ".tmp"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(req.data, f, indent=2, ensure_ascii=False)
        os.replace(tmp_path, filepath)

        st2 = os.stat(filepath)
        return {
            "ok": True,
            "source": source,
            "path": filepath,
            "mtime": st2.st_mtime,
        }
    except PermissionError:
        raise HTTPException(423, "File is locked. Another user may have it open.")
    except Exception as e:
        raise HTTPException(500, f"Failed to save: {str(e)}")

SUMMARY_CACHE_DIR = _get_data_base()  # Dynamic from ui_config.json

def get_cache_path(csv_path: str) -> str:
    """Get the cache file path for a given CSV file."""
    # Extract filename without extension
    filename = os.path.basename(csv_path)
    name_without_ext = os.path.splitext(filename)[0]
    
    # Try to extract project name from path (e.g., "ResQ_Channel_202512" from the path)
    # Use the CSV filename as the cache name
    cache_filename = f"{name_without_ext}_summary.json"
    return os.path.join(SUMMARY_CACHE_DIR, cache_filename)

def is_cache_valid(csv_path: str, cache_path: str) -> bool:
    """Check if cache exists and is newer than the CSV file."""
    if not os.path.exists(cache_path):
        return False
    csv_mtime = os.stat(csv_path).st_mtime
    cache_mtime = os.stat(cache_path).st_mtime
    return cache_mtime > csv_mtime

def generate_table_summary(path: str) -> Dict[str, Any]:
    """Generate summary info for a CSV file."""
    st = os.stat(path)
    file_size = st.st_size

    # Read full CSV
    df = pd.read_csv(path)
    row_count = len(df)

    # Get column info
    columns = []
    for col in df.columns:
        dtype = str(df[col].dtype)
        col_data = df[col].dropna()

        # Determine type and compute values summary
        if "int" in dtype:
            friendly_type = "Integer"
            if len(col_data) > 0:
                min_val = int(col_data.min())
                max_val = int(col_data.max())
                values_str = f"Range: ({min_val:,}, {max_val:,})"
            else:
                values_str = "(empty)"
        elif "float" in dtype:
            friendly_type = "Float"
            if len(col_data) > 0:
                min_val = col_data.min()
                max_val = col_data.max()
                # Format based on magnitude
                if abs(max_val) >= 1000 or abs(min_val) >= 1000:
                    values_str = f"Range: ({min_val:,.2f}, {max_val:,.2f})"
                else:
                    values_str = f"Range: ({min_val:.4f}, {max_val:.4f})"
            else:
                values_str = "(empty)"
        elif "object" in dtype:
                friendly_type = "String"
                # Get all distinct values
                distinct = col_data.unique().tolist()
                distinct_count = len(distinct)
                if distinct_count <= 10:
                    # Show all distinct values
                    values_str = ", ".join(str(v) for v in sorted(distinct, key=str))
                else:
                    # Too many - show count and sample
                    sample = sorted(distinct, key=str)[:10]
                    values_str = f"{distinct_count} distinct: {', '.join(str(v) for v in sample)}..."
        elif "datetime" in dtype:
            friendly_type = "DateTime"
            if len(col_data) > 0:
                min_val = col_data.min()
                max_val = col_data.max()
                values_str = f"Range: {min_val} - {max_val}"
            else:
                values_str = "(empty)"
        elif "bool" in dtype:
            friendly_type = "Boolean"
            values_str = "True, False"
        else:
            friendly_type = dtype
            values_str = "(unknown)"

        columns.append({
            "name": str(col),
            "dtype": dtype,
            "type": friendly_type,
            "values": values_str,
        })

    # Format file size
    if file_size < 1024:
        size_str = f"{file_size} B"
    elif file_size < 1024 * 1024:
        size_str = f"{file_size / 1024:.1f} KB"
    else:
        size_str = f"{file_size / (1024 * 1024):.2f} MB"

    return {
        "ok": True,
        "path": path,
        "row_count": row_count,
        "column_count": len(columns),
        "file_size": file_size,
        "file_size_str": size_str,
        "columns": columns,
        "csv_mtime": st.st_mtime,
    }

@app.get("/table_summary")
def get_table_summary(path: str) -> Dict[str, Any]:
    """Get summary info about a CSV/data table file. Uses cached JSON if available."""
    if not path:
        raise HTTPException(400, "Missing path parameter")

    if not os.path.exists(path):
        raise HTTPException(404, f"File not found: {path}")

    try:
        cache_path = get_cache_path(path)
        
        # Check if valid cache exists
        if is_cache_valid(path, cache_path):
            # Load from cache
            with open(cache_path, "r", encoding="utf-8") as f:
                cached_data = json.load(f)
            cached_data["from_cache"] = True
            return cached_data
        
        # Generate new summary
        summary = generate_table_summary(path)
        summary["from_cache"] = False
        
        # Save to cache
        os.makedirs(SUMMARY_CACHE_DIR, exist_ok=True)
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(summary, f, indent=2)
        
        return summary
    except Exception as e:
        raise HTTPException(500, f"Error reading file: {str(e)}")

@app.get("/project_settings")
def list_project_settings_sources() -> Dict[str, Any]:
    """List available project settings sources."""
    sources = []
    for key, filename in PROJECT_SETTINGS_SOURCES.items():
        filepath = os.path.join(PROJECT_SETTINGS_DIR, filename)
        exists = os.path.exists(filepath)
        sources.append({
            "key": key,
            "filename": filename,
            "path": filepath,
            "exists": exists,
        })
    return {"sources": sources}

@app.get("/project_book/meta")
def project_book_meta() -> Dict[str, Any]:
    if not os.path.exists(PROJECT_BOOK):
        raise HTTPException(404, f"Project workbook not found: {PROJECT_BOOK}")

    st = os.stat(PROJECT_BOOK)
    wb = openpyxl.load_workbook(PROJECT_BOOK, read_only=True, keep_vba=True)
    return {
        "path": PROJECT_BOOK,
        "mtime": st.st_mtime,
        "sheets": wb.sheetnames,
    }

@app.get("/project_book/sheet")
def project_book_sheet(sheet: str) -> Dict[str, Any]:
    if not os.path.exists(PROJECT_BOOK):
        raise HTTPException(404, f"Project workbook not found: {PROJECT_BOOK}")

    st = os.stat(PROJECT_BOOK)
    values = read_sheet_matrix(PROJECT_BOOK, sheet_name=sheet)
    return {"sheet": sheet, "values": values, "mtime": st.st_mtime}

@app.post("/project_book/patch")
def project_book_patch(req: XlsmPatchRequest) -> Dict[str, Any]:
    if not os.path.exists(PROJECT_BOOK):
        raise HTTPException(404, f"Project workbook not found: {PROJECT_BOOK}")

    st = os.stat(PROJECT_BOOK)
    if req.file_mtime is not None and abs(st.st_mtime - req.file_mtime) > 1e-6:
        raise HTTPException(409, "Workbook changed on disk. Reload and retry.")

    # If Excel has it open, save may fail; treat as locked.
    try:
        wb = openpyxl.load_workbook(PROJECT_BOOK, data_only=True, keep_vba=True)
        if req.sheet not in wb.sheetnames:
            raise HTTPException(404, f"Sheet not found: {req.sheet}")
        ws = wb[req.sheet]

        applied = 0
        rejected = []

        for it in req.items:
            rr = it.r + 1  # openpyxl is 1-based
            cc = it.c + 1
            if rr < 1 or cc < 1:
                rejected.append({"r": it.r, "c": it.c, "reason": "out_of_range"})
                continue

            v = it.value
            # Keep formulas if user sends "=..."
            if isinstance(v, str) and v.startswith("="):
                ws.cell(row=rr, column=cc).value = v
            else:
                ws.cell(row=rr, column=cc).value = v
            applied += 1

        # Atomic-ish save: write to temp then replace
        tmp_path = PROJECT_BOOK + ".tmp"
        wb.save(tmp_path)
        os.replace(tmp_path, PROJECT_BOOK)

        st2 = os.stat(PROJECT_BOOK)
        return {"ok": True, "applied": applied, "rejected": rejected, "mtime": st2.st_mtime}

    except PermissionError:
        raise HTTPException(423, "Workbook is locked (possibly open in Excel). Close it and retry.")

@app.post("/workflow/save")
def workflow_save(req: WorkflowSaveRequest) -> Dict[str, Any]:
    name = _sanitize_filename(req.name or "workflow")
    os.makedirs(WORKFLOW_DIR, exist_ok=True)
    path = os.path.join(WORKFLOW_DIR, f"{name}{WORKFLOW_EXT}")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(req.data, f, ensure_ascii=True, indent=2)
    prev = req.prev_path
    if prev:
      try:
        prev_path = Path(prev).resolve()
        new_path = Path(path).resolve()
        root = Path(WORKFLOW_DIR).resolve()
        if prev_path != new_path and str(prev_path).startswith(str(root) + os.sep):
          if prev_path.exists():
            prev_path.unlink()
      except Exception:
        pass
    return {"ok": True, "path": path}

@app.post("/workflow/save_as")
def workflow_save_as(req: WorkflowSaveAsRequest) -> Dict[str, Any]:
    if not req.path:
        raise HTTPException(400, "Missing path")
    p = Path(req.path)
    if not p.is_absolute():
        p = Path(WORKFLOW_DIR) / p
    if p.suffix.lower() not in (".json", WORKFLOW_EXT):
        p = p.with_suffix(WORKFLOW_EXT)
    os.makedirs(p.parent, exist_ok=True)
    with open(str(p), "w", encoding="utf-8") as f:
        json.dump(req.data, f, ensure_ascii=True, indent=2)
    return {"ok": True, "path": str(p)}

@app.get("/workflow/default_dir")
def workflow_default_dir() -> Dict[str, Any]:
    os.makedirs(WORKFLOW_DIR, exist_ok=True)
    return {"path": WORKFLOW_DIR}

@app.post("/app/restart")
def app_restart() -> Dict[str, Any]:
    try:
        RESTART_FLAG.write_text(str(time.time()), encoding="utf-8")
    except Exception:
        pass

    def _shutdown() -> None:
        time.sleep(0.25)
        os._exit(0)

    threading.Thread(target=_shutdown, daemon=True).start()
    return {"ok": True}

@app.post("/app/restart_electron")
def app_restart_electron() -> Dict[str, Any]:
    try:
        ELECTRON_RESTART_FLAG.write_text(str(time.time()), encoding="utf-8")
    except Exception:
        pass
    return {"ok": True}

@app.post("/app/shutdown_electron")
def app_shutdown_electron() -> Dict[str, Any]:
    try:
        ELECTRON_SHUTDOWN_FLAG.write_text(str(time.time()), encoding="utf-8")
    except Exception:
        pass
    return {"ok": True}

@app.post("/app/shutdown")
def app_shutdown() -> Dict[str, Any]:
    try:
        SHUTDOWN_FLAG.write_text(str(time.time()), encoding="utf-8")
    except Exception:
        pass

    def _shutdown() -> None:
        time.sleep(0.25)
        os._exit(0)

    threading.Thread(target=_shutdown, daemon=True).start()
    return {"ok": True}

@app.post("/workflow/load")
def workflow_load(req: WorkflowLoadRequest) -> Dict[str, Any]:
    if not req.path:
        raise HTTPException(400, "Missing path")
    p = Path(req.path)
    if not p.is_absolute():
        p = Path(WORKFLOW_DIR) / p
    if p.suffix == "":
        p = p.with_suffix(WORKFLOW_EXT)
    if not p.exists():
        raise HTTPException(404, f"Workflow not found: {p}")
    with open(str(p), "r", encoding="utf-8") as f:
        data = json.load(f)
    return {"ok": True, "path": str(p), "data": data}

# ---- Frontend (same folder, no /static) ----
# Mount AFTER API routes to avoid conflicts

app.mount("/ui", StaticFiles(directory=str(BASE_DIR), html=True), name="ui")

@app.get("/")
def home():
    return RedirectResponse(url="/ui/")
